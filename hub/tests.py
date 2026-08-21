import io
import json
from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import patch
from urllib.parse import parse_qs, urlparse

from django.contrib.auth import authenticate
from django.contrib.messages import get_messages
from django.contrib.messages.storage.fallback import FallbackStorage
from django.contrib.sessions.middleware import SessionMiddleware
from django.http import HttpResponse
from django.core.management import call_command
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from hub.forms import (
    AdminRequestFilterForm,
    EngineerActivityLogForm,
    RequestAdminForm,
    RequestForm,
    RequestStatusForm,
    SqrSubmissionForm,
)
from hub.models import (
    Account,
    EngineerActivityLog,
    Request,
    RequestCommunication,
    SqrSubmission,
    SqrSubmissionChange,
    SqrSubmissionHistory,
)
from hub.views import (
    AssignmentEmailResult,
    DashboardLiveDataView,
    DashboardView,
    RequestAdminUpdateView,
    RequestCollaborativeManageView,
    SqrListView,
    UserManagementView,
    clear_engineer_outlook_lock_on_reassignment,
    notify_engineer_assignment_email,
)


class AssignmentEmailTests(TestCase):
    def setUp(self):
        self.requestor = User.objects.create_user(
            username="requestor1",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="requestor@example.com",
        )
        self.engineer = User.objects.create_user(
            username="engineer1",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer@example.com",
        )

    def test_notify_engineer_assignment_email_returns_no_new_assignee_without_assignment(self):
        request_obj = Request.objects.create(
            requestor=self.requestor,
            account_manager="Requestor One",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            account_id=self._create_account_id("Account A"),
        )

        result = notify_engineer_assignment_email(request_obj, actor_user=self.requestor)

        self.assertEqual(result.status, "no_new_assignee")

    def test_notify_engineer_assignment_email_returns_missing_email_when_assignee_has_no_email(self):
        self.engineer.email = ""
        self.engineer.save(update_fields=["email"])
        request_obj = Request.objects.create(
            requestor=self.requestor,
            account_manager="Requestor One",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            account_id=self._create_account_id("Account B"),
        )

        result = notify_engineer_assignment_email(request_obj, actor_user=self.requestor)

        self.assertEqual(result.status, "missing_assignee_email")

    def test_dashboard_post_surfaces_assignment_delivery_failure(self):
        factory = RequestFactory()
        request = factory.post(
            reverse("hub:dashboard"),
            data={
                "account_name": "Account C",
                "needed_by": "2026-04-19",
                "product_category": "Azure",
                "engagement_type": Request.Engagement.SUPPORT,
                "priority": Request.Priority.MEDIUM,
                "description": "Need deployment help.",
                "engineer": str(self.engineer.pk),
            },
        )
        request.user = self.requestor
        self._attach_session_and_messages(request)

        with patch("hub.views.notify_engineer_assignment_email", return_value=AssignmentEmailResult("delivery_failed")):
            response = DashboardView.as_view()(request)

        self.assertEqual(response.status_code, 302)
        messages = [message.message for message in get_messages(request)]
        self.assertIn("Request submitted", messages)
        self.assertTrue(any("assignment email could not be delivered" in message for message in messages))

    @staticmethod
    def _attach_session_and_messages(request):
        session_middleware = SessionMiddleware(lambda req: None)
        session_middleware.process_request(request)
        request.session.save()
        setattr(request, "_messages", FallbackStorage(request))

    def _create_account_id(self, name: str) -> int:
        return Account.objects.create(name=name).pk


class AuthenticationBackendTests(TestCase):
    def test_exact_case_admin_username_takes_priority_over_case_insensitive_duplicate(self):
        canonical_admin = User.objects.get(username="Admin")
        canonical_admin.set_password("@Password")
        canonical_admin.role = User.Roles.ADMIN
        canonical_admin.is_staff = True
        canonical_admin.is_superuser = True
        canonical_admin.save(update_fields=["password", "role", "is_staff", "is_superuser"])

        User.objects.create_user(
            username="admin",
            password="lowerpass",
            role=User.Roles.REQUESTOR,
            email="lowercase.admin@example.com",
            is_staff=True,
            is_superuser=True,
        )

        authenticated = authenticate(username="Admin", password="@Password")

        self.assertIsNotNone(authenticated)
        self.assertEqual(authenticated.pk, canonical_admin.pk)
        self.assertEqual(authenticated.username, "Admin")


class RequestReportViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="reports-admin",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="reports-admin@example.com",
        )
        self.requestor = User.objects.create_user(
            username="reports-requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="reports-requestor@example.com",
        )
        self.engineer = User.objects.create_user(
            username="reports-engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="reports-engineer@example.com",
            first_name="Report",
            last_name="Engineer",
        )
        self.account = Account.objects.create(name="Reports Account")

    def _create_request(self, **overrides):
        values = {
            "requestor": self.requestor,
            "account": self.account,
            "account_manager": "Account Manager 01",
            "product_category": "Azure",
            "engagement_type": Request.Engagement.SUPPORT,
            "priority": Request.Priority.MEDIUM,
            "engineer": self.engineer,
        }
        values.update(overrides)
        return Request.objects.create(**values)

    def _create_activity_log(self, **overrides):
        values = {
            "engineer": self.engineer,
            "account": self.account,
            "request_date": date(2026, 8, 15),
            "activity_type": EngineerActivityLog.ActivityType.INTERNAL_SUPPORT,
            "actual_hours": Decimal("2.50"),
            "details": "Activity report regression entry",
            "location": EngineerActivityLog.Location.OFFICE,
            "is_billable": True,
            "status": EngineerActivityLog.Status.COMPLETED,
        }
        values.update(overrides)
        return EngineerActivityLog.objects.create(**values)

    def test_report_requires_admin_or_pm_esg_role(self):
        self.client.force_login(self.requestor)
        response = self.client.get(reverse("hub:report"))
        self.assertEqual(response.status_code, 403)

        pm_esg = User.objects.create_user(
            username="reports-pm-esg",
            password="pass12345",
            role=User.Roles.PM_ESG,
        )
        self.client.force_login(pm_esg)
        response = self.client.get(reverse("hub:report"))
        self.assertEqual(response.status_code, 200)

    def test_operational_report_excludes_soft_deleted_requests(self):
        self._create_request(status=Request.Status.ONGOING)
        deleted = self._create_request(
            account_manager="Deleted Manager",
            status=Request.Status.COMPLETED,
        )
        deleted.is_deleted = True
        deleted.save(update_fields=["is_deleted"])

        self.client.force_login(self.admin)
        response = self.client.get(reverse("hub:report"))

        self.assertEqual(response.context["totals"]["requests"], 1)
        self.assertEqual(response.context["totals"]["ongoing"], 1)
        self.assertEqual(response.context["totals"]["completed"], 0)
        self.assertNotIn("Deleted Manager", response.context["account_manager_chart"]["labels"])

    def test_operational_charts_include_all_configured_categories(self):
        self._create_request(
            engagement_type=Request.Engagement.PROJECT_MANAGEMENT,
            product_category="VMware",
        )

        self.client.force_login(self.admin)
        response = self.client.get(reverse("hub:report"))

        self.assertEqual(
            response.context["engagement_chart"]["labels"],
            [label for _, label in Request.Engagement.choices],
        )
        self.assertEqual(
            response.context["product_chart"]["labels"],
            [label for _, label in Request._meta.get_field("product_category").choices],
        )
        engagement_index = response.context["engagement_chart"]["labels"].index("Project Management")
        product_index = response.context["product_chart"]["labels"].index("VMware")
        self.assertEqual(response.context["engagement_chart"]["totals"][engagement_index], 1)
        self.assertEqual(response.context["product_chart"]["totals"][product_index], 1)

    def test_ranked_charts_compact_after_top_ten_but_tables_remain_complete(self):
        for index in range(12):
            for _ in range(12 - index):
                self._create_request(
                    account_manager=f"Account Manager {index + 1:02d}",
                    engineer=None,
                    status=Request.Status.COMPLETED,
                )

        self.client.force_login(self.admin)
        response = self.client.get(reverse("hub:report"))

        chart = response.context["account_manager_chart"]
        self.assertEqual(len(chart["labels"]), 11)
        self.assertEqual(chart["labels"][-1], "Others")
        self.assertEqual(chart["totals"][-1], 3)
        self.assertEqual(len(response.context["account_manager_stats"]), 12)
        self.assertEqual(len(response.context["account_manager_chart_full"]["labels"]), 12)
        self.assertNotIn("Others", response.context["account_manager_chart_full"]["labels"])

    def test_activity_report_uses_operational_layout_and_retains_recent_activity(self):
        activity_log = self._create_activity_log()
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("hub:report"),
            {
                "report_view": "activity",
                "start_month": "2026-09",
                "end_month": "2026-08",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["report_view"], "activity")
        self.assertContains(response, 'class="rpt-tab rpt-tab--active"')
        self.assertContains(response, 'aria-current="page"')
        self.assertContains(response, 'class="glass-card rpt2-filter rpt2-activity-filter')
        self.assertContains(response, 'class="row row-cols-2 row-cols-md-3 row-cols-xl-5 g-3 mb-4 lg-enter rpt2-kpi-grid"')
        self.assertContains(response, 'class="glass-card rpt-chart-card rpt2-activity-log-card')
        self.assertContains(response, "Recent Engineer Activity")
        self.assertContains(response, "Activity report regression entry")
        self.assertEqual(response.context["activity_start_month"], "2026-08")
        self.assertEqual(response.context["activity_end_month"], "2026-09")
        self.assertTrue(response.context["activity_range_swapped"])
        self.assertContains(response, "Range was reversed")
        self.assertContains(response, "August 2026 to September 2026")
        self.assertContains(response, "start_month=2026-08&end_month=2026-09")
        self.assertContains(response, f"edit_activity={activity_log.pk}")

    def test_activity_report_pagination_preserves_month_filters(self):
        for index in range(51):
            self._create_activity_log(details=f"Paginated activity {index:02d}")
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("hub:report"),
            {
                "report_view": "activity",
                "start_month": "2026-08",
                "end_month": "2026-08",
            },
        )

        self.assertEqual(response.context["activity_logs_page_obj"].paginator.num_pages, 2)
        self.assertContains(
            response,
            "?report_view=activity&start_month=2026-08&end_month=2026-08&activity_log_page=2",
        )


class UserManagementSearchTests(TestCase):
    def test_user_search_prioritizes_best_matches_and_sorts_results(self):
        alina = User.objects.create_user(
            username="alina",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="alina@example.com",
            first_name="Alina",
            last_name="Smith",
        )
        alice = User.objects.create_user(
            username="alice",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="alice@example.com",
            first_name="Alice",
            last_name="Wong",
        )
        User.objects.create_user(
            username="zoe",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="zoe@example.com",
            first_name="Zoe",
            last_name="Brown",
        )

        alina.date_joined = timezone.datetime(2024, 1, 1, tzinfo=timezone.utc)
        alice.date_joined = timezone.datetime(2024, 1, 3, tzinfo=timezone.utc)
        alina.save(update_fields=["date_joined"])
        alice.save(update_fields=["date_joined"])

        request = RequestFactory().get("/management/", {"user_q": "ali"})

        page_obj, query = UserManagementView()._paginate_users(request)
        usernames = [user.username for user in page_obj.object_list]

        self.assertEqual(query, "ali")
        self.assertEqual(usernames[:2], ["alice", "alina"])


class RequestStatusValidationTests(TestCase):
    def setUp(self):
        self.requestor = User.objects.create_user(
            username="status_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="status.requestor@example.com",
        )
        self.engineer = User.objects.create_user(
            username="status_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="status.engineer@example.com",
        )
        self.backup_engineer = User.objects.create_user(
            username="status_backup_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="status.backup@example.com",
        )
        self.account = Account.objects.create(name="Status Validation Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Status Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            backup_engineer=self.backup_engineer,
            status=Request.Status.ONGOING,
            description="Status validation request.",
        )

    def test_assigned_engineer_cannot_complete_without_related_activity_logs(self):
        form = RequestStatusForm(
            data={"status": Request.Status.COMPLETED, "end_date": timezone.now().date()},
            instance=self.request_obj,
            actor_user=self.engineer,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("status", form.errors)
        self.assertIn("related activity", form.errors["status"][0].lower())

    def test_backup_engineer_cannot_complete_without_related_activity_logs(self):
        form = RequestStatusForm(
            data={"status": Request.Status.COMPLETED, "end_date": timezone.now().date()},
            instance=self.request_obj,
            actor_user=self.backup_engineer,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("status", form.errors)
        self.assertIn("related activity", form.errors["status"][0].lower())


class OutlookThreadingTests(TestCase):
    def setUp(self):
        self.requestor = User.objects.create_user(
            username="thread_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR_ESS,
            email="thread.requestor@example.com",
            first_name="Thread",
            last_name="Requestor",
        )
        self.engineer = User.objects.create_user(
            username="thread_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="thread.engineer@example.com",
            first_name="Thread",
            last_name="Engineer",
        )
        self.backup_engineer = User.objects.create_user(
            username="thread_backup",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="thread.backup@example.com",
        )
        self.account = Account.objects.create(name="Thread Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Thread Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            backup_engineer=self.backup_engineer,
            status=Request.Status.COMPLETED,
            description="Threading support request.",
        )

    def test_closing_draft_reuses_acknowledgement_subject(self):
        contexts = self._post_acknowledgement_and_closing_with_captured_contexts()
        acknowledgement_query = self._mailto_query(contexts[0]["mailto_url"])
        closing_query = self._mailto_query(contexts[1]["mailto_url"])

        self.assertEqual(acknowledgement_query["subject"], [f"Re: {self.request_obj.reference_code} · Thread Account"])
        self.assertEqual(closing_query["subject"], acknowledgement_query["subject"])
        self.assertNotIn("Advisory Only", closing_query["subject"][0])
        self.assertIn("Advisory Only", closing_query["body"][0])

    def test_closing_draft_keeps_requestor_ess_cc_consistent_with_acknowledgement(self):
        contexts = self._post_acknowledgement_and_closing_with_captured_contexts()
        acknowledgement_cc = set(self._mailto_query(contexts[0]["mailto_url"])["cc"][0].split(","))
        closing_cc = set(self._mailto_query(contexts[1]["mailto_url"])["cc"][0].split(","))

        self.assertIn("JoanI@phildata.com", acknowledgement_cc)
        self.assertIn("JoanI@phildata.com", closing_cc)
        self.assertEqual(closing_cc, acknowledgement_cc)

    def _post_acknowledgement_and_closing_with_captured_contexts(self) -> list[dict]:
        self.client.force_login(self.engineer)
        contexts = []

        def capture_render(request, template_name, context=None, *args, **kwargs):
            contexts.append(context or {})
            return HttpResponse("OK")

        with patch("hub.views.render", side_effect=capture_render):
            acknowledgement_response = self.client.post(
                reverse("hub:request-outlook", args=[self.request_obj.pk]),
                HTTP_REFERER=reverse("hub:dashboard"),
            )
            closing_response = self.client.post(
                reverse("hub:request-outlook-closing", args=[self.request_obj.pk]),
                HTTP_REFERER=reverse("hub:request-manage-collab", args=[self.request_obj.pk]),
            )

        self.assertEqual(acknowledgement_response.status_code, 200)
        self.assertEqual(closing_response.status_code, 200)
        self.assertEqual(len(contexts), 2)
        return contexts

    @staticmethod
    def _mailto_query(mailto_url: str) -> dict[str, list[str]]:
        return parse_qs(urlparse(mailto_url).query)


class RequestTeamsUrlTests(TestCase):
    def test_teams_url_includes_only_requestor_assigned_and_backup(self):
        requestor = User.objects.create_user(
            username="teams_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="teams.requestor@example.com",
        )
        engineer = User.objects.create_user(
            username="teams_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="teams.engineer@example.com",
        )
        backup_engineer = User.objects.create_user(
            username="teams_backup",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="teams.backup@example.com",
        )
        request_obj = Request.objects.create(
            requestor=requestor,
            account=Account.objects.create(name="Teams Account"),
            account_manager="Teams Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=engineer,
            backup_engineer=backup_engineer,
        )

        participants = set(parse_qs(urlparse(request_obj.teams_chat_url).query)["users"][0].split(","))

        self.assertEqual(
            participants,
            {"teams.requestor@example.com", "teams.engineer@example.com", "teams.backup@example.com"},
        )
        self.assertNotIn("JeanM@phildata.com", participants)


class DashboardViewTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.pm_ess = User.objects.create_user(
            username="pmess1",
            password="pass12345",
            role=User.Roles.PM_ESS,
            email="pmess@example.com",
        )
        self.requestor_ess = User.objects.create_user(
            username="requestoress1",
            password="pass12345",
            role=User.Roles.REQUESTOR_ESS,
            email="requestoress@example.com",
        )
        self.requestor = User.objects.create_user(
            username="requestor2",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="requestor2@example.com",
        )
        self.pm_esg = User.objects.create_user(
            username="pmesg1",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="pmesg@example.com",
            first_name="PM",
            last_name="ESG",
        )
        self.engineer = User.objects.create_user(
            username="engineer_dashboard",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer.dashboard@example.com",
        )
        self.account = Account.objects.create(name="Dashboard Account")

    def test_dashboard_includes_live_keyword_search_controls(self):
        admin = User.objects.create_user(
            username="admin_dashboard_search",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="admin.dashboard.search@example.com",
        )
        Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Regular Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            description="Searchable dashboard request",
        )

        self.client.force_login(admin)
        response = self.client.get(reverse("hub:dashboard"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('id="dashboard-request-search"', content)
        self.assertIn('id="dashboard-request-count-badge"', content)
        self.assertIn('Search requests...', content)

    def test_dashboard_search_preserves_server_sort_when_query_is_empty(self):
        admin = User.objects.create_user(
            username="admin_dashboard_sort",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="admin.dashboard.sort@example.com",
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse("hub:dashboard"),
            {"sort": "engineer", "direction": "asc"},
        )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("const filtered = !q", content)
        self.assertIn("? state.records.map(function (record) {", content)

    def test_sqr_page_includes_compact_toolbar_search_control(self):
        self.client.force_login(self.pm_esg)
        response = self.client.get(reverse("hub:sqr"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('id="sqr-header-search"', content)
        self.assertIn('placeholder="Search SQR..."', content)

    def test_pm_ess_dashboard_includes_own_and_requestor_ess_requests(self):
        own_request = Request.objects.create(
            requestor=self.pm_ess,
            account=self.account,
            account_manager="PM ESS",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
        )
        requestor_ess_request = Request.objects.create(
            requestor=self.requestor_ess,
            account=self.account,
            account_manager="Requestor ESS",
            product_category="M365",
            engagement_type=Request.Engagement.TRAINING,
            priority=Request.Priority.MEDIUM,
        )
        other_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Regular Requestor",
            product_category="Dell",
            engagement_type=Request.Engagement.INQUIRY,
            priority=Request.Priority.MEDIUM,
        )

        request = self.factory.get(reverse("hub:dashboard"))
        request.user = self.pm_ess

        view = DashboardView()
        view.setup(request)
        context = view.get_context_data()

        request_ids = {item.pk for item in context["requests"]}

        self.assertIn(own_request.pk, request_ids)
        self.assertIn(requestor_ess_request.pk, request_ids)
        self.assertNotIn(other_request.pk, request_ids)

    def test_pm_ess_dashboard_my_requests_tab_shows_only_own_requests(self):
        own_request = Request.objects.create(
            requestor=self.pm_ess,
            account=self.account,
            account_manager="PM ESS",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
        )
        requestor_ess_request = Request.objects.create(
            requestor=self.requestor_ess,
            account=self.account,
            account_manager="Requestor ESS",
            product_category="M365",
            engagement_type=Request.Engagement.TRAINING,
            priority=Request.Priority.MEDIUM,
        )

        request = self.factory.get(reverse("hub:dashboard"), data={"request_tab": "mine"})
        request.user = self.pm_ess

        view = DashboardView()
        view.setup(request)
        context = view.get_context_data()
        request_ids = {item.pk for item in context["requests"]}

        self.assertEqual(context["pm_ess_request_tab"], "mine")
        self.assertIn(own_request.pk, request_ids)
        self.assertNotIn(requestor_ess_request.pk, request_ids)

    def test_pm_esg_can_submit_request_from_dashboard(self):
        request = self.factory.post(
            reverse("hub:dashboard"),
            data={
                "account_name": self.account.name,
                "needed_by": "2026-04-19",
                "product_category": "Azure",
                "engagement_type": Request.Engagement.SUPPORT,
                "priority": Request.Priority.MEDIUM,
                "description": "PM ESG submitted request.",
                "engineer": str(self.engineer.pk),
            },
        )
        request.user = self.pm_esg
        AssignmentEmailTests._attach_session_and_messages(request)

        with patch("hub.views.notify_engineer_assignment_email", return_value=AssignmentEmailResult("sent")):
            response = DashboardView.as_view()(request)

        self.assertEqual(response.status_code, 302)
        request_obj = Request.objects.get(description="PM ESG submitted request.")
        self.assertEqual(request_obj.requestor, self.pm_esg)
        self.assertEqual(request_obj.account_manager, "PM ESG")
        self.assertEqual(request_obj.assignment_revision, 1)
        self.assertEqual(request_obj.lifecycle_stage, Request.LifecycleStage.ASSIGNED)
        self.assertEqual(
            list(request_obj.lifecycle_events.values_list("event_type", flat=True)),
            ["created", "assigned"],
        )

    def test_requestor_dashboard_post_initializes_unassigned_lifecycle(self):
        request = self.factory.post(
            reverse("hub:dashboard"),
            data={
                "account_name": self.account.name,
                "product_category": "Azure",
                "engagement_type": Request.Engagement.SUPPORT,
                "priority": Request.Priority.MEDIUM,
                "description": "Unassigned request lifecycle defaults.",
            },
        )
        request.user = self.requestor
        AssignmentEmailTests._attach_session_and_messages(request)

        response = DashboardView.as_view()(request)

        self.assertEqual(response.status_code, 302)
        request_obj = Request.objects.get(description="Unassigned request lifecycle defaults.")
        self.assertEqual(request_obj.assignment_revision, 0)
        self.assertEqual(request_obj.lifecycle_stage, Request.LifecycleStage.CREATED)
        self.assertEqual(
            list(request_obj.lifecycle_events.values_list("event_type", flat=True)),
            ["created"],
        )

    def test_engineer_completed_filter_orders_recent_first(self):
        engineer = User.objects.create_user(
            username="engineer_completed_order",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer.completed@example.com",
        )

        older_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Regular Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=engineer,
            status=Request.Status.COMPLETED,
            end_date=timezone.now().date(),
            description="older completed request",
        )
        newer_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Regular Requestor",
            product_category="M365",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=engineer,
            status=Request.Status.COMPLETED,
            end_date=timezone.now().date(),
            description="newer completed request",
        )

        older_timestamp = timezone.now() - timedelta(days=2)
        newer_timestamp = timezone.now() - timedelta(days=1)
        Request.objects.filter(pk=older_request.pk).update(created_at=older_timestamp, updated_at=older_timestamp)
        Request.objects.filter(pk=newer_request.pk).update(created_at=newer_timestamp, updated_at=newer_timestamp)

        request = self.factory.get(reverse("hub:dashboard"), data={"metric_filter": "completed"})
        request.user = engineer

        view = DashboardView()
        view.setup(request)
        context = view.get_context_data()
        request_ids = [item.pk for item in context["requests"]]

        self.assertEqual(context["active_metric_filter"], "completed")
        self.assertEqual(request_ids, [newer_request.pk, older_request.pk])


class DashboardLiveDataViewTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.admin = User.objects.create_user(
            username="admin_live",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="admin.live@example.com",
        )
        self.engineer = User.objects.create_user(
            username="engineer_live",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer.live@example.com",
        )
        self.requestor = User.objects.create_user(
            username="requestor_live",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="requestor.live@example.com",
        )
        self.account = Account.objects.create(name="Live Dashboard Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Live Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            description="live dashboard request",
        )

    def test_non_admin_denied(self):
        request = self.factory.get(reverse("hub:dashboard-live"))
        request.user = self.engineer
        response = DashboardLiveDataView.as_view()(request)
        self.assertEqual(response.status_code, 403)
        payload = json.loads(response.content)
        self.assertFalse(payload["ok"])

    def test_admin_receives_rows_and_metrics(self):
        request = self.factory.get(reverse("hub:dashboard-live"))
        request.user = self.admin
        response = DashboardLiveDataView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.content)
        self.assertTrue(payload["ok"])
        self.assertTrue(payload["changed"])
        self.assertIn(self.request_obj.reference_code, payload["html_rows"])
        self.assertIn("dashboard-live-metrics", payload["html_metrics"])
        self.assertIn("dashboard-live-status-pills", payload["html_status_pills"])
        self.assertTrue(payload["version"])

    def test_unchanged_version_short_circuits(self):
        version = DashboardView._admin_dashboard_live_version(self.admin)
        request = self.factory.get(reverse("hub:dashboard-live"), data={"version": version})
        request.user = self.admin
        response = DashboardLiveDataView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.content)
        self.assertTrue(payload["ok"])
        self.assertFalse(payload["changed"])
        self.assertEqual(payload["version"], version)
        self.assertNotIn("html_rows", payload)

    def test_soft_delete_removes_row_from_live_payload(self):
        Request.objects.filter(pk=self.request_obj.pk).update(
            is_deleted=True,
            deleted_at=timezone.now(),
            updated_at=timezone.now(),
        )
        request = self.factory.get(reverse("hub:dashboard-live"))
        request.user = self.admin
        response = DashboardLiveDataView.as_view()(request)
        payload = json.loads(response.content)
        self.assertTrue(payload["ok"])
        self.assertNotIn(self.request_obj.reference_code, payload["html_rows"])

    def test_completed_metric_filter_applied(self):
        self.request_obj.status = Request.Status.COMPLETED
        self.request_obj.end_date = timezone.now().date()
        self.request_obj.save(update_fields=["status", "end_date", "updated_at"])
        request = self.factory.get(
            reverse("hub:dashboard-live"),
            data={"metric_filter": "completed"},
        )
        request.user = self.admin
        response = DashboardLiveDataView.as_view()(request)
        payload = json.loads(response.content)
        self.assertTrue(payload["ok"])
        self.assertIn(self.request_obj.reference_code, payload["html_rows"])
        self.assertGreaterEqual(payload["request_count"], 1)


class OnHoldRoleTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.requestor = User.objects.create_user(
            username="requestor_onhold",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="requestor.onhold@example.com",
        )
        self.engineer = User.objects.create_user(
            username="engineer_active",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer.active@example.com",
        )
        self.on_hold_engineer = User.objects.create_user(
            username="engineer_onhold",
            password="pass12345",
            role=User.Roles.ON_HOLD,
            email="engineer.onhold@example.com",
        )
        self.account = Account.objects.create(name="On Hold Account")

    def test_new_request_form_excludes_on_hold_engineers_from_assignment_choices(self):
        form = RequestForm(actor_role=User.Roles.REQUESTOR)
        engineer_ids = set(form.fields["engineer"].queryset.values_list("id", flat=True))

        self.assertIn(self.engineer.pk, engineer_ids)
        self.assertNotIn(self.on_hold_engineer.pk, engineer_ids)

    def test_edit_request_form_keeps_current_on_hold_assignee_visible(self):
        request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Requestor On Hold",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.on_hold_engineer,
        )

        form = RequestForm(instance=request_obj, actor_role=User.Roles.REQUESTOR)
        engineer_ids = set(form.fields["engineer"].queryset.values_list("id", flat=True))

        self.assertIn(self.on_hold_engineer.pk, engineer_ids)

    def test_on_hold_user_dashboard_shows_assigned_requests(self):
        assigned_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Requestor On Hold",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.on_hold_engineer,
        )
        other_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Requestor On Hold",
            product_category="M365",
            engagement_type=Request.Engagement.TRAINING,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
        )

        request = self.factory.get(reverse("hub:dashboard"))
        request.user = self.on_hold_engineer

        view = DashboardView()
        view.setup(request)
        context = view.get_context_data()
        request_ids = {item.pk for item in context["requests"]}

        self.assertEqual(context["role"], User.Roles.ENGINEER)
        self.assertIn(assigned_request.pk, request_ids)
        self.assertNotIn(other_request.pk, request_ids)

    def test_on_hold_user_can_manage_previously_assigned_request(self):
        request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Requestor On Hold",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.on_hold_engineer,
        )

        request = self.factory.get(reverse("hub:request-manage-collab", args=[request_obj.pk]))
        request.user = self.on_hold_engineer

        view = RequestCollaborativeManageView()
        view.setup(request, pk=request_obj.pk)
        view.kwargs = {"pk": request_obj.pk}

        self.assertEqual(view.get_object(), request_obj)


class RequestActivityLogEditModalTests(TestCase):
    def setUp(self):
        self.requestor = User.objects.create_user(
            username="activity_modal_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="activity.modal.requestor@example.com",
        )
        self.engineer = User.objects.create_user(
            username="activity_modal_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="activity.modal.engineer@example.com",
        )
        self.account = Account.objects.create(name="Activity Modal Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="Activity Modal Manager",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
            description="Request for activity log editing modal test.",
        )
        self.activity_log = EngineerActivityLog.objects.create(
            engineer=self.engineer,
            account=self.account,
            request=self.request_obj,
            request_date=date.today(),
            activity_type=EngineerActivityLog.ActivityType.DEPLOYMENT,
            actual_hours=Decimal("2.50"),
            details="Fix deployment follow-up bug.",
            location=EngineerActivityLog.Location.OFFICE,
            is_billable=False,
            status=EngineerActivityLog.Status.IN_PROGRESS,
        )

    def test_manage_request_shows_edit_activity_modal_for_selected_log(self):
        self.client.force_login(self.engineer)

        response = self.client.get(
            reverse("hub:request-manage-collab", args=[self.request_obj.pk]),
            {"edit_activity": self.activity_log.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Edit Activity Report")
        self.assertContains(response, 'id="activity-log-form"')
        self.assertContains(response, 'name="form_type" value="activity_log"')
        self.assertContains(response, f'name="log_id" value="{self.activity_log.pk}"')


class SqrEngineerLinkedRequestAccessTests(TestCase):
    def setUp(self):
        self.engineer = User.objects.create_user(
            username="sqr_link_owner",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.link.owner@example.com",
        )
        self.other_engineer = User.objects.create_user(
            username="sqr_link_request_assignee",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.link.assignee@example.com",
        )
        self.pm = User.objects.create_user(
            username="sqr_link_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.link.pm@example.com",
            first_name="Linked",
            last_name="PM",
        )
        self.requestor = User.objects.create_user(
            username="sqr_link_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="sqr.link.requestor@example.com",
        )
        self.account = Account.objects.create(name="SQR Linked Request Account")
        self.linked_request = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="SQR Link Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.other_engineer,
        )
        self.submission = SqrSubmission.objects.create(
            linked_request=self.linked_request,
            engineer=self.engineer,
            pm_esg_reviewer=self.pm,
            customer_name="Linked Customer",
            customer_company="ESS",
            customer_contact="Linked Contact",
            project_title="Linked Project",
            project_details="Implementation",
            sse_manhrs=Decimal("20"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr",
        )

    def test_engineer_can_open_request_linked_to_owned_sqr(self):
        request = RequestFactory().get(reverse("hub:request-manage-collab", args=[self.linked_request.pk]))
        request.user = self.engineer
        view = RequestCollaborativeManageView()
        view.setup(request, pk=self.linked_request.pk)
        view.kwargs = {"pk": self.linked_request.pk}

        self.assertEqual(view.get_object(), self.linked_request)

    def test_engineer_can_save_own_sqr_with_preserved_linked_request(self):
        self.client.force_login(self.engineer)

        response = self.client.post(
            reverse("hub:sqr-edit", args=[self.submission.pk]),
            data={
                "linked_request": str(self.linked_request.pk),
                "customer_company": "ESS",
                "customer_contact": "Linked Contact",
                "pm_esg_reviewer": str(self.pm.pk),
                "customer_name": "Linked Customer Updated",
                "project_title": "Linked Project Updated",
                "project_details": "Implementation",
                "sse_manhrs": "20",
                "sqr_folder_link": "https://example.com/sqr",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.submission.refresh_from_db()
        self.assertEqual(self.submission.linked_request, self.linked_request)
        self.assertEqual(self.submission.customer_name, "Linked Customer Updated")


class SqrInlineUndoTests(TestCase):
    def setUp(self):
        self.engineer = User.objects.create_user(
            username="sqr_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.engineer@example.com",
        )
        self.pm = User.objects.create_user(
            username="sqr_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.pm@example.com",
            first_name="SQR",
            last_name="PM",
        )
        self.requestor = User.objects.create_user(
            username="sqr_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="sqr.requestor@example.com",
        )
        self.account = Account.objects.create(name="SQR Undo Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="SQR Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
        )
        self.submission = SqrSubmission.objects.create(
            linked_request=self.request_obj,
            engineer=self.engineer,
            pm_esg_reviewer=self.pm,
            customer_name="Undo Customer",
            customer_company="ESS",
            customer_contact="Undo Contact",
            project_title="Undo Project",
            project_details="Implementation",
            sse_manhrs=Decimal("20"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr",
        )
        self.client.force_login(self.pm)

    def _inline_update(self, field, value):
        return self.client.post(
            reverse("hub:sqr-inline-update", args=[self.submission.pk]),
            data=json.dumps({"field": field, "value": value}),
            content_type="application/json",
        )

    def _undo(self, change):
        return self.client.post(reverse("hub:sqr-inline-undo", args=[self.submission.pk, change.pk]))

    def test_inline_update_creates_undo_change(self):
        response = self._inline_update("pm_manhrs", "16")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["ok"])
        self.assertIn("undo_url", data)
        change = SqrSubmissionChange.objects.get(pk=data["change_id"])
        self.assertEqual(change.field, "pm_manhrs")
        self.assertEqual(change.old_values["pm_manhrs"], "8")
        self.assertEqual(change.new_values["pm_manhrs"], "16")
        self.assertEqual(change.old_values["pm_amount"], "24000")
        self.assertEqual(change.new_values["pm_amount"], "48000")

    def test_undo_restores_dependent_sse_values(self):
        response = self._inline_update("sse_manhrs", "50")
        self.assertEqual(response.status_code, 200)
        change = SqrSubmissionChange.objects.get(pk=response.json()["change_id"])
        self.submission.refresh_from_db()
        self.assertEqual(self.submission.sse_manhrs, Decimal("50"))
        self.assertEqual(self.submission.pm_manhrs, Decimal("24"))

        undo_response = self._undo(change)

        self.assertEqual(undo_response.status_code, 200)
        undo_data = undo_response.json()
        self.assertTrue(undo_data["ok"])
        self.assertTrue(undo_data["undone"])
        self.submission.refresh_from_db()
        change.refresh_from_db()
        self.assertEqual(self.submission.sse_manhrs, Decimal("20"))
        self.assertEqual(self.submission.sse_amount, Decimal("40000.00"))
        self.assertEqual(self.submission.pm_manhrs, Decimal("8"))
        self.assertEqual(self.submission.pm_amount, Decimal("24000.00"))
        self.assertIsNotNone(change.undone_at)
        self.assertEqual(change.undone_by, self.pm)

    def test_status_undo_restores_status_metadata_with_warning(self):
        response = self._inline_update("status", SqrSubmission.Status.APPROVED)
        self.assertEqual(response.status_code, 200)
        change = SqrSubmissionChange.objects.get(pk=response.json()["change_id"])
        self.submission.refresh_from_db()
        self.assertEqual(self.submission.status, SqrSubmission.Status.APPROVED)
        self.assertIsNotNone(self.submission.reviewed_at)
        self.assertIsNotNone(self.submission.validity_due_date)

        undo_response = self._undo(change)

        self.assertEqual(undo_response.status_code, 200)
        undo_data = undo_response.json()
        self.assertTrue(undo_data["workflow_side_effect_warning"])
        self.submission.refresh_from_db()
        self.assertEqual(self.submission.status, SqrSubmission.Status.FOR_PROCESSING)
        self.assertIsNone(self.submission.reviewed_at)
        self.assertIsNone(self.submission.validity_due_date)

    def test_stale_undo_is_rejected(self):
        response = self._inline_update("pm_manhrs", "16")
        self.assertEqual(response.status_code, 200)
        change = SqrSubmissionChange.objects.get(pk=response.json()["change_id"])
        SqrSubmission.objects.filter(pk=self.submission.pk).update(
            pm_manhrs=Decimal("24"),
            pm_amount=Decimal("72000.00"),
        )

        undo_response = self._undo(change)

        self.assertEqual(undo_response.status_code, 409)
        self.assertFalse(undo_response.json()["ok"])

    def test_non_pm_admin_cannot_inline_update(self):
        self.client.force_login(self.engineer)

        response = self._inline_update("pm_manhrs", "16")

        self.assertEqual(response.status_code, 403)
        self.assertEqual(SqrSubmissionChange.objects.count(), 0)


class SqrHistoryTests(TestCase):
    def setUp(self):
        self.engineer = User.objects.create_user(
            username="sqr_history_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.history.engineer@example.com",
        )
        self.pm = User.objects.create_user(
            username="sqr_history_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.history.pm@example.com",
            first_name="History",
            last_name="PM",
        )
        self.admin = User.objects.create_user(
            username="sqr_history_admin",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="sqr.history.admin@example.com",
        )
        self.requestor = User.objects.create_user(
            username="sqr_history_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="sqr.history.requestor@example.com",
        )
        self.account = Account.objects.create(name="SQR History Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.account,
            account_manager="SQR History Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            engineer=self.engineer,
        )
        self.submission = SqrSubmission.objects.create(
            linked_request=self.request_obj,
            engineer=self.engineer,
            pm_esg_reviewer=self.pm,
            customer_name="History Customer",
            customer_company="ESS",
            customer_contact="History Contact",
            project_title="History Project",
            project_details="Implementation",
            sse_manhrs=Decimal("20"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr",
        )
        self.client.force_login(self.pm)

    def _inline_update(self, field, value):
        return self.client.post(
            reverse("hub:sqr-inline-update", args=[self.submission.pk]),
            data=json.dumps({"field": field, "value": value}),
            content_type="application/json",
        )

    def test_inline_update_creates_permanent_history(self):
        response = self._inline_update("pm_manhrs", "16")

        self.assertEqual(response.status_code, 200)
        history = SqrSubmissionHistory.objects.get(submission=self.submission)
        self.assertEqual(history.action, SqrSubmissionHistory.Action.UPDATED)
        self.assertEqual(history.field, "pm_manhrs")
        self.assertEqual(history.old_values["pm_manhrs"], "8")
        self.assertEqual(history.new_values["pm_manhrs"], "16")
        self.assertIn("PM Man-hrs", history.summary)

    def test_pm_admin_can_fetch_history(self):
        self._inline_update("pm_manhrs", "16")

        pm_response = self.client.get(reverse("hub:sqr-history", args=[self.submission.pk]))
        self.assertEqual(pm_response.status_code, 200)
        self.assertTrue(pm_response.json()["ok"])
        self.assertEqual(len(pm_response.json()["entries"]), 1)

        self.client.force_login(self.admin)
        admin_response = self.client.get(reverse("hub:sqr-history", args=[self.submission.pk]))
        self.assertEqual(admin_response.status_code, 200)
        self.assertTrue(admin_response.json()["ok"])

    def test_engineer_cannot_fetch_history(self):
        self._inline_update("pm_manhrs", "16")
        self.client.force_login(self.engineer)

        response = self.client.get(reverse("hub:sqr-history", args=[self.submission.pk]))

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()["ok"])

    def test_restore_reverts_history_entry(self):
        response = self._inline_update("pm_manhrs", "16")
        self.assertEqual(response.status_code, 200)
        history = SqrSubmissionHistory.objects.get(action=SqrSubmissionHistory.Action.UPDATED)

        restore_response = self.client.post(reverse("hub:sqr-history-restore", args=[self.submission.pk, history.pk]))

        self.assertEqual(restore_response.status_code, 200)
        self.assertTrue(restore_response.json()["ok"])
        self.submission.refresh_from_db()
        history.refresh_from_db()
        self.assertEqual(self.submission.pm_manhrs, Decimal("8"))
        self.assertEqual(self.submission.pm_amount, Decimal("24000.00"))
        self.assertIsNotNone(history.restored_at)
        self.assertEqual(history.restored_by, self.pm)
        restore_entry = SqrSubmissionHistory.objects.get(action=SqrSubmissionHistory.Action.RESTORED)
        self.assertEqual(restore_entry.metadata["restored_history_id"], history.pk)

    def test_stale_restore_is_rejected(self):
        response = self._inline_update("pm_manhrs", "16")
        self.assertEqual(response.status_code, 200)
        history = SqrSubmissionHistory.objects.get(action=SqrSubmissionHistory.Action.UPDATED)
        SqrSubmission.objects.filter(pk=self.submission.pk).update(
            pm_manhrs=Decimal("24"),
            pm_amount=Decimal("72000.00"),
        )

        restore_response = self.client.post(reverse("hub:sqr-history-restore", args=[self.submission.pk, history.pk]))

        self.assertEqual(restore_response.status_code, 409)
        self.assertFalse(restore_response.json()["ok"])


class SqrExportTests(TestCase):
    ESG_HEADERS = [
        "Account Name",
        "Service Description",
        "Project Name",
        "SSE1",
        "SSE2",
        "SSE3",
        "Project?",
        "PM",
        "Status",
        "Start and Finish Dates",
        "% Complete",
        "Key Updates",
        "Active Risks/Issues",
        "Scope Status",
        "Schedule Status",
        "Budget Status",
        "Folder",
        "Quoted SSE hrs",
        "Billed SSE hrs",
        "Quoted PM hrs",
        "Billed PM hrs",
        "Project Management",
        "Deployment",
        "On-Call",
        "Total Revenue",
        "Revenue?",
        "SI Type",
        "SI Reference",
        "SI Date",
        "Remarks",
        "Action",
    ]

    SQR_HEADERS = [
        "SQR Date",
        "SQR ID",
        "Account Name",
        "Service Description",
        "Scope of Services",
        "RQ ID",
        "Group Name",
        "Account Manager",
        "Requester Name",
        "Approver Name",
        "SQR Doc. Ref. Link",
        "SSE Man-hrs",
        "SSE Amount",
        "PM Man-hrs",
        "PM Amount",
        "Maintenance Amt.",
        "Discount Rate (%)",
        "Discount Amount",
        "Total Price",
        "SQR Status",
        "Approval Date",
        "Validity Due Date",
        "Proposal Status",
        "Assigned PM",
        "Assigned SSE",
        "Start Date",
        "Target Finish Date",
        "Overall Status",
        "Health Status",
        "Overall Progress %",
        "Key Updates / Risks",
        "Actual Finish Date",
        "Completion Signed Date",
        "Completion Warranty End Date",
        "Maintenance Start Date",
        "Maintenance End Date",
        "PO / PNL Date",
        "PO Remarks",
        "Billing Type",
        "Billed Date",
        "Billing Reference",
        "Billing Status",
        "Billing Remarks",
    ]

    def setUp(self):
        self.engineer = User.objects.create_user(
            username="sqr_export_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.export.engineer@example.com",
            first_name="Export",
            last_name="Engineer",
        )
        self.sse = User.objects.create_user(
            username="sqr_export_sse",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.export.sse@example.com",
            first_name="Assigned",
            last_name="SSE",
        )
        self.pm = User.objects.create_user(
            username="sqr_export_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.export.pm@example.com",
            first_name="Export",
            last_name="PM",
        )
        self.assigned_pm = User.objects.create_user(
            username="sqr_export_assigned_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.export.assigned.pm@example.com",
            first_name="Assigned",
            last_name="PM",
        )
        self.client.force_login(self.pm)

    def test_esg_export_headers_and_row_match_esg_tracker_table(self):
        older = self._create_submission(customer_name="Older Account", project_title="Older Service")
        target = self._create_submission(
            customer_name="Acme Corp",
            project_title="Cloud Migration",
            project_details="Deployment Only",
            assigned_pm=self.assigned_pm,
            assigned_sse=self.sse,
            delivery_start_date=date(2026, 7, 1),
            delivery_target_finish_date=date(2026, 7, 31),
            delivery_progress=75,
            overall_status=SqrSubmission.OverallStatus.IN_PROGRESS,
            key_updates_risks="Migration is on track.",
            sse_manhrs=Decimal("12"),
            pm_manhrs=Decimal("8"),
            managed_support_amount=Decimal("5000.00"),
            revenue_date=date(2026, 8, 3),
            revenue_source="invoiced",
            revenue_reference_no="si-001",
            revenue_status=SqrSubmission.RevenueStatus.BILLED,
            revenue_remarks="Ready for billing.",
            revenue_declaration=SqrSubmission.RevenueDeclaration.DECLARED,
            sqr_folder_link="https://example.com/folder",
        )
        SqrSubmission.objects.filter(pk=older.pk).update(created_at=timezone.now() - timedelta(days=1))
        SqrSubmission.objects.filter(pk=target.pk).update(created_at=timezone.now())

        response = self.client.get(f"{reverse('hub:sqr-export')}?report=esg")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = workbook.active
        assert sheet is not None
        self.assertEqual(sheet.title, "ESG Performance Tracker")
        self.assertEqual([cell.value for cell in sheet[1]], self.ESG_HEADERS)
        self.assertEqual([sheet.cell(row=2, column=idx).value for idx in range(1, 32)], [
            "Acme Corp",
            "Cloud Migration",
            "Acme Corp - Cloud Migration",
            "Assigned SSE",
            None,
            None,
            "Yes",
            "Assigned PM",
            "ONGOING",
            "7/1/26-7/31/26",
            75,
            "Migration is on track.",
            "Migration is on track.",
            "Deployment Only",
            "7/31/26",
            None,
            "https://example.com/folder",
            12,
            None,
            8,
            None,
            24000,
            24000,
            5000,
            53000,
            "Yes",
            "Invoiced",
            "SI-001",
            "3-Aug-26",
            "Ready for billing.",
            "History",
        ])

    def test_sqr_export_keeps_original_headers(self):
        self._create_submission(customer_name="SQR Account", project_title="SQR Service")

        response = self.client.get(f"{reverse('hub:sqr-export')}?report=sqr")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = workbook.active
        assert sheet is not None
        self.assertEqual(sheet.title, "SQR")
        self.assertEqual([cell.value for cell in sheet[1]], self.SQR_HEADERS)

    def _create_submission(self, **overrides):
        data = {
            "engineer": self.engineer,
            "pm_esg_reviewer": self.pm,
            "customer_name": "Export Account",
            "customer_company": "ESS",
            "customer_contact": "Export Contact",
            "project_title": "Export Service",
            "project_details": "Implementation",
            "sse_manhrs": Decimal("20"),
            "documentation_links": "https://example.com/doc",
            "sqr_folder_link": "https://example.com/sqr",
        }
        data.update(overrides)
        return SqrSubmission.objects.create(**data)


class SqrMyAssignedFilterTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.engineer = User.objects.create_user(
            username="sqr_assigned_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.assigned.engineer@example.com",
        )
        self.pm = User.objects.create_user(
            username="sqr_assigned_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.assigned.pm@example.com",
            first_name="Assigned",
            last_name="Approver",
        )
        self.other_pm = User.objects.create_user(
            username="sqr_other_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.other.pm@example.com",
            first_name="Other",
            last_name="Approver",
        )
        self.admin = User.objects.create_user(
            username="sqr_assigned_admin",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="sqr.assigned.admin@example.com",
        )
        self.mine = SqrSubmission.objects.create(
            engineer=self.engineer,
            pm_esg_reviewer=self.pm,
            customer_name="Mine Account",
            customer_company="ESS",
            customer_contact="Mine Contact",
            project_title="Mine Service",
            project_details="Implementation",
            sse_manhrs=Decimal("10"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr-mine",
        )
        self.other = SqrSubmission.objects.create(
            engineer=self.engineer,
            pm_esg_reviewer=self.other_pm,
            customer_name="Other Account",
            customer_company="ESG",
            customer_contact="Other Contact",
            project_title="Other Service",
            project_details="Implementation",
            sse_manhrs=Decimal("12"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr-other",
        )

    def _context_for(self, user):
        request = self.factory.get(reverse("hub:sqr"))
        request.user = user
        view = SqrListView()
        view.setup(request)
        return view.get_context_data()

    def test_pm_esg_defaults_to_my_assigned_filter_but_keeps_all_rows(self):
        context = self._context_for(self.pm)

        self.assertTrue(context["default_my_assigned_filter"])
        self.assertTrue(context["is_pm_esg"])
        self.assertEqual(context["my_assigned_sqr_count"], 1)
        self.assertEqual(context["sqr_current_user_id"], self.pm.pk)
        # Full report access remains: both assigned and unassigned rows are still in the page.
        self.assertEqual(len(context["proposal_submissions"]), 2)
        self.assertEqual(context["proposal_counts"]["total"], 2)

    def test_admin_does_not_default_to_my_assigned_filter(self):
        context = self._context_for(self.admin)

        self.assertFalse(context["default_my_assigned_filter"])
        self.assertFalse(context["is_pm_esg"])
        self.assertEqual(len(context["proposal_submissions"]), 2)


class SqrReportsDashboardTests(TestCase):
    retained_chart_ids = (
        "sqrFunnelChart",
        "sqrDealStageChart",
        "sqrRevenueOverviewChart",
        "sqrMonthlyTrendChart",
        "sqrGroupDistributionChart",
    )
    removed_chart_ids = (
        "sqrQuickOverviewChart",
        "sqrStatusChart",
        "sqrDeliveryHealthChart",
        "sqrScopeDistributionChart",
    )

    def setUp(self):
        self.admin = User.objects.create_user(
            username="sqr_reports_admin",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="sqr.reports.admin@example.com",
        )
        self.pm = User.objects.create_user(
            username="sqr_reports_pm",
            password="pass12345",
            role=User.Roles.PM_ESG,
            email="sqr.reports.pm@example.com",
        )
        self.engineer = User.objects.create_user(
            username="sqr_reports_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="sqr.reports.engineer@example.com",
        )
        SqrSubmission.objects.create(
            engineer=self.engineer,
            pm_esg_reviewer=self.pm,
            customer_name="CRM Account",
            customer_company="Enterprise",
            customer_contact="CRM Contact",
            project_title="CRM Service",
            project_details="Implementation",
            sse_manhrs=Decimal("10"),
            documentation_links="https://example.com/doc",
            sqr_folder_link="https://example.com/sqr-report",
        )

    def assert_compact_dashboard_contract(self, html):
        self.assertIn("Sales Pipeline", html)
        self.assertIn("Won Value", html)
        self.assertIn("Top Won Accounts", html)
        self.assertIn("Highest-Value Wins", html)
        for chart_id in self.retained_chart_ids:
            self.assertIn(f'id="{chart_id}"', html)
        for chart_id in self.removed_chart_ids:
            self.assertNotIn(f'id="{chart_id}"', html)
        self.assertNotIn("What the dashboard is saying right now", html)
        self.assertNotIn("Executive summary", html)

    def test_admin_reports_page_renders_compact_sales_dashboard(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("hub:sqr"), {"tab": "reports"})

        self.assertEqual(response.status_code, 200)
        self.assert_compact_dashboard_contract(response.content.decode())

    def test_pm_reports_fragment_uses_same_dashboard_contract(self):
        self.client.force_login(self.pm)

        response = self.client.get(reverse("hub:sqr-reports-data"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assert_compact_dashboard_contract(payload["html"])

    def test_engineer_cannot_fetch_reports_fragment(self):
        self.client.force_login(self.engineer)

        response = self.client.get(reverse("hub:sqr-reports-data"))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json(), {"ok": False, "error": "Permission denied"})


class AdminAccountChangeTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.admin = User.objects.create_user(
            username="admin_account_change",
            password="pass12345",
            role=User.Roles.ADMIN,
            email="admin.account@example.com",
            first_name="Admin",
            last_name="Account",
        )
        self.requestor = User.objects.create_user(
            username="requestor_account_change",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="requestor.account@example.com",
            first_name="Req",
            last_name="Account",
        )
        self.original_account = Account.objects.create(name="Original Account")
        self.request_obj = Request.objects.create(
            requestor=self.requestor,
            account=self.original_account,
            account_manager="Req Account",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            description="Admin can reassign account.",
            start_date=date(2026, 8, 1),
        )

    def _make_used_account(self, name: str) -> Account:
        account = Account.objects.create(name=name)
        Request.objects.create(
            requestor=self.requestor,
            account=account,
            account_manager="Req Account",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            description=f"Anchor request for {name}",
            start_date=date(2026, 8, 1),
        )
        return account

    def test_admin_form_can_change_account_via_dropdown(self):
        existing = self._make_used_account("Existing Target Account")
        form = RequestAdminForm(
            data={
                "account_name": existing.name,
                "request_date": "2026-08-01",
                "requestor": str(self.requestor.pk),
                "priority": Request.Priority.MEDIUM,
                "status": self.request_obj.status,
                "engineer": "",
                "backup_engineer": "",
                "due_date": "",
                "end_date": "",
                "description": self.request_obj.description,
            },
            instance=self.request_obj,
            allow_capacity_override=True,
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.account_id, existing.pk)

    def test_admin_manage_view_exposes_account_field_and_saves_change(self):
        target = self._make_used_account("Admin Reassigned Account")
        request = self.factory.get(reverse("hub:request-manage", args=[self.request_obj.pk]))
        request.user = self.admin
        AssignmentEmailTests._attach_session_and_messages(request)
        response = RequestAdminUpdateView.as_view()(request, pk=self.request_obj.pk)
        self.assertEqual(response.status_code, 200)
        response.render()
        html = response.content.decode()
        self.assertIn('name="account_name"', html)
        self.assertIn("form-control", html)
        self.assertIn('data-admin-account-field="true"', html)
        self.assertIn('data-account-autocomplete="true"', html)
        self.assertRegex(html, r'<input[^>]*name="account_name"[^>]*class="[^"]*form-control')
        self.assertIn('id="account-name-options"', html)
        self.assertIn("Original Account", html)
        self.assertIn(target.name, html)
        self.assertIn("account-autocomplete", html)
        self.assertIn("initAccountAutocompletes", html)

        post = self.factory.post(
            reverse("hub:request-manage", args=[self.request_obj.pk]),
            data={
                "form_type": "details",
                "override_capacity": "0",
                "account_name": target.name,
                "request_date": "2026-08-01",
                "requestor": str(self.requestor.pk),
                "priority": Request.Priority.MEDIUM,
                "status": self.request_obj.status,
                "engineer": "",
                "backup_engineer": "",
                "due_date": "",
                "end_date": "",
                "description": self.request_obj.description,
            },
        )
        post.user = self.admin
        AssignmentEmailTests._attach_session_and_messages(post)
        with patch("hub.views.notify_engineer_assignment_email", return_value=AssignmentEmailResult("no_new_assignee")):
            with patch("hub.views.notify_engineer_assignment_notification"):
                response = RequestAdminUpdateView.as_view()(post, pk=self.request_obj.pk)
        self.assertEqual(response.status_code, 302)
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.account_id, target.pk)
        self.assertEqual(self.request_obj.account.name, "Admin Reassigned Account")

    def test_delete_user_account_blocks_when_user_has_protected_requests(self):
        engineer = User.objects.create_user(
            username="engineer_protected_delete",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="engineer.protected@example.com",
            first_name="Protected",
            last_name="Engineer",
        )
        account = Account.objects.create(name="Protected Request Account")
        request_obj = Request.objects.create(
            requestor=self.requestor,
            account=account,
            account_manager="Req Account",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            description="Engineer is referenced by a request.",
            start_date=date(2026, 8, 1),
            engineer=engineer,
        )

        request = self.factory.post(reverse("hub:management"), data={})
        request.user = self.admin
        AssignmentEmailTests._attach_session_and_messages(request)

        response = UserManagementView()._delete_user_account(request, engineer)

        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(pk=engineer.pk).exists())
        self.assertEqual(Request.objects.get(pk=request_obj.pk).engineer_id, engineer.pk)
        messages = [message.message for message in get_messages(request)]
        self.assertTrue(any("cannot delete" in message.lower() for message in messages))
        self.assertTrue(any("request" in message.lower() for message in messages))


class UnusedAccountPruneTests(TestCase):
    def setUp(self):
        self.requestor = User.objects.create_user(
            username="unused_acct_requestor",
            password="pass12345",
            role=User.Roles.REQUESTOR,
            email="unused.acct.req@example.com",
            first_name="Unused",
            last_name="Requestor",
        )
        self.engineer = User.objects.create_user(
            username="unused_acct_engineer",
            password="pass12345",
            role=User.Roles.ENGINEER,
            email="unused.acct.eng@example.com",
            first_name="Unused",
            last_name="Engineer",
        )
        self.used_account = Account.objects.create(name="Used Live Account")
        self.live_request = Request.objects.create(
            requestor=self.requestor,
            account=self.used_account,
            account_manager="Unused Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            description="Live request keeps account used.",
            start_date=date(2026, 8, 1),
        )
        self.soft_deleted_account = Account.objects.create(name="Soft Deleted Only Account")
        self.soft_request = Request.all_objects.create(
            requestor=self.requestor,
            account=self.soft_deleted_account,
            account_manager="Unused Requestor",
            product_category="Azure",
            engagement_type=Request.Engagement.SUPPORT,
            priority=Request.Priority.MEDIUM,
            description="Soft-deleted request still counts.",
            start_date=date(2026, 8, 1),
            is_deleted=True,
            deleted_at=timezone.now(),
        )
        self.unused_account = Account.objects.create(name="Truly Unused Account")
        self.activity_only_account = Account.objects.create(name="Activity Log Only Account")
        EngineerActivityLog.objects.create(
            engineer=self.engineer,
            account=self.activity_only_account,
            request_date=date(2026, 8, 1),
            activity_type=EngineerActivityLog.ActivityType.INTERNAL_SUPPORT,
            actual_hours=Decimal("1.00"),
            details="Log without request",
            location=EngineerActivityLog.Location.OFFICE,
            is_billable=False,
            status=EngineerActivityLog.Status.COMPLETED,
        )

    def test_used_queryset_includes_live_and_soft_deleted_excludes_unused(self):
        names = set(Account.used_queryset().values_list("name", flat=True))
        self.assertIn(self.used_account.name, names)
        self.assertIn(self.soft_deleted_account.name, names)
        self.assertNotIn(self.unused_account.name, names)
        self.assertNotIn(self.activity_only_account.name, names)

    def test_request_form_suggestions_exclude_unused(self):
        form = RequestForm(actor_role=User.Roles.REQUESTOR, actor_user=self.requestor)
        self.assertIn(self.used_account.name, form.account_name_suggestions)
        self.assertIn(self.soft_deleted_account.name, form.account_name_suggestions)
        self.assertNotIn(self.unused_account.name, form.account_name_suggestions)
        self.assertNotIn(self.activity_only_account.name, form.account_name_suggestions)

    def test_admin_form_keeps_current_and_excludes_unused(self):
        form = RequestAdminForm(instance=self.live_request, allow_capacity_override=True)
        self.assertIn(self.used_account.name, form.account_name_suggestions)
        self.assertNotIn(self.unused_account.name, form.account_name_suggestions)
        self.assertEqual(form.fields["account_name"].initial, self.used_account.name)
        self.assertEqual(form.fields["account_name"].widget.attrs.get("data-account-autocomplete"), "true")
        self.assertEqual(form.fields["account_name"].__class__.__name__, "CharField")

    def test_sqr_and_filter_and_activity_forms_use_used_accounts(self):
        sqr_form = SqrSubmissionForm()
        self.assertIn(self.used_account.name, sqr_form.account_name_options)
        self.assertNotIn(self.unused_account.name, sqr_form.account_name_options)

        filter_form = AdminRequestFilterForm()
        filter_names = set(filter_form.fields["account"].queryset.values_list("name", flat=True))
        self.assertIn(self.used_account.name, filter_names)
        self.assertNotIn(self.unused_account.name, filter_names)

        activity_form = EngineerActivityLogForm(engineer=self.engineer)
        activity_names = set(activity_form.fields["account"].queryset.values_list("name", flat=True))
        self.assertIn(self.used_account.name, activity_names)
        self.assertNotIn(self.unused_account.name, activity_names)
        self.assertNotIn(self.activity_only_account.name, activity_names)

    def test_prune_command_deletes_unused_skips_protected_and_soft_deleted(self):
        call_command("prune_unused_accounts", "--apply")
        remaining = set(Account.objects.values_list("name", flat=True))
        self.assertIn(self.used_account.name, remaining)
        self.assertIn(self.soft_deleted_account.name, remaining)
        self.assertIn(self.activity_only_account.name, remaining)
        self.assertNotIn(self.unused_account.name, remaining)

    def test_management_baseline_does_not_reseed_when_empty(self):
        EngineerActivityLog.objects.all().delete()
        Request.all_objects.all().delete()
        Account.objects.all().delete()
        self.assertEqual(Account.objects.count(), 0)
        UserManagementView._sync_account_baseline()
        self.assertEqual(Account.objects.count(), 0)
