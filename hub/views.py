import csv
import base64
import html
from django.db import IntegrityError
import json
import logging
import mimetypes
import re
import threading
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal
from datetime import date, datetime, timedelta
from email.charset import Charset, QP
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import msal
import requests

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.sessions.models import Session
from django.core.cache import cache
from django.db import close_old_connections
from django.core.paginator import InvalidPage, Paginator
from django.db import transaction
from django.db.models import Count, Min, Q, Sum
from django.db.models.functions import TruncMonth
from django.forms import modelformset_factory
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import DetailView, DeleteView, ListView, TemplateView, UpdateView
from urllib.parse import quote, urlencode

from accounts.forms import UserManagementForm
from accounts.models import User

REQUESTOR_ROLES = set(User.REQUESTOR_ROLES)
REQUEST_CREATOR_ROLES = set(getattr(User, "REQUEST_CREATOR_ROLES", User.REQUESTOR_ROLES))
ENGINEER_ACCESS_ROLES = set(getattr(User, "ENGINEER_ACCESS_ROLES", (User.Roles.ENGINEER,)))
ASSIGNABLE_ENGINEER_ROLES = set(getattr(User, "ASSIGNABLE_ENGINEER_ROLES", (User.Roles.ENGINEER,)))
PM_ESS_ROLE = User.Roles.PM_ESS
PM_ESG_ROLE = User.Roles.PM_ESG
ADMIN_PANEL_ROLES = {User.Roles.ADMIN, PM_ESG_ROLE}
SQR_ACCESS_ROLES = ENGINEER_ACCESS_ROLES | {User.Roles.ADMIN, PM_ESG_ROLE}

from .forms import (
    AccountManagementForm,
    AdminRequestFilterForm,
    EngineerActivityLogForm,
    RequestAdminForm,
    SqrDeliveryForm,
    SqrProposalStatusForm,
    SqrRevenueForm,
    SqrRevenueOrderForm,
    SqrRevenueQuotationForm,
    SqrImportForm,
    RequestForm,
    RequestStatusForm,
    SqrReviewForm,
    SqrSubmissionForm,
    StatusLogForm,
)
from .constants import ACCOUNT_NAME_RAW
from .models import (
    Account,
    EngineerActivityLog,
    Notification,
    Request,
    RequestCommunication,
    SqrSubmission,
    SqrSubmissionChange,
    SqrSubmissionHistory,
    StatusLog,
)
from .mixins import (
    AdminOrEngineerRequiredMixin,
    AdminOrPmEsgRequiredMixin,
    AdminRequiredMixin,
    EngineerRequiredMixin,
)

MANILA_TZ = ZoneInfo("Asia/Manila")
logger = logging.getLogger(__name__)

SQR_APPROVAL_SERVICE_COMPONENT = "Systems Support & Maintenance Service - 1 Year"
SQR_APPROVAL_LOGO_URL = "https://raw.githubusercontent.com/lloydismael/request-hub/dev/static/img/phil-data-full-logo.png"
SQR_APPROVAL_LOGO_PATH = Path(settings.BASE_DIR) / "static" / "img" / "phil-data-full-logo.png"
SQR_APPROVAL_LOGO_CID = "sqr-approval-logo@request-hub"
SQR_APPROVAL_INCLUDED_SERVICES = [
    "Proactive System Health Checks",
    "System/Platform Patching (scheduled)",
    "Incident Support",
    "Basic Troubleshooting and Issue Isolation",
    "Monthly System Status Report",
    "Advisory Support",
]


def _format_sqr_approval_greeting(requestor_name: str) -> str:
    cleaned_name = (requestor_name or "").strip()
    return f"Hi @{cleaned_name}" if cleaned_name else "Hi"


def _split_sqr_scope_items(scope_text: str) -> list[str]:
    cleaned = (scope_text or "").replace("\r", "").strip()
    if not cleaned:
        return list(SQR_APPROVAL_INCLUDED_SERVICES)

    if "•" in cleaned:
        raw_items = cleaned.split("•")
    elif "\n" in cleaned:
        raw_items = cleaned.splitlines()
    elif ";" in cleaned:
        raw_items = cleaned.split(";")
    else:
        raw_items = [cleaned]

    items: list[str] = []
    for raw_item in raw_items:
        item = re.sub(r"^[\-•\s]+", "", raw_item).strip()
        if item:
            items.append(item)
    return items or list(SQR_APPROVAL_INCLUDED_SERVICES)


def _get_sqr_scope_items(submission: "SqrSubmission") -> list[str]:
    return _split_sqr_scope_items(submission.project_details or "")


def _get_sqr_service_component(submission: "SqrSubmission") -> str:
    return SQR_APPROVAL_SERVICE_COMPONENT


def _get_sqr_logo_attachment() -> Optional[dict]:
    try:
        logo_bytes = SQR_APPROVAL_LOGO_PATH.read_bytes()
    except OSError:
        logger.warning("SQR approval logo file is unavailable at %s", SQR_APPROVAL_LOGO_PATH)
        return None

    content_type = mimetypes.guess_type(str(SQR_APPROVAL_LOGO_PATH))[0] or "image/png"
    return {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": SQR_APPROVAL_LOGO_PATH.name,
        "contentType": content_type,
        "isInline": True,
        "contentId": SQR_APPROVAL_LOGO_CID,
        "contentBytes": base64.b64encode(logo_bytes).decode("ascii"),
    }


def _build_sqr_approval_mime_message(context: dict, sender_email: str, *, mark_as_unsent: bool = False) -> str:
    windows_charset = Charset("windows-1252")
    windows_charset.body_encoding = QP

    message = MIMEMultipart("related")
    message.set_param("type", "multipart/alternative")
    message["Subject"] = Header(context["subject"], "windows-1252")
    message["From"] = sender_email
    recipients = context.get("recipients") or []
    if recipients[:1]:
        message["To"] = recipients[0]
    if recipients[1:]:
        message["Cc"] = ", ".join(recipients[1:])
    if mark_as_unsent:
        message["X-Unsent"] = "1"

    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(context["plain_text"], "plain", windows_charset))
    alternative.attach(MIMEText(context.get("html_draft") or context["html_preview"], "html", windows_charset))
    message.attach(alternative)

    for attachment in context.get("attachments") or []:
        raw_bytes = base64.b64decode(attachment["contentBytes"])
        maintype, _, subtype = (attachment.get("contentType") or "application/octet-stream").partition("/")
        if maintype == "image":
            part = MIMEImage(raw_bytes, _subtype=subtype or "png")
        else:
            part = MIMEApplication(raw_bytes, _subtype=subtype or "octet-stream")
        part.add_header("Content-Disposition", "inline", filename=attachment.get("name") or "attachment")
        if attachment.get("contentId"):
            part.add_header("Content-ID", f'<{attachment["contentId"]}>')
        message.attach(part)

    return base64.b64encode(message.as_bytes()).decode("ascii")


def _format_sqr_approval_subject(submission: "SqrSubmission") -> str:
    rq_id = submission.linked_request.reference_code if submission.linked_request else ""
    if rq_id:
        return f"{submission.reference_code} • {rq_id} • {submission.customer_name}"
    return f"{submission.reference_code} • {submission.customer_name}"


def _format_sqr_approval_date(submission: "SqrSubmission") -> str:
    if not submission.reviewed_at:
        return "—"
    return submission.reviewed_at.astimezone(MANILA_TZ).strftime("%B %d, %Y")


def _format_sqr_approval_total_price(submission: "SqrSubmission") -> str:
    if submission.computed_total_price is not None:
        return f"PHP {submission.computed_total_price:,.2f}"
    if submission.quotation_total_price is not None:
        return f"PHP {submission.quotation_total_price:,.2f}"
    return "—"


def _format_sqr_approval_validity_date(submission: "SqrSubmission") -> str:
    return submission.validity_due_date.strftime("%B %d, %Y") if submission.validity_due_date else "—"


def _build_sqr_approval_body_lines(submission: "SqrSubmission", requestor_name: str) -> list[str]:
    service_component = _get_sqr_service_component(submission)
    total_price = _format_sqr_approval_total_price(submission)
    approval_date = _format_sqr_approval_date(submission)
    validity_date = _format_sqr_approval_validity_date(submission)
    customer_name = submission.customer_name or "—"
    service_description = (submission.project_title or "").strip() or "—"
    account_manager = (submission.customer_contact or "").strip() or "—"
    scope_of_services = (submission.project_details or "").strip() or "—"
    summary_rows = [
        ("SQR ID", submission.reference_code or "—"),
        ("Customer Name", customer_name),
        ("Service Description", service_description),
        ("Account Manager", account_manager),
        ("Scope of Services", scope_of_services),
    ]
    summary_width = max(len(label) for label, _ in summary_rows) + 2

    def format_label_value(label: str, value: str) -> str:
        return f"{label:<{summary_width}}: {value}"

    return [
        "Submitted SQR is now approved, please refer to the ff. details below.",
        "",
        format_label_value("SQR ID", submission.reference_code or "—"),
        format_label_value("Customer Name", customer_name),
        format_label_value("Service Description", service_description),
        format_label_value("Account Manager", account_manager),
        format_label_value("Scope of Services", scope_of_services),
        format_label_value("Add-On Service", service_component),
        "Included Services:",
        "• Proactive System Health Checks",
        "• System/Platform Patching (scheduled)",
        "• Incident Support",
        "• Basic Troubleshooting and Issue Isolation",
        "• Monthly System Status Report",
        "• Advisory Support",
        "",
        format_label_value("Quantity", "1 Lot"),
        format_label_value("Total Price", total_price),
        format_label_value("Approval Date", approval_date),
        format_label_value("Quotation Validity Until", validity_date),
        "",
        "**TERMS AND CONDITIONS**",
        "VAT: This quote excludes Value Added Tax (VAT).",
        "For P&L documentation purposes, a VAT-inclusive total may be applied. Internal billing and revenue reporting remain VAT-exclusive.",
        "This is a budgetary quote.",
        "This quote is issued for internal billing purposes only.",
        "Travel costs within Metro Manila are included.",
        "This quote does not include hardware, software licenses, or subscriptions unless stated.",
        "",
        "For any questions or to discuss this quote further, please don't hesitate to contact us:",
        "EnterpriseServices@phildata.com",
    ]


def _build_sqr_approval_html(submission: "SqrSubmission", requestor_name: str) -> str:
    return _build_sqr_approval_html_markup(submission, requestor_name, logo_src=SQR_APPROVAL_LOGO_URL)


def _build_sqr_approval_draft_html(submission: "SqrSubmission", requestor_name: str) -> str:
    return _build_sqr_approval_html_markup(submission, requestor_name, logo_src=f"cid:{SQR_APPROVAL_LOGO_CID}")


def _build_sqr_approval_html_markup(submission: "SqrSubmission", requestor_name: str, *, logo_src: str) -> str:
    summary_rows = [
        ("SQR ID", submission.reference_code or "—"),
        ("Account / Customer", submission.customer_name or "—"),
        ("Service Description", (submission.project_title or "").strip() or "—"),
        ("Account Manager", (submission.customer_contact or "").strip() or "—"),
    ]
    service_component = html.escape(_get_sqr_service_component(submission))
    approval_date = html.escape(_format_sqr_approval_date(submission))
    total_price = html.escape(_format_sqr_approval_total_price(submission))
    validity_date = html.escape(_format_sqr_approval_validity_date(submission))
    greeting = html.escape(_format_sqr_approval_greeting(requestor_name))
    logo_url = html.escape(logo_src)

    summary_headers_html = "".join(
        f'<th style="direction:ltr;text-align:left;border-width:1px;border-style:solid;border-color:rgb(15,77,103);background-color:rgb(21,96,130);padding:12px 14px;color:rgb(255,255,255);"><div class="elementToProof" style="direction:ltr;text-align:left;font-family:Arial,Helvetica,sans-serif;font-size:12px;"><span style="font-weight:700;">{html.escape(label)}</span></div></th>'
        for label, _value in summary_rows
    )
    summary_values_html = "".join(
        f'<td style="direction:ltr;border-width:1px;border-style:solid;border-color:rgb(207,216,227);padding:14px;vertical-align:top;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;">{html.escape(value)}</div></td>'
        for _label, value in summary_rows
    )

    return "".join(
        [
            '<html><head><meta http-equiv="Content-Type" content="text/html; charset=Windows-1252"><style type="text/css" style="display:none;"> P {margin-top:0;margin-bottom:0;} </style></head><body dir="ltr">',
            '<div class="elementToProof" style="background-color:rgb(244,247,251);padding:24px 12px;">',
            '<div class="elementToProof" style="background-color:rgb(255,255,255);margin:0 auto;border-width:1px;border-style:solid;border-color:rgb(215,222,232);max-width:860px;">',
            '<div class="elementToProof" style="padding:28px 30px 32px;">',
            f'<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 12px;font-size:14px;"><span style="font-family:Arial,Helvetica,sans-serif;color:rgb(16,24,40);">{greeting}</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 22px;font-size:14px;"><span style="font-family:Arial,Helvetica,sans-serif;color:rgb(16,24,40);">Submitted SQR is now approved. Please refer to the quotation details below.</span></p>',
            '<div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:14px;color:rgb(16,24,40);">',
            f'<img width="193" height="80" style="width:193.667px;height:80px;max-width:1070px;" src="{logo_url}" alt="Phil-Data">',
            '</div>',
            '<table role="presentation" style="direction:ltr;margin:0 0 18px;width:100%;box-sizing:border-box;border-collapse:collapse;border-spacing:0;">',
            '<tbody><tr>',
            '<td style="direction:ltr;vertical-align:bottom;">',
            '<div class="elementToProof" style="direction:ltr;letter-spacing:0.08em;font-family:Arial,Helvetica,sans-serif;font-size:11px;font-weight:700;color:rgb(71,84,103);"><span style="letter-spacing:0.08em;">Date</span></div>',
            f'<div class="elementToProof" style="direction:ltr;margin-top:6px;font-family:Arial,Helvetica,sans-serif;font-size:13px;color:rgb(16,24,40);">{approval_date}</div>',
            '</td>',
            '<td style="direction:ltr;text-align:right;vertical-align:bottom;">',
            '<div class="elementToProof" style="direction:ltr;text-align:right;letter-spacing:0.05em;font-family:Arial,Helvetica,sans-serif;font-size:28px;font-weight:700;color:rgb(17,24,39);"><span style="letter-spacing:0.05em;">SERVICE QUOTATION</span></div>',
            '</td>',
            '</tr></tbody></table>',
            '<table role="presentation" style="direction:ltr;width:100%;table-layout:fixed;box-sizing:border-box;border-collapse:collapse;border-spacing:0;">',
            '<tbody>',
            f'<tr>{summary_headers_html}</tr>',
            f'<tr>{summary_values_html}</tr>',
            '</tbody></table>',
            '<table role="presentation" style="direction:ltr;margin-top:18px;width:100%;box-sizing:border-box;border-collapse:collapse;border-spacing:0;">',
            '<tbody>',
            '<tr>',
            '<th style="direction:ltr;text-align:left;border-width:1px;border-style:solid;border-color:rgb(15,77,103);background-color:rgb(21,96,130);padding:12px 14px;color:rgb(255,255,255);"><div class="elementToProof" style="direction:ltr;text-align:left;font-family:Arial,Helvetica,sans-serif;font-size:12px;"><span style="font-weight:700;">Service Component</span></div></th>',
            '<th style="direction:ltr;text-align:left;border-width:1px;border-style:solid;border-color:rgb(15,77,103);background-color:rgb(21,96,130);padding:12px 14px;color:rgb(255,255,255);width:120px;"><div class="elementToProof" style="direction:ltr;text-align:left;font-family:Arial,Helvetica,sans-serif;font-size:12px;"><span style="font-weight:700;">Qty</span></div></th>',
            '<th style="direction:ltr;text-align:left;border-width:1px;border-style:solid;border-color:rgb(15,77,103);background-color:rgb(21,96,130);padding:12px 14px;color:rgb(255,255,255);width:180px;"><div class="elementToProof" style="direction:ltr;text-align:left;font-family:Arial,Helvetica,sans-serif;font-size:12px;"><span style="font-weight:700;">Total Price</span></div></th>',
            '</tr>',
            '<tr>',
            f'<td style="direction:ltr;border-width:1px;border-style:solid;border-color:rgb(207,216,227);padding:14px;vertical-align:top;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;">{service_component}</div></td>',
            '<td style="direction:ltr;border-width:1px;border-style:solid;border-color:rgb(207,216,227);padding:14px;vertical-align:top;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;">1 lot</div></td>',
            f'<td style="direction:ltr;border-width:1px;border-style:solid;border-color:rgb(207,216,227);padding:14px;vertical-align:top;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;">{total_price}</div></td>',
            '</tr>',
            '<tr>',
            '<td style="direction:ltr;border-right:1px solid rgb(207,216,227);border-bottom:1px solid rgb(207,216,227);border-left:1px solid rgb(207,216,227);padding:14px;"></td>',
            '<td style="direction:ltr;border-right:1px solid rgb(207,216,227);border-bottom:1px solid rgb(207,216,227);border-left:1px solid rgb(207,216,227);padding:14px;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;"><span style="font-weight:700;">Total</span></div></td>',
            f'<td style="direction:ltr;border-right:1px solid rgb(207,216,227);border-bottom:1px solid rgb(207,216,227);border-left:1px solid rgb(207,216,227);padding:14px;color:rgb(16,24,40);"><div class="elementToProof" style="direction:ltr;font-family:Arial,Helvetica,sans-serif;font-size:13px;"><span style="font-weight:700;">{total_price}</span></div></td>',
            '</tr>',
            '</tbody></table>',
            f'<p class="elementToProof" style="direction:ltr;margin:14px 0 0;font-size:13px;color:rgb(71,84,103);"><span style="font-family:Arial,Helvetica,sans-serif;">Quotation Validity Until: </span><span style="font-family:Arial,Helvetica,sans-serif;color:rgb(16,24,40);"><b>{validity_date}</b></span></p>',
            '<div class="elementToProof" style="margin-top:24px;">',
            '<div class="elementToProof" style="direction:ltr;margin:0 0 12px;font-family:Arial,Helvetica,sans-serif;font-size:16px;color:rgb(16,24,40);"><span style="font-weight:700;">Terms and Conditions</span></div>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">VAT: This quote excludes Value Added Tax (VAT).</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;"><i>For P&amp;L documentation purposes, a VAT-inclusive total may be applied. Internal billing and revenue reporting remain VAT-exclusive.</i></span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">This is a budgetary quote.</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">This quote is issued for internal billing purposes only.</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">Travel costs within Metro Manila are included.</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 8px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">This quote does not include hardware, software licenses, or subscriptions unless stated.</span></p>',
            '</div>',
            '<div class="elementToProof" style="margin-top:22px;padding-top:16px;border-top:1px solid rgb(228,231,236);">',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0 0 6px;font-size:13px;color:rgb(16,24,40);"><span style="font-family:Arial,Helvetica,sans-serif;">For any questions or to discuss this quote further, please don\'t hesitate to contact us:</span></p>',
            '<p class="elementToProof" style="direction:ltr;line-height:1.6;margin:0;font-size:13px;"><span style="font-family:Arial,Helvetica,sans-serif;color:rgb(21,96,130);font-weight:700;"><a href="mailto:EnterpriseServices@phildata.com" style="color:rgb(21,96,130);text-decoration:none;margin-top:0;margin-bottom:0;">EnterpriseServices@phildata.com</a></span></p>',
            '</div>',
            '</div>',
            '</div>',
            '</div>',
            '</body></html>',
        ]
    )


@dataclass(frozen=True)
class AssignmentEmailResult:
    status: str
    recipients: tuple[str, ...] = ()


def flash_assignment_email_feedback(request, result: AssignmentEmailResult, *, action_label: str, notify_on_no_assignee: bool = False) -> None:
    if result.status == "sent":
        return

    if result.status == "no_new_assignee":
        if notify_on_no_assignee:
            messages.info(
                request,
                f"Request {action_label}, but no assignment email was sent because no engineer or backup engineer was assigned.",
            )
        return

    if result.status == "missing_assignee_email":
        messages.warning(
            request,
            f"Request {action_label}, but no assignment email was sent because the assigned engineer or backup engineer has no email address configured.",
        )
        return

    if result.status == "missing_acs_config":
        messages.warning(
            request,
            f"Request {action_label}, but Azure Communication Services email is not configured.",
        )
        return

    if result.status == "delivery_failed":
        messages.error(
            request,
            f"Request {action_label}, but the assignment email could not be delivered from {settings.ACS_EMAIL_SENDER or 'the configured sender address'}. Check Azure Communication Services email configuration and logs.",
        )


class EngineerActivityLogView(EngineerRequiredMixin, LoginRequiredMixin, TemplateView):
    template_name = "hub/activity_logs.html"
    form_class = EngineerActivityLogForm

    def _parse_month_filter(self):
        month_value = (self.request.GET.get("month") or "").strip()
        if not month_value:
            return "", None, None
        try:
            parsed = datetime.strptime(month_value, "%Y-%m")
        except ValueError:
            return "", None, None
        return month_value, parsed.year, parsed.month

    def get_queryset(self):
        queryset = (
            EngineerActivityLog.objects.filter(engineer=self.request.user)
            .select_related("account", "request")
            .order_by("-request_date", "-created_at")
        )
        _, year, month = self._parse_month_filter()
        if year and month:
            queryset = queryset.filter(request_date__year=year, request_date__month=month)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = kwargs.get("form")
        editing_log = kwargs.get("editing_log")
        if form is None:
            form = self.form_class(engineer=self.request.user)

        # Full queryset — used for aggregates, month options, charts
        full_qs = self.get_queryset()
        logs_total_count = full_qs.count()

        agg = full_qs.aggregate(
            total_hours=Sum("actual_hours"),
            billable_hours=Sum("actual_hours", filter=Q(is_billable=True)),
        )
        total_hours_val = agg.get("total_hours") or Decimal("0")
        billable_hours_val = agg.get("billable_hours") or Decimal("0")

        # Paginate to 50 rows per page
        try:
            page_num = max(1, int(self.request.GET.get("page") or 1))
        except (TypeError, ValueError):
            page_num = 1
        paginator = Paginator(full_qs, 50)
        try:
            logs_page_obj = paginator.page(page_num)
        except InvalidPage:
            logs_page_obj = paginator.page(1)
        logs = list(logs_page_obj.object_list)

        selected_month, _, _ = self._parse_month_filter()
        month_rows = (
            EngineerActivityLog.objects.filter(engineer=self.request.user)
            .annotate(month=TruncMonth("request_date"))
            .values("month")
            .annotate(total=Count("id"))
            .order_by("-month")
        )
        month_options = [
            {
                "value": row["month"].strftime("%Y-%m"),
                "label": row["month"].strftime("%B %Y"),
                "count": row["total"],
            }
            for row in month_rows
            if row.get("month")
        ]
        selected_month_label = ""
        if selected_month:
            selected_month_label = next(
                (option["label"] for option in month_options if option["value"] == selected_month),
                "",
            )

        related_requests = Request.objects.filter(
            Q(engineer=self.request.user) | Q(backup_engineer=self.request.user)
        ).distinct()

        request_status_counts = related_requests.values("status").annotate(total=Count("id")).order_by("status")
        request_priority_counts = related_requests.values("priority").annotate(total=Count("id")).order_by("priority")
        activity_type_counts = (
            EngineerActivityLog.objects.filter(engineer=self.request.user)
            .values("activity_type")
            .annotate(total=Count("id"))
            .order_by("activity_type")
        )

        monthly_activity_rows = (
            EngineerActivityLog.objects.filter(engineer=self.request.user)
            .annotate(month=TruncMonth("request_date"))
            .values("month")
            .annotate(total_logs=Count("id"), total_hours=Sum("actual_hours"))
            .order_by("month")
        )

        activity_report_data = {
            "request_status": {
                "labels": [Request.Status(value["status"]).label for value in request_status_counts],
                "values": [value["total"] for value in request_status_counts],
            },
            "request_priority": {
                "labels": [Request.Priority(value["priority"]).label for value in request_priority_counts],
                "values": [value["total"] for value in request_priority_counts],
            },
            "activity_type": {
                "labels": [EngineerActivityLog.ActivityType(value["activity_type"]).label for value in activity_type_counts],
                "values": [value["total"] for value in activity_type_counts],
            },
            "monthly": {
                "labels": [value["month"].strftime("%b %Y") for value in monthly_activity_rows if value["month"]],
                "hours": [float(value["total_hours"] or 0) for value in monthly_activity_rows if value["month"]],
                "entries": [value["total_logs"] for value in monthly_activity_rows if value["month"]],
            },
        }

        context.update(
            {
                "form": form,
                "logs": logs,
                "logs_page_obj": logs_page_obj,
                "logs_total_count": logs_total_count,
                "hours_summary": {
                    "total": total_hours_val,
                    "billable": billable_hours_val,
                    "non_billable": total_hours_val - billable_hours_val,
                },
                "requests_summary": {
                    "total": related_requests.count(),
                    "ongoing": related_requests.filter(status=Request.Status.ONGOING).count(),
                    "completed": related_requests.filter(status=Request.Status.COMPLETED).count(),
                },
                "activity_report_data": activity_report_data,
                "editing_log": editing_log,
                "month_options": month_options,
                "selected_month": selected_month,
                "selected_month_label": selected_month_label,
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        edit_id = request.GET.get("edit")
        editing_log = None
        form = None
        if edit_id:
            try:
                editing_log = self.get_queryset().get(pk=edit_id)
            except (EngineerActivityLog.DoesNotExist, ValueError):
                messages.error(request, "We could not find that activity log to edit.")
                return redirect("hub:activity-logs")
            form = self.form_class(engineer=request.user, instance=editing_log)
        context = self.get_context_data(form=form, editing_log=editing_log)
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        log_id = request.POST.get("log_id")
        instance = None
        if log_id:
            try:
                instance = self.get_queryset().get(pk=log_id)
            except (EngineerActivityLog.DoesNotExist, ValueError):
                messages.error(request, "Unable to update the selected activity log.")
                return redirect("hub:activity-logs")

        form = self.form_class(data=request.POST, engineer=request.user, instance=instance)
        if form.is_valid():
            activity_log = form.save(commit=False)
            activity_log.engineer = request.user
            if not activity_log.request_date:
                activity_log.request_date = timezone.now().date()
            activity_log.save()
            if instance:
                messages.success(request, "Activity log updated successfully.")
            else:
                messages.success(request, "Activity logged successfully.")
            return redirect("hub:activity-logs")
        context = self.get_context_data(form=form, editing_log=instance)
        return self.render_to_response(context)


class EngineerActivityLogDeleteView(EngineerRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        log = get_object_or_404(
            EngineerActivityLog.objects.filter(engineer=request.user),
            pk=pk,
        )
        log.delete()
        messages.success(request, "Activity log deleted successfully.")
        return redirect("hub:activity-logs")


class ReportExportView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, View):
    """Export operational or activity report data as CSV."""

    def get(self, request, *args, **kwargs):
        report_view = (request.GET.get("report_view") or "operational").lower()
        if report_view == "activity":
            return self._export_activity_logs()
        return self._export_operational_report()

    @staticmethod
    def _format_user(user: Optional[User]) -> str:
        if not user:
            return ""
        full_name = user.get_full_name()
        return full_name or user.username

    @staticmethod
    def _format_date(value):
        if not value:
            return ""
        return value.strftime("%Y-%m-%d")

    @staticmethod
    def _format_datetime(value):
        if not value:
            return ""
        localized = timezone.localtime(value)
        return localized.strftime("%Y-%m-%d %H:%M:%S")

    def _export_operational_report(self):
        filename = f"operational-report-{timezone.now().strftime('%Y%m%d-%H%M%S')}.csv"
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "Reference",
                "Requestor",
                "Account",
                "Engagement",
                "Product Category",
                "Priority",
                "Status",
                "Engineer",
                "Backup Engineer",
                "Start Date",
                "Due Date",
                "Created",
            ]
        )

        queryset = (
            Request.objects.select_related("requestor", "account", "engineer", "backup_engineer")
            .order_by("created_at")
        )
        status_labels = dict(Request.Status.choices)
        engagement_labels = dict(Request.Engagement.choices)
        for request_obj in queryset:
            writer.writerow(
                [
                    request_obj.reference_code,
                    self._format_user(getattr(request_obj, "requestor", None)),
                    request_obj.account.name if request_obj.account else "",
                    engagement_labels.get(request_obj.engagement_type, request_obj.engagement_type or ""),
                    request_obj.product_category or "",
                    request_obj.get_priority_display(),
                    status_labels.get(request_obj.status, request_obj.status or ""),
                    self._format_user(getattr(request_obj, "engineer", None)),
                    self._format_user(getattr(request_obj, "backup_engineer", None)),
                    self._format_date(request_obj.start_date),
                    self._format_date(request_obj.due_date),
                    self._format_datetime(request_obj.created_at),
                ]
            )

        return response

    def _export_activity_logs(self):
        filename = f"activity-logs-{timezone.now().strftime('%Y%m%d-%H%M%S')}.csv"
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "Request Date",
                "Engineer",
                "Account",
                "Request Reference",
                "Activity Type",
                "Details",
                "Hours",
                "Location",
                "Billable",
                "Created",
            ]
        )

        logs = (
            EngineerActivityLog.objects.select_related("engineer", "account", "request")
            .order_by("-request_date", "-created_at")
        )
        activity_labels = dict(EngineerActivityLog.ActivityType.choices)
        location_labels = dict(EngineerActivityLog.Location.choices)
        for log in logs:
            writer.writerow(
                [
                    self._format_date(log.request_date),
                    self._format_user(log.engineer),
                    log.account.name if log.account else "",
                    log.request.reference_code if log.request else "",
                    activity_labels.get(log.activity_type, log.activity_type or ""),
                    log.details or "",
                    f"{log.actual_hours}",
                    location_labels.get(log.location, log.location or ""),
                    "Yes" if log.is_billable else "No",
                    self._format_datetime(log.created_at),
                ]
            )

        return response


def summarize_request_changes(original, updated, changed_fields):
    choice_maps = {
        "priority": dict(Request.Priority.choices),
        "status": dict(Request.Status.choices),
        "engagement_type": dict(Request.Engagement.choices),
    }

    def fmt_user(u):
        return (u.get_full_name() or u.username) if u else None

    def fmt_date(v):
        return v.strftime("%b %d, %Y") if v else None

    def fmt_choice(field, v):
        return choice_maps[field].get(v, v) if v else None

    change_summaries = []
    for field in changed_fields:
        old = getattr(original, field, None)
        new = getattr(updated, field, None)
        if old == new:
            continue

        if field == "status":
            label = fmt_choice("status", new) or "Unknown"
            change_summaries.append(f"Marked as {label}")

        elif field == "engineer":
            if new:
                change_summaries.append(f"Assigned engineer: {fmt_user(new)}")
            else:
                change_summaries.append("Engineer unassigned")

        elif field == "backup_engineer":
            if new:
                change_summaries.append(f"Assigned backup engineer: {fmt_user(new)}")
            else:
                change_summaries.append("Backup engineer unassigned")

        elif field == "priority":
            change_summaries.append(f"Priority set to {fmt_choice('priority', new)}")

        elif field == "due_date":
            if new:
                change_summaries.append(f"Due date set to {fmt_date(new)}")
            else:
                change_summaries.append("Due date cleared")

        elif field == "end_date":
            if new:
                change_summaries.append(f"Closed on {fmt_date(new)}")
            else:
                change_summaries.append("End date cleared")

        elif field == "start_date":
            if new:
                change_summaries.append(f"Start date set to {fmt_date(new)}")

        elif field == "account":
            if new:
                change_summaries.append(f"Account changed to {new.name}")

        elif field == "engagement_type":
            change_summaries.append(f"Engagement type set to {fmt_choice('engagement_type', new)}")

        elif field == "product_category":
            if new:
                change_summaries.append(f"Product category set to {new}")

        elif field == "description":
            change_summaries.append("Description updated")

        elif field == "account_manager":
            if new:
                change_summaries.append(f"Requestor set to {new}")

    return "; ".join(change_summaries)


def notify_status_update(log, source_label):
    """Notify relevant stakeholders that a new status update was posted."""
    request_obj = log.request
    author = log.author
    author_name = author.get_full_name() or author.username

    recipients: dict[int, User] = {}
    if request_obj.engineer:
        recipients[request_obj.engineer.pk] = request_obj.engineer
    if request_obj.requestor:
        recipients[request_obj.requestor.pk] = request_obj.requestor

    for admin in User.objects.filter(role__in=ADMIN_PANEL_ROLES):
        recipients[admin.pk] = admin

    recipients.pop(author.pk, None)

    for user in recipients.values():
        Notification.objects.create(
            recipient=user,
            message=f"{author_name} posted an update on {request_obj.reference_code or 'a request'}.",
            related_request=request_obj,
            actor=author_name,
            source=source_label,
        )


def get_request_activity_log_context(request_obj: Request) -> dict:
    related_activity_logs = (
        request_obj.activity_logs.select_related("engineer", "account")
        .order_by("-request_date", "-created_at")
    )
    activity_summary = related_activity_logs.aggregate(
        total_hours=Sum("actual_hours"),
        billable_hours=Sum("actual_hours", filter=Q(is_billable=True)),
    )
    total_hours = activity_summary.get("total_hours") or Decimal("0")
    billable_hours = activity_summary.get("billable_hours") or Decimal("0")
    return {
        "related_activity_logs": related_activity_logs,
        "related_activity_summary": {
            "count": related_activity_logs.count(),
            "total_hours": total_hours,
            "billable_hours": billable_hours,
            "non_billable_hours": total_hours - billable_hours,
        },
    }


def create_change_status_log(request_obj: Request, actor_user: User, source_label: str, summary_text: str) -> None:
    """Persist a status log entry describing automatic updates."""
    if not summary_text:
        return

    actor_name = actor_user.get_full_name() or actor_user.username
    message = f"{actor_name} — {summary_text}"
    StatusLog.objects.create(
        request=request_obj,
        author=actor_user,
        message=message,
    )


def normalize_request_form_changed_fields(changed_fields):
    normalized = []
    for field in changed_fields:
        if field == "needed_by":
            normalized.append("start_date")
        elif field == "account_name":
            continue
        else:
            normalized.append(field)
    return normalized


def notify_account_manager_request_update(actor_user, original, updated, changed_fields, source_label):
    summary_text = summarize_request_changes(original, updated, changed_fields)
    if not summary_text:
        return

    create_change_status_log(updated, actor_user, source_label, summary_text)

    actor = actor_user.get_full_name() or actor_user.username
    recipients: dict[int, User] = {}

    for admin in User.objects.filter(role__in=ADMIN_PANEL_ROLES):
        recipients[admin.pk] = admin
    if updated.requestor:
        recipients[updated.requestor.pk] = updated.requestor
    if updated.engineer:
        recipients[updated.engineer.pk] = updated.engineer
    if updated.backup_engineer:
        recipients[updated.backup_engineer.pk] = updated.backup_engineer

    recipients.pop(actor_user.pk, None)

    for recipient in recipients.values():
        Notification.objects.create(
            recipient=recipient,
            message=f"{actor} updated {updated.reference_code}: {summary_text}",
            related_request=updated,
            actor=actor,
            source=source_label,
        )


def notify_engineer_assignment_email(
    request_obj: Request,
    *,
    actor_user: User,
    request=None,
    previous_engineer_id: int | None = None,
    previous_backup_id: int | None = None,
) -> AssignmentEmailResult:
    """Email the assigned engineer or backup when they receive a request."""

    recipients: list[str] = []
    missing_email_roles: list[str] = []

    if request_obj.engineer_id and request_obj.engineer_id != previous_engineer_id:
        email = (request_obj.engineer.email or "").strip()
        if email:
            recipients.append(email)
        else:
            missing_email_roles.append("engineer")

    if request_obj.backup_engineer_id and request_obj.backup_engineer_id != previous_backup_id:
        email = (request_obj.backup_engineer.email or "").strip()
        if email:
            recipients.append(email)
        else:
            missing_email_roles.append("backup engineer")

    if not recipients:
        if missing_email_roles:
            logger.warning(
                "Assignment email skipped: assigned recipient missing email for %s (%s)",
                request_obj.reference_code,
                ", ".join(missing_email_roles),
            )
            return AssignmentEmailResult("missing_assignee_email")
        logger.info("Assignment email skipped: no engineer/backup recipient for %s", request_obj.reference_code)
        return AssignmentEmailResult("no_new_assignee")

    actor_name = actor_user.get_full_name() or actor_user.username or "Request Hub"
    due_display = request_obj.due_date.strftime("%b %d, %Y") if request_obj.due_date else "Not set"
    detail_url = ""

    if request:
        try:
            detail_url = request.build_absolute_uri(request_obj.get_absolute_url())
        except Exception:
            detail_url = ""

    subject = f"[Request Hub] New assignment - {request_obj.reference_code or 'Request'}"
    body_lines = [
        "You have been assigned a request in Request Hub.",
        "",
        f"Reference: {request_obj.reference_code or 'Request'}",
        f"Account: {request_obj.account.name if request_obj.account else 'Not set'}",
        f"Engagement: {request_obj.get_engagement_type_display()}",
        f"Priority: {request_obj.get_priority_display()}",
        f"Due date: {due_display}",
        f"Assigned by: {actor_name}",
    ]

    description = (request_obj.description or "").strip()
    if description:
        body_lines.extend(["", "Description:", description])

    if detail_url:
        body_lines.extend(["", f"View request: {detail_url}"])

    use_acs = bool(settings.ACS_EMAIL_CONNECTION_STRING and settings.ACS_EMAIL_SENDER)

    if not use_acs:
        logger.warning("ACS email not configured; set ACS_EMAIL_CONNECTION_STRING and ACS_EMAIL_SENDER")
        return AssignmentEmailResult("missing_acs_config", tuple(recipients))

    try:
        from azure.communication.email import EmailClient

        client = EmailClient.from_connection_string(settings.ACS_EMAIL_CONNECTION_STRING)
        message = {
            "senderAddress": settings.ACS_EMAIL_SENDER,
            "recipients": {"to": [{"address": addr} for addr in recipients]},
            "content": {
                "subject": subject,
                "plainText": "\n".join(body_lines),
            },
        }
        poller = client.begin_send(message)
        poller.result()
        return AssignmentEmailResult("sent", tuple(recipients))
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "ACS email send failed for %s to %s",
            request_obj.reference_code,
            ", ".join(recipients),
            exc_info=exc,
        )
        return AssignmentEmailResult("delivery_failed", tuple(recipients))


def notify_engineer_assignment_notification(
    request_obj: Request,
    *,
    actor_user: User,
    previous_engineer_id: int | None = None,
    previous_backup_id: int | None = None,
) -> None:
    """Create in-app notifications when an engineer or backup is newly assigned."""

    recipients: dict[int, User] = {}
    if request_obj.engineer and request_obj.engineer_id != previous_engineer_id:
        recipients[request_obj.engineer_id] = request_obj.engineer
    if request_obj.backup_engineer and request_obj.backup_engineer_id != previous_backup_id:
        recipients[request_obj.backup_engineer_id] = request_obj.backup_engineer

    recipients.pop(actor_user.pk, None)

    if not recipients:
        return

    actor_name = actor_user.get_full_name() or actor_user.username or "Request Hub"
    account_name = request_obj.account.name if request_obj.account else "Account"
    engagement = request_obj.get_engagement_type_display()
    for recipient in recipients.values():
        Notification.objects.create(
            recipient=recipient,
            message=f"You were assigned to {request_obj.reference_code} · {account_name} ({engagement}).",
            related_request=request_obj,
            actor=actor_name,
            source="Assignment",
        )


def clear_engineer_outlook_lock_on_reassignment(
    request_obj: Request,
    *,
    previous_engineer_id: int | None,
) -> int:
    """Clear engineer Outlook lock records when the primary engineer changes.
    (Disabled per requirement: the new engineer shouldn't need to acknowledge again if already done)"""
    return 0

def _admin_sort_account_manager_key(request_obj):
    manager_name = ""
    manager_user = getattr(request_obj, "requestor", None)
    if manager_user:
        manager_name = (manager_user.get_full_name() or manager_user.username or "").strip()
    elif request_obj.account_manager:
        manager_name = request_obj.account_manager.strip()
    normalized = manager_name.lower()
    has_name = 0 if normalized else 1
    return (has_name, normalized)


def _send_sqr_new_submission_email(
    submission: SqrSubmission,
    *,
    http_request=None,
) -> None:
    """Send an email to the assigned PM-ESG reviewer when a new SQR is submitted."""
    reviewer = submission.pm_esg_reviewer
    if not reviewer:
        logger.warning(
            "SQR new-submission email skipped: no PM-ESG reviewer assigned (SQR %s)",
            submission.reference_code,
        )
        return

    reviewer_email = (getattr(reviewer, "email", "") or "").strip()
    if not reviewer_email:
        # Fall back to username if it looks like an email address
        username = (getattr(reviewer, "username", "") or "").strip()
        if "@" in username:
            reviewer_email = username
    if not reviewer_email:
        logger.warning(
            "SQR new-submission email skipped: reviewer %s has no email (SQR %s)",
            getattr(reviewer, "username", "?"),
            submission.reference_code,
        )
        return

    use_acs = bool(settings.ACS_EMAIL_CONNECTION_STRING and settings.ACS_EMAIL_SENDER)
    if not use_acs:
        logger.warning(
            "ACS email not configured; SQR new-submission email skipped for %s",
            submission.reference_code,
        )
        return

    reviewer_name = reviewer.get_full_name() or reviewer.username
    engineer_name = submission.engineer.get_full_name() or submission.engineer.username
    sqr_path = reverse("hub:sqr")
    if http_request:
        try:
            sqr_url = http_request.build_absolute_uri(sqr_path)
        except Exception:
            sqr_url = sqr_path
    else:
        sqr_url = sqr_path

    customer_company = (submission.customer_company or "").strip()
    submitted_date = (
        submission.created_at.astimezone(MANILA_TZ).strftime("%B %d, %Y  %I:%M %p")
        if submission.created_at
        else "—"
    )
    divider = "─" * 56

    body_lines = [
        f"Hi {reviewer_name},",
        "",
        f"A new SQR submission has been created and assigned to you for review.",
        f"Please review it at your earliest convenience.",
        "",
        divider,
        "  SUBMISSION DETAILS",
        divider,
        f"  Reference:      {submission.reference_code}",
        f"  Submitted By:   {engineer_name}",
        f"  Customer:       {submission.customer_name}",
    ]
    if customer_company:
        body_lines.append(f"  Company/Group:  {customer_company}")
    body_lines += [
        f"  Project Title:  {submission.project_title}",
        f"  Scope of Svc:   {submission.project_details or '—'}",
        f"  Date Submitted: {submitted_date}",
        "",
        divider,
        "  NEXT STEPS",
        divider,
        "  1. Log in to Request Hub and open the SQR Review page.",
        "  2. Review the submission details and quotation.",
        "  3. Approve or request revision with your comments.",
        "",
        f"  View SQR Tracker:  {sqr_url}",
        "",
        divider,
        "This is an automated notification from Request Hub.",
        "Please do not reply directly to this message.",
    ]

    subject = f"[Request Hub] New SQR Submitted — {submission.reference_code} ({submission.customer_name})"
    plain_text = "\n".join(body_lines)

    try:
        from azure.communication.email import EmailClient

        client = EmailClient.from_connection_string(settings.ACS_EMAIL_CONNECTION_STRING)
        message = {
            "senderAddress": settings.ACS_EMAIL_SENDER,
            "recipients": {"to": [{"address": reviewer_email}]},
            "content": {
                "subject": subject,
                "plainText": plain_text,
            },
        }
        poller = client.begin_send(message)
        poller.result()
        logger.info(
            "SQR new-submission email sent to %s for %s",
            reviewer_email,
            submission.reference_code,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "SQR new-submission email send failed for %s to %s",
            submission.reference_code,
            reviewer_email,
            exc_info=exc,
        )


def _send_sqr_for_revision_email(
    submission: SqrSubmission,
    reviewer_name: str,
    *,
    http_request=None,
) -> None:
    """Send an email to the SQR creator when the status is set to For Revision."""
    engineer = submission.engineer
    engineer_email = (getattr(engineer, "email", "") or "").strip()
    if not engineer_email:
        logger.warning(
            "SQR for-revision email skipped: engineer %s has no email (SQR %s)",
            getattr(engineer, "username", "?"),
            submission.reference_code,
        )
        return

    use_acs = bool(settings.ACS_EMAIL_CONNECTION_STRING and settings.ACS_EMAIL_SENDER)
    if not use_acs:
        logger.warning(
            "ACS email not configured; SQR for-revision email skipped for %s",
            submission.reference_code,
        )
        return

    engineer_name = engineer.get_full_name() or engineer.username
    sqr_edit_path = reverse("hub:sqr")
    if http_request:
        try:
            sqr_url = http_request.build_absolute_uri(sqr_edit_path)
        except Exception:
            sqr_url = sqr_edit_path
    else:
        sqr_url = sqr_edit_path

    review_notes = (submission.review_notes or "").strip()
    customer_company = (submission.customer_company or "").strip()
    reviewed_date = (
        submission.updated_at.strftime("%B %d, %Y  %I:%M %p")
        if submission.updated_at
        else "—"
    )
    divider = "─" * 56

    body_lines = [
        f"Hi {engineer_name},",
        "",
        "Your SQR submission has been reviewed and marked For Revision.",
        "Please address the reviewer's comments and resubmit at your earliest convenience.",
        "",
        divider,
        "  SQR DETAILS",
        divider,
        f"  Reference:      {submission.reference_code}",
        f"  Customer:       {submission.customer_name}",
    ]
    if customer_company:
        body_lines.append(f"  Company/Group:  {customer_company}")
    body_lines += [
        f"  Project Title:  {submission.project_title}",
        f"  Reviewed By:    {reviewer_name}",
        f"  Date Reviewed:  {reviewed_date}",
        "",
        divider,
        "  REVIEWER COMMENTS",
        divider,
        (
            review_notes
            if review_notes
            else "No specific comments were provided. Please follow up with the reviewer for details."
        ),
        "",
        divider,
        "  NEXT STEPS",
        divider,
        "  1. Log in to Request Hub and navigate to your SQR submissions.",
        "  2. Carefully review the comments above.",
        "  3. Update your SQR submission accordingly.",
        "  4. Resubmit for review once the changes are complete.",
        "",
        f"  View SQR Tracker:  {sqr_url}",
        "",
        divider,
        "If you have questions, please reach out to your reviewer directly.",
        "",
        "This is an automated notification from Request Hub.",
        "ESG Request Hub  |  ESGRequestHub@phildata.com",
    ]

    subject = f"[Request Hub] {submission.reference_code} \u2014 Action Required: For Revision"
    plain_text = "\n".join(body_lines)

    try:
        from azure.communication.email import EmailClient

        client = EmailClient.from_connection_string(settings.ACS_EMAIL_CONNECTION_STRING)
        message = {
            "senderAddress": settings.ACS_EMAIL_SENDER,
            "recipients": {
                "to": [{"address": engineer_email}],
            },
            "content": {
                "subject": subject,
                "plainText": plain_text,
            },
        }
        poller = client.begin_send(message)
        poller.result()
        logger.info(
            "SQR for-revision email sent to %s for %s",
            engineer_email,
            submission.reference_code,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "SQR for-revision email send failed for %s to %s",
            submission.reference_code,
            engineer_email,
            exc_info=exc,
        )


def _send_sqr_approved_email(
    submission: SqrSubmission,
    reviewer_name: str,
    *,
    http_request=None,
) -> None:
    """Send an email to the SQR creator when the status is set to Approved."""
    engineer = submission.engineer
    engineer_email = (getattr(engineer, "email", "") or "").strip()
    if not engineer_email:
        logger.warning(
            "SQR approved email skipped: engineer %s has no email (SQR %s)",
            getattr(engineer, "username", "?"),
            submission.reference_code,
        )
        return

    use_acs = bool(settings.ACS_EMAIL_CONNECTION_STRING and settings.ACS_EMAIL_SENDER)
    if not use_acs:
        logger.warning(
            "ACS email not configured; SQR approved email skipped for %s",
            submission.reference_code,
        )
        return

    engineer_name = (engineer.get_full_name() or engineer.username or "").strip()
    subject = _format_sqr_approval_subject(submission)
    body_lines = _build_sqr_approval_body_lines(submission, engineer_name)
    plain_text = "\n".join(body_lines)
    html_text = _build_sqr_approval_html(submission, engineer_name)

    try:
        from azure.communication.email import EmailClient

        client = EmailClient.from_connection_string(settings.ACS_EMAIL_CONNECTION_STRING)
        message = {
            "senderAddress": settings.ACS_EMAIL_SENDER,
            "recipients": {
                "to": [
                    {"address": engineer_email},
                    {"address": "ESGRequestHub@phildata.com"},
                ]
            },
            "content": {
                "subject": subject,
                "plainText": plain_text,
                "html": html_text,
            },
        }
        poller = client.begin_send(message)
        poller.result()
        logger.info(
            "SQR approved email sent to %s for %s",
            engineer_email,
            submission.reference_code,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "SQR approved email send failed for %s to %s",
            submission.reference_code,
            engineer_email,
            exc_info=exc,
        )


def _get_sqr_approval_email_context(submission: SqrSubmission) -> dict:
    engineer = submission.engineer
    engineer_email = (getattr(engineer, "email", "") or "").strip()
    engineer_name = (engineer.get_full_name() or engineer.username or "").strip() if engineer else ""
    recipients = [email for email in (engineer_email, "ESGRequestHub@phildata.com") if email]
    logo_attachment = _get_sqr_logo_attachment()
    attachments = [logo_attachment] if logo_attachment else []
    return {
        "submission": submission,
        "subject": _format_sqr_approval_subject(submission),
        "recipients": recipients,
        "plain_text": "\n".join(_build_sqr_approval_body_lines(submission, engineer_name)),
        "html_preview": _build_sqr_approval_html(submission, engineer_name),
        "html_draft": _build_sqr_approval_draft_html(submission, engineer_name),
        "attachments": attachments,
    }


def _get_graph_access_token() -> tuple[str, str]:
    if not (settings.PHILDATA_TENANT_ID and settings.PHILDATA_CLIENT_ID and settings.PHILDATA_CLIENT_SECRET):
        return "", "Microsoft Graph draft creation is not configured. Set PHILDATA_TENANT_ID, PHILDATA_CLIENT_ID, and PHILDATA_CLIENT_SECRET."

    app = msal.ConfidentialClientApplication(
        client_id=settings.PHILDATA_CLIENT_ID,
        client_credential=settings.PHILDATA_CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{settings.PHILDATA_TENANT_ID}",
    )
    token_result = app.acquire_token_for_client(scopes=settings.PHILDATA_GRAPH_SCOPE)
    access_token = token_result.get("access_token")
    if access_token:
        return access_token, ""
    error_description = token_result.get("error_description") or token_result.get("error") or "Unable to acquire Microsoft Graph access token."
    return "", error_description


def _extract_graph_error_message(response: requests.Response) -> str:
    try:
        error_payload = response.json()
        return error_payload.get("error", {}).get("message") or response.text
    except ValueError:
        return response.text


def _create_sqr_approval_graph_draft_json(context: dict, sender_email: str, access_token: str) -> tuple[str, str]:
    graph_url = f"https://graph.microsoft.com/v1.0/users/{quote(sender_email, safe='')}/messages"
    recipients = context.get("recipients") or []
    payload = {
        "subject": context["subject"],
        "body": {
            "contentType": "HTML",
            "content": context.get("html_draft") or context["html_preview"],
        },
        "toRecipients": [
            {"emailAddress": {"address": email}}
            for email in recipients[:1]
            if email
        ],
        "ccRecipients": [
            {"emailAddress": {"address": email}}
            for email in recipients[1:]
            if email
        ],
    }
    attachments = context.get("attachments") or []
    if attachments:
        payload["attachments"] = attachments

    try:
        response = requests.post(
            graph_url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
    except requests.RequestException as exc:
        return "", f"Unable to contact Microsoft Graph: {exc}"

    if response.status_code not in (200, 201):
        message = _extract_graph_error_message(response)
        return "", f"Microsoft Graph JSON draft creation failed ({response.status_code}): {message}"

    draft = response.json()
    return draft.get("webLink") or "", ""


def _create_sqr_approval_graph_draft_mime(context: dict, sender_email: str, access_token: str) -> tuple[str, str]:
    graph_url = f"https://graph.microsoft.com/v1.0/users/{quote(sender_email, safe='')}/messages"
    mime_payload = _build_sqr_approval_mime_message(context, sender_email)
    try:
        response = requests.post(
            graph_url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "text/plain",
            },
            data=mime_payload,
            timeout=30,
        )
    except requests.RequestException as exc:
        return "", f"Unable to contact Microsoft Graph: {exc}"

    if response.status_code not in (200, 201):
        message = _extract_graph_error_message(response)
        return "", f"Microsoft Graph MIME draft creation failed ({response.status_code}): {message}"

    draft = response.json()
    return draft.get("webLink") or "", ""


def _create_sqr_approval_graph_draft(context: dict, sender_email: str) -> tuple[str, str]:
    sender_email = (sender_email or "").strip()
    if not sender_email:
        return "", "The signed-in user does not have an email address configured for Outlook draft creation."

    access_token, token_error = _get_graph_access_token()
    if token_error:
        return "", token_error

    web_link, json_error = _create_sqr_approval_graph_draft_json(context, sender_email, access_token)
    if web_link:
        return web_link, ""

    web_link, mime_error = _create_sqr_approval_graph_draft_mime(context, sender_email, access_token)
    if web_link:
        return web_link, ""

    if json_error and mime_error:
        return "", f"{json_error} Fallback attempt also failed: {mime_error}"
    return "", json_error or mime_error or "Microsoft Graph draft creation failed."


def _build_sqr_approval_mailto_url(context: dict) -> str:
    recipients = context.get("recipients") or []
    to_recipient = recipients[0] if recipients else ""
    cc_recipients = ";".join(recipients[1:])
    mailto_parts = [
        f"mailto:{quote(to_recipient)}",
        f"subject={quote(context['subject'])}",
        f"body={quote(context['plain_text'])}",
    ]
    if cc_recipients:
        mailto_parts.insert(1, f"cc={quote(cc_recipients)}")
    return "?".join([mailto_parts[0], "&".join(mailto_parts[1:])])


def _admin_sort_engineer_key(request_obj):
    engineer = getattr(request_obj, "engineer", None)
    if engineer:
        engineer_name = (engineer.get_full_name() or engineer.username or "").strip().lower()
        return (0, engineer_name)
    return (1, "")


def _admin_sort_backup_engineer_key(request_obj):
    backup = getattr(request_obj, "backup_engineer", None)
    if backup:
        name = (backup.get_full_name() or backup.username or "").strip().lower()
        return (0, name)
    return (1, "")


def _admin_sort_date_key(value):
    if value is None:
        return (1, date.max)
    return (0, value)


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "hub/dashboard.html"

    @staticmethod
    def _build_request_report_data(requests: list[Request]) -> dict:
        status_labels = dict(Request.Status.choices)
        priority_labels = dict(Request.Priority.choices)
        engagement_labels = dict(Request.Engagement.choices)

        status_counts: dict[str, int] = {}
        priority_counts: dict[str, int] = {}
        engagement_counts: dict[str, int] = {}
        monthly_counts: dict[str, int] = {}

        for request_obj in requests:
            status_key = request_obj.status or ""
            priority_key = request_obj.priority or ""
            engagement_key = request_obj.engagement_type or ""

            if status_key:
                status_counts[status_key] = status_counts.get(status_key, 0) + 1
            if priority_key:
                priority_counts[priority_key] = priority_counts.get(priority_key, 0) + 1
            if engagement_key:
                engagement_counts[engagement_key] = engagement_counts.get(engagement_key, 0) + 1

            month_source = request_obj.created_at
            month_key = month_source.strftime("%b %Y")
            monthly_counts[month_key] = monthly_counts.get(month_key, 0) + 1

        return {
            "request_status": {
                "labels": [status_labels.get(key, key) for key in status_counts.keys()],
                "values": list(status_counts.values()),
            },
            "request_priority": {
                "labels": [priority_labels.get(key, key) for key in priority_counts.keys()],
                "values": list(priority_counts.values()),
            },
            "request_engagement": {
                "labels": [engagement_labels.get(key, key) for key in engagement_counts.keys()],
                "values": list(engagement_counts.values()),
            },
            "request_monthly": {
                "labels": list(monthly_counts.keys()),
                "values": list(monthly_counts.values()),
            },
        }

    @staticmethod
    def _build_engineer_report_data(requests: list[Request], user) -> dict:
        from django.db.models import Min
        assigned_requests = [req for req in requests if req.engineer_id == user.id]

        # 1. Count of requests assigned by requestor
        requestor_counts = {}
        for req in assigned_requests:
            name = "Unknown"
            if req.requestor:
                name = req.requestor.get_full_name() or req.requestor.username or "Unknown"
            requestor_counts[name] = requestor_counts.get(name, 0) + 1

        # 2. Avg Ack Response Time
        request_ids = [req.pk for req in assigned_requests if req.pk]
        ack_map = {}
        if request_ids:
            ack_rows = (
                RequestCommunication.objects.filter(
                    request_id__in=request_ids,
                    user__role__in=ENGINEER_ACCESS_ROLES,
                    channel__in=[RequestCommunication.Channel.OUTLOOK, RequestCommunication.Channel.TEAMS],
                )
                .values("request_id")
                .annotate(first_ack=Min("created_at"))
            )
            ack_map = {row["request_id"]: row["first_ack"] for row in ack_rows}

        ack_seconds_list = []
        for req in assigned_requests:
            ack_time = ack_map.get(req.pk)
            if ack_time:
                start = req.updated_at if req.updated_at and req.updated_at > req.created_at else req.created_at
                delta = (ack_time - start).total_seconds()
                if delta >= 0:
                    # simplistic fallback without working hour logic to maintain performance on db scale
                    # and decouple from the helper method inside `_annotate...`
                    ack_seconds_list.append(delta)

        if ack_seconds_list:
            avg_ack_seconds = sum(ack_seconds_list) / len(ack_seconds_list)
            hours = int(avg_ack_seconds // 3600)
            minutes = int((avg_ack_seconds % 3600) // 60)
            if hours > 0:
                avg_ack_time = f"{hours}h {minutes}m"
            else:
                avg_ack_time = f"{minutes}m"
        else:
            avg_ack_time = "N/A"

        # 3. Avg Resolution Time
        completed_reqs = [req for req in assigned_requests if req.status == Request.Status.COMPLETED]
        if completed_reqs:
            avg_resolution_days = sum(req.days_since_creation for req in completed_reqs) / len(completed_reqs)
            avg_resolution_time = f"{avg_resolution_days:.1f} days"
        else:
            avg_resolution_time = "N/A"

        # 4. Avg Requests per week
        if assigned_requests:
            from django.utils import timezone
            min_date = min(req.created_at for req in assigned_requests)
            days_span = (timezone.now() - min_date).days
            weeks = max(1, days_span / 7)
            avg_requests_per_week = f"{len(assigned_requests) / weeks:.1f}"
        else:
            avg_requests_per_week = "0"

        return {
            "requests_by_requestor": {
                "labels": list(requestor_counts.keys()),
                "values": list(requestor_counts.values()),
            },
            "avg_ack_time": avg_ack_time,
            "avg_resolution_time": avg_resolution_time,
            "avg_requests_per_week": avg_requests_per_week,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        if user.role in ADMIN_PANEL_ROLES:
            context["role"] = User.Roles.ADMIN
        elif user.role in ENGINEER_ACCESS_ROLES:
            context["role"] = User.Roles.ENGINEER
        else:
            context["role"] = user.role
        context["is_admin_ui"] = user.role in ADMIN_PANEL_ROLES
        context["is_requestor_role"] = user.role in REQUESTOR_ROLES
        context["is_pm_ess"] = user.role == PM_ESS_ROLE
        context["is_pm_esg"] = user.role == PM_ESG_ROLE
        context["is_requestor_ui"] = user.role in REQUESTOR_ROLES or user.role == PM_ESS_ROLE
        context["is_engineer_access_role"] = user.role in ENGINEER_ACCESS_ROLES
        context["can_create_request"] = user.role in REQUEST_CREATOR_ROLES
        context["notifications"] = user.notifications.filter(is_read=False)[:10]

        if user.role == PM_ESS_ROLE:
            form = kwargs.get("form")
            if form is None:
                form = RequestForm(actor_role=user.role, actor_user=user)
            context["form"] = form
            context["account_name_choices"] = form.account_name_suggestions
            metric_filter = self.request.GET.get("metric_filter") or ""
            metric_keys = {"ongoing", "completed"}
            if metric_filter not in metric_keys:
                metric_filter = ""

            request_tab = (self.request.GET.get("request_tab") or "all").strip().lower()
            if request_tab not in {"all", "mine"}:
                request_tab = "all"

            all_requests = list(
                Request.objects.filter(
                    Q(requestor__role=User.Roles.REQUESTOR_ESS) | Q(requestor=user)
                )
                .select_related("account", "engineer", "requestor")
                .order_by("-created_at")
            )

            my_requests = [req for req in all_requests if req.requestor_id == user.id]
            requests = my_requests if request_tab == "mine" else all_requests

            metrics = {
                "ongoing": sum(1 for req in requests if req.status == Request.Status.ONGOING),
                "completed": sum(1 for req in requests if req.status == Request.Status.COMPLETED),
            }

            filtered_requests = requests
            if metric_filter == "ongoing":
                filtered_requests = [req for req in requests if req.status == Request.Status.ONGOING]
            elif metric_filter == "completed":
                filtered_requests = [req for req in requests if req.status == Request.Status.COMPLETED]

            metric_links = {}
            for key in ["ongoing", "completed"]:
                params = self.request.GET.copy()
                if params.get("metric_filter") == key:
                    params.pop("metric_filter", None)
                else:
                    params["metric_filter"] = key
                encoded = params.urlencode()
                metric_links[key] = f"?{encoded}" if encoded else "?"

            request_tab_links = {}
            for key in ("all", "mine"):
                params = self.request.GET.copy()
                if key == "all":
                    params.pop("request_tab", None)
                else:
                    params["request_tab"] = "mine"
                encoded = params.urlencode()
                request_tab_links[key] = f"?{encoded}" if encoded else "?"

            context["requests"] = filtered_requests
            context["metrics"] = metrics
            context["metric_links"] = metric_links
            context["active_metric_filter"] = metric_filter
            context["pm_ess_request_tab"] = request_tab
            context["pm_ess_request_tab_links"] = request_tab_links
            context["form_has_errors"] = form.is_bound and bool(form.errors)
            context["request_report_summary"] = {
                "total": len(requests),
                "ongoing": metrics["ongoing"],
                "completed": metrics["completed"],
            }
            context["request_report_data"] = self._build_request_report_data(requests)
        elif user.role in REQUEST_CREATOR_ROLES and user.role not in ADMIN_PANEL_ROLES:
            form = kwargs.get("form")
            if form is None:
                form = RequestForm(actor_role=user.role, actor_user=user)
            context["form"] = form
            context["account_name_choices"] = form.account_name_suggestions
            metric_filter = self.request.GET.get("metric_filter") or ""
            metric_keys = {"ongoing", "completed"}
            if metric_filter not in metric_keys:
                metric_filter = ""

            requests = list(
                Request.objects.filter(requestor=user)
                .select_related("account", "engineer")
                .order_by("-created_at")
            )

            metrics = {
                "ongoing": sum(1 for req in requests if req.status == Request.Status.ONGOING),
                "completed": sum(1 for req in requests if req.status == Request.Status.COMPLETED),
            }

            filtered_requests = requests
            if metric_filter == "ongoing":
                filtered_requests = [req for req in requests if req.status == Request.Status.ONGOING]
            elif metric_filter == "completed":
                filtered_requests = [req for req in requests if req.status == Request.Status.COMPLETED]

            metric_links = {}
            for key in ["ongoing", "completed"]:
                params = self.request.GET.copy()
                if params.get("metric_filter") == key:
                    params.pop("metric_filter", None)
                else:
                    params["metric_filter"] = key
                encoded = params.urlencode()
                metric_links[key] = f"?{encoded}" if encoded else "?"

            request_tab = (self.request.GET.get("request_tab") or "mine").strip().lower()
            if request_tab not in {"all", "mine"}:
                request_tab = "mine"

            request_tab_links = {}
            for key in ("all", "mine"):
                params = self.request.GET.copy()
                if key == "all":
                    params.pop("request_tab", None)
                else:
                    params["request_tab"] = "mine"
                encoded = params.urlencode()
                request_tab_links[key] = f"?{encoded}" if encoded else "?"

            context["requests"] = filtered_requests
            context["metrics"] = metrics
            context["metric_links"] = metric_links
            context["active_metric_filter"] = metric_filter
            context["pm_ess_request_tab"] = request_tab
            context["pm_ess_request_tab_links"] = request_tab_links
            context["form_has_errors"] = form.is_bound and bool(form.errors)
            context["request_report_summary"] = {
                "total": len(requests),
                "ongoing": metrics["ongoing"],
                "completed": metrics["completed"],
            }
            context["request_report_data"] = self._build_request_report_data(requests)
        elif user.role in ENGINEER_ACCESS_ROLES:
            metric_filter = (self.request.GET.get("metric_filter") or "").strip()
            tab = self.request.GET.get("tab") or "assigned"
            valid_metrics = {"ongoing", "due_soon", "overdue", "completed"}
            if metric_filter not in valid_metrics:
                metric_filter = ""
            effective_metric_filter = metric_filter or "ongoing"

            if tab == "backup":
                requests = list(
                    Request.objects.filter(backup_engineer=user)
                    .select_related("account", "requestor")
                    .order_by("status", "due_date")
                )
            else:
                requests = list(
                    Request.objects.filter(engineer=user)
                    .select_related("account", "requestor")
                    .order_by("status", "due_date")
                )

            context["active_tab"] = tab
            if tab == "report":
                context["engineer_report_data"] = self._build_engineer_report_data(
                    requests=[req for req in Request.objects.filter(engineer=user).select_related("requestor")],
                    user=user
                )

            request_ids = [req.pk for req in requests]
            outlook_limited: set[int] = set()
            teams_limited: set[int] = set()
            if request_ids:
                communications = RequestCommunication.objects.filter(
                    request_id__in=request_ids,
                    user__role__in=ENGINEER_ACCESS_ROLES,
                ).only("request_id", "channel", "user_id")
                for comm in communications:
                    if comm.channel == RequestCommunication.Channel.OUTLOOK:
                        outlook_limited.add(comm.request_id)
                    elif comm.channel == RequestCommunication.Channel.TEAMS:
                        if getattr(comm, "user_id", None) == user.id:
                            teams_limited.add(comm.request_id)
            for req in requests:
                setattr(req, "outlook_limit_reached", req.pk in outlook_limited)
                setattr(req, "teams_limit_reached", req.pk in teams_limited)

            self._annotate_acknowledgement_status(requests)
            self._annotate_engineer_activity(requests)

            today = timezone.now().astimezone(MANILA_TZ).date()
            metrics = {
                "ongoing": sum(1 for req in requests if req.status == Request.Status.ONGOING),
                "due_soon": sum(
                    1
                    for req in requests
                    if req.status == Request.Status.ONGOING
                and req.engagement_type not in {
                    Request.Engagement.DEPLOYMENT, Request.Engagement.CERTIFICATION
                }
                    and req.due_date
                    and 0 <= (req.due_date - today).days <= 3
                ),
                "overdue": sum(
                    1
                    for req in requests
                    if req.status == Request.Status.ONGOING and req.due_date and req.due_date < today
                ),
                "completed": sum(1 for req in requests if req.status == Request.Status.COMPLETED),
            }

            filtered_requests = requests
            if effective_metric_filter == "ongoing":
                filtered_requests = [req for req in requests if req.status == Request.Status.ONGOING]
            elif effective_metric_filter == "due_soon":
                filtered_requests = [
                    req
                    for req in requests
                    if req.status == Request.Status.ONGOING
                    and req.engagement_type not in {
                        Request.Engagement.DEPLOYMENT, Request.Engagement.CERTIFICATION
                    }
                    and req.due_date
                    and 0 <= (req.due_date - today).days <= 3
                ]
            elif effective_metric_filter == "overdue":
                filtered_requests = [
                    req
                    for req in requests
                    if req.status == Request.Status.ONGOING and req.due_date and req.due_date < today
                ]
            elif effective_metric_filter == "completed":
                filtered_requests = [req for req in requests if req.status == Request.Status.COMPLETED]
                filtered_requests = sorted(filtered_requests, key=lambda req: req.created_at, reverse=True)

            metric_links = {}
            for key in ("ongoing", "due_soon", "overdue", "completed"):
                params = self.request.GET.copy()
                is_active_key = metric_filter == key or (key == "ongoing" and metric_filter == "")
                if is_active_key:
                    params.pop("metric_filter", None)
                else:
                    params["metric_filter"] = key
                encoded = params.urlencode()
                metric_links[key] = f"?{encoded}" if encoded else "?"

            context["requests"] = filtered_requests
            context["metrics"] = metrics
            context["metric_links"] = metric_links
            context["active_metric_filter"] = effective_metric_filter
            context["active_tab"] = tab
            context["new_ticket_count"] = user.notifications.filter(
                is_read=False,
                source__icontains="assignment",
            ).count()
        else:
            metric_keys = [
                "open",
                "overdue",
                "due_soon",
                "completed",
                "new_this_week",
            ]
            metric_filter = self.request.GET.get("metric_filter")
            if not metric_filter or metric_filter not in metric_keys:
                metric_filter = "open"

            pm_esg_tab = ""
            if user.role == PM_ESG_ROLE:
                pm_esg_tab = (self.request.GET.get("pm_esg_tab") or "all").strip().lower()
                if pm_esg_tab not in {"all", "assigned", "my_requests"}:
                    pm_esg_tab = "all"

            queryset = Request.objects.select_related("account", "engineer", "backup_engineer", "requestor")
            if pm_esg_tab == "assigned":
                queryset = queryset.filter(engineer=user)
            elif pm_esg_tab == "my_requests":
                queryset = queryset.filter(requestor=user)
            filter_form = AdminRequestFilterForm(self.request.GET or None)
            filtered_queryset = filter_form.filter_queryset(queryset)
            requests = filter_form.filter_sequence(filtered_queryset)
            filters_active = filter_form.has_active_filters()
            show_filters = self.request.GET.get("show_filters") == "1"

            if not isinstance(requests, list):
                requests = list(requests)

            sort_map = {
                "reference_code": lambda req: (req.reference_code or "").lower(),
                "account": lambda req: (req.account.name.lower() if req.account else ""),
                "account_manager": _admin_sort_account_manager_key,
                "engineer": _admin_sort_engineer_key,
                "backup_engineer": _admin_sort_backup_engineer_key,
                "engagement": lambda req: req.engagement_type or "",
                "product_category": lambda req: req.product_category or "",
                "status": lambda req: req.status or "",
                "created": lambda req: req.created_at,
                "end_date": lambda req: _admin_sort_date_key(req.end_date),
                "days": lambda req: req.days_since_creation,
                "due": lambda req: _admin_sort_date_key(req.due_date),
            }

            default_sort = "created"
            default_direction = "desc"
            sort_key = self.request.GET.get("sort", default_sort)
            direction = self.request.GET.get("direction", default_direction)
            if sort_key not in sort_map:
                sort_key = default_sort
            if direction not in {"asc", "desc"}:
                direction = default_direction
            reverse = direction == "desc"

            key_fn = sort_map[sort_key]
            requests.sort(key=key_fn, reverse=reverse)

            null_field_map = {"end_date": "end_date", "due": "due_date"}
            if sort_key in null_field_map:
                field_name = null_field_map[sort_key]
                ordered = [req for req in requests if getattr(req, field_name) is not None]
                ordered.extend(req for req in requests if getattr(req, field_name) is None)
                requests = ordered

            self._annotate_acknowledgement_status(requests)

            today = timezone.now().astimezone(MANILA_TZ).date()
            all_requests = list(requests)
            filtered_requests = self._filter_requests_by_metric(all_requests, metric_filter, today)

            overdue_count = sum(1 for req in all_requests if req.admin_days_overdue)
            status_counts = {
                "all": len(all_requests),
                "ongoing": sum(1 for req in all_requests if req.status == Request.Status.ONGOING),
                "completed": sum(1 for req in all_requests if req.status == Request.Status.COMPLETED),
            }

            columns = list(sort_map.keys())
            next_directions = {}
            for column in columns:
                if column == sort_key:
                    next_directions[column] = "asc" if direction == "desc" else "desc"
                else:
                    next_directions[column] = "asc"

            base_params = self.request.GET.copy()
            base_params.pop("sort", None)
            base_params.pop("direction", None)
            sort_links = {}
            for column in columns:
                params = base_params.copy()
                params["sort"] = column
                params["direction"] = next_directions[column]
                encoded = params.urlencode()
                sort_links[column] = f"?{encoded}" if encoded else "?"

            toggle_params = self.request.GET.copy()
            if show_filters:
                toggle_params.pop("show_filters", None)
            else:
                toggle_params["show_filters"] = "1"
            toggle_encoded = toggle_params.urlencode()
            filter_toggle_link = f"?{toggle_encoded}" if toggle_encoded else ("?" if show_filters else "?show_filters=1")

            metric_links = {}
            metric_buckets = {}
            for key in metric_keys:
                params = self.request.GET.copy()
                if params.get("metric_filter") == key:
                    params.pop("metric_filter", None)
                else:
                    params["metric_filter"] = key
                encoded_params = params.urlencode()
                metric_links[key] = f"?{encoded_params}" if encoded_params else "?"
                metric_buckets[key] = self._filter_requests_by_metric(all_requests, key, today)

            context["requests"] = filtered_requests
            context["overdue_count"] = overdue_count
            context["status_counts"] = status_counts
            context["new_ticket_count"] = user.notifications.filter(
                is_read=False,
                source__icontains="new request",
            ).count()
            context["current_sort"] = sort_key
            context["current_direction"] = direction
            context["sort_next"] = next_directions
            context["sort_links"] = sort_links
            context["filter_form"] = filter_form
            context["filters_active"] = filters_active or (metric_filter and metric_filter != "open")
            context["show_filters"] = show_filters
            context["filter_toggle_link"] = filter_toggle_link
            context["metrics"] = {key: len(metric_buckets[key]) for key in metric_keys}
            context["metric_links"] = metric_links
            context["active_metric_filter"] = metric_filter

            pm_esg_tab_links = {}
            if user.role == PM_ESG_ROLE:
                for key in ("all", "assigned", "my_requests"):
                    params = self.request.GET.copy()
                    if key == "all":
                        params.pop("pm_esg_tab", None)
                    else:
                        params["pm_esg_tab"] = key
                    encoded = params.urlencode()
                    pm_esg_tab_links[key] = f"?{encoded}" if encoded else "?"
            context["pm_esg_tab"] = pm_esg_tab
            context["pm_esg_tab_links"] = pm_esg_tab_links

        if context.get("can_create_request") and "form" not in context:
            form = kwargs.get("form")
            if form is None:
                form = RequestForm(actor_role=user.role, actor_user=user)
            context["form"] = form
            context["account_name_choices"] = form.account_name_suggestions
            context["form_has_errors"] = form.is_bound and bool(form.errors)
        return context

    @staticmethod
    def _filter_requests_by_metric(requests: list[Request], metric_filter: str, today: date) -> list[Request]:
        if not metric_filter:
            return requests

        if metric_filter == "open":
            return [req for req in requests if req.status == Request.Status.ONGOING]
        if metric_filter == "overdue":
            return [
                req
                for req in requests
                if req.admin_days_overdue
            ]
        if metric_filter == "due_soon":
            return [
                req
                for req in requests
                if req.status == Request.Status.ONGOING
                and req.engagement_type not in {
                    Request.Engagement.DEPLOYMENT, Request.Engagement.CERTIFICATION
                }
                and req.due_date
                and 0 <= (req.due_date - today).days <= 3
            ]
        if metric_filter == "completed":
            return [
                req
                for req in requests
                if req.status == Request.Status.COMPLETED
            ]
        if metric_filter == "new_this_week":
            return [
                req
                for req in requests
                if (today - req.created_at.date()).days <= 7
            ]
        return requests

    @staticmethod
    def _format_duration(delta: timedelta) -> str:
        total_seconds = int(delta.total_seconds())
        if total_seconds <= 0:
            return "<1 minute"
        minutes = total_seconds // 60
        if minutes < 1:
            return "<1 minute"
        hours, rem_minutes = divmod(minutes, 60)
        parts = []
        if hours:
            parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
        if rem_minutes:
            parts.append(f"{rem_minutes} minute{'s' if rem_minutes != 1 else ''}")
        if not parts:
            return "<1 minute"
        return " ".join(parts)

    def _annotate_acknowledgement_status(self, requests: list[Request]) -> None:
        if not requests:
            return
        request_ids = [req.pk for req in requests if req.pk]
        if not request_ids:
            return

        def working_seconds_between(start_dt: datetime, end_dt: datetime) -> int:
            """Return working seconds between two datetimes within 8:30–18:30, Monday–Friday (Manila)."""
            if end_dt <= start_dt:
                return 0

            # Normalize to Manila timezone to align with working hours definition.
            start = timezone.localtime(start_dt, MANILA_TZ)
            end = timezone.localtime(end_dt, MANILA_TZ)

            work_start_hour = 8
            work_start_minute = 30
            work_end_hour = 18
            work_end_minute = 30

            def next_work_start(dt: datetime) -> datetime:
                dt = dt.astimezone(MANILA_TZ)
                # Move to today's start if before work start.
                candidate = dt.replace(hour=work_start_hour, minute=work_start_minute, second=0, microsecond=0)
                if dt.weekday() < 5 and dt < candidate:
                    return candidate
                # Otherwise move to next weekday 8:30.
                offset_days = 1
                next_day = dt + timedelta(days=offset_days)
                while next_day.weekday() >= 5:  # skip weekends
                    offset_days += 1
                    next_day = dt + timedelta(days=offset_days)
                return next_day.replace(hour=work_start_hour, minute=work_start_minute, second=0, microsecond=0, tzinfo=MANILA_TZ)

            total_seconds = 0
            current = start

            # If starting outside working hours, jump to next window.
            work_start_today = current.replace(hour=work_start_hour, minute=work_start_minute, second=0, microsecond=0)
            work_end_today = current.replace(hour=work_end_hour, minute=work_end_minute, second=0, microsecond=0)
            if current.weekday() >= 5 or current >= work_end_today or current < work_start_today:
                current = next_work_start(current)

            while current < end:
                if current.weekday() >= 5:
                    current = next_work_start(current)
                    continue

                work_start = current.replace(hour=work_start_hour, minute=work_start_minute, second=0, microsecond=0)
                work_end = current.replace(hour=work_end_hour, minute=work_end_minute, second=0, microsecond=0)

                if current < work_start:
                    current = work_start

                if current >= work_end:
                    current = next_work_start(current)
                    continue

                slice_end = min(work_end, end)
                total_seconds += int((slice_end - current).total_seconds())
                current = slice_end

                if current >= work_end:
                    current = next_work_start(current)

            return total_seconds

        ack_rows = (
            RequestCommunication.objects.filter(
                request_id__in=request_ids,
                user__role__in=[*ENGINEER_ACCESS_ROLES, User.Roles.ADMIN, PM_ESG_ROLE],
                channel__in=[
                    RequestCommunication.Channel.OUTLOOK,
                    RequestCommunication.Channel.TEAMS,
                ],
            )
            .values("request_id")
            .annotate(first_ack=Min("created_at"))
        )
        ack_map = {row["request_id"]: row["first_ack"] for row in ack_rows}

        now = timezone.now()
        amber_threshold_seconds = int(timedelta(minutes=45).total_seconds())
        sla_threshold_seconds = int(timedelta(hours=1).total_seconds())

        for req in requests:
            if not req.engineer:
                req.ack_sla_status = ""
                req.ack_sla_tooltip = "Awaiting engineer assignment; acknowledgement SLA starts after assignment."
                req.ack_start_iso = ""
                req.ack_first_iso = ""
                req.is_acknowledged = False
                continue

            start_anchor = req.created_at
            if req.updated_at and req.updated_at > req.created_at:
                start_anchor = req.updated_at

            ack_time = ack_map.get(req.pk)
            status = ""
            tooltip = ""
            if ack_time:
                delta_seconds = working_seconds_between(start_anchor, ack_time)
                # If delta is negative, the start_anchor (likely updated_at) is later than ack_time.
                # This happens on reassignment. We treat it as acknowledged.
                if delta_seconds <= sla_threshold_seconds:
                    status = "green"
                    if delta_seconds <= 0:
                        # Fallback calculation using created_at to give a roughly meaningful duration
                        fallback_delta = working_seconds_between(req.created_at, ack_time)
                        duration_str = self._format_duration(timedelta(seconds=max(0, fallback_delta)))
                        tooltip = f"Acknowledged previously ({duration_str} from creation)"
                    else:
                        tooltip = f"Acknowledged within SLA ({self._format_duration(timedelta(seconds=delta_seconds))})"
                else:
                    status = "red"
                    tooltip = f"Acknowledged after 1-hour SLA ({self._format_duration(timedelta(seconds=delta_seconds))})"
            else:
                age_seconds = working_seconds_between(start_anchor, now)
                if age_seconds <= 0:
                    tooltip = "Awaiting acknowledgement (outside working hours)"
                elif age_seconds >= sla_threshold_seconds:
                    status = "red"
                    tooltip = f"No acknowledgement after {self._format_duration(timedelta(seconds=age_seconds))} (working hours)"
                elif age_seconds >= amber_threshold_seconds:
                    status = "amber"
                    tooltip = f"Awaiting acknowledgement ({self._format_duration(timedelta(seconds=age_seconds))} elapsed in working hours)"
                else:
                    tooltip = f"Awaiting acknowledgement ({self._format_duration(timedelta(seconds=age_seconds))} elapsed in working hours)"

            req.ack_sla_status = status
            req.ack_sla_tooltip = tooltip or "Acknowledgement status unavailable"
            req.ack_start_iso = start_anchor.isoformat() if start_anchor else ""
            req.ack_first_iso = ack_time.isoformat() if ack_time else ""
            req.is_acknowledged = bool(ack_time)

    def _annotate_engineer_activity(self, requests: list[Request]) -> None:
        """Attach status-log and recent-change hints used by the engineer dashboard."""
        if not requests:
            return
        request_ids = [req.pk for req in requests if req.pk]
        log_counts: dict[int, int] = {}
        latest_logs: dict[int, StatusLog] = {}
        if request_ids:
            from django.db.models import Count

            count_rows = (
                StatusLog.objects.filter(request_id__in=request_ids)
                .values("request_id")
                .annotate(total=Count("id"))
            )
            log_counts = {row["request_id"]: row["total"] for row in count_rows}

            latest_qs = (
                StatusLog.objects.filter(request_id__in=request_ids)
                .select_related("author")
                .order_by("request_id", "-created_at")
            )
            seen: set[int] = set()
            for log in latest_qs:
                if log.request_id in seen:
                    continue
                seen.add(log.request_id)
                latest_logs[log.request_id] = log

        for req in requests:
            count = log_counts.get(req.pk, 0)
            latest = latest_logs.get(req.pk)
            req.status_log_count = count
            if latest:
                author_name = (
                    latest.author.get_full_name()
                    or latest.author.username
                    or "Unknown"
                )
                message = (latest.message or "").strip()
                excerpt = message[:140] + ("…" if len(message) > 140 else "")
                created_local = timezone.localtime(latest.created_at, MANILA_TZ)
                req.latest_status_log_tooltip = (
                    f"Latest update by {author_name} on "
                    f"{created_local.strftime('%b %d, %Y %I:%M %p')}"
                    + (f" — {excerpt}" if excerpt else "")
                )
            else:
                req.latest_status_log_tooltip = ""

            # Consider the request "changed" when updated_at is clearly after created_at.
            has_changes = bool(
                req.updated_at
                and req.created_at
                and (req.updated_at - req.created_at).total_seconds() > 60
            )
            req.has_recent_changes = has_changes
            if has_changes:
                updated_local = timezone.localtime(req.updated_at, MANILA_TZ)
                req.recent_change_tooltip = (
                    f"Request last updated on "
                    f"{updated_local.strftime('%b %d, %Y %I:%M %p')}"
                )
            else:
                req.recent_change_tooltip = ""

    def post(self, request, *args, **kwargs):
        if request.user.role not in REQUEST_CREATOR_ROLES:
            return redirect("hub:dashboard")
        form = RequestForm(request.POST, actor_role=request.user.role, actor_user=request.user)
        if form.is_valid():
            req = form.save(commit=False)
            req.requestor = request.user
            full_name = request.user.get_full_name().strip()
            req.account_manager = full_name or request.user.username
            req._actor_user = request.user
            req._actor_source = "Dashboard · New Request"
            if request.user.role in ADMIN_PANEL_ROLES:
                req._allow_capacity_override = True
            req.save()
            notify_engineer_assignment_notification(req, actor_user=request.user)
            assignment_email_result = notify_engineer_assignment_email(
                req,
                actor_user=request.user,
                request=request,
            )
            self._notify_admins_new_request(req)
            messages.success(request, "Request submitted", extra_tags="request-success")
            flash_assignment_email_feedback(
                request,
                assignment_email_result,
                action_label="submitted",
                notify_on_no_assignee=True,
            )
            return redirect("hub:dashboard")
        context = self.get_context_data(form=form)
        return self.render_to_response(context)

    @staticmethod
    def _notify_admins_new_request(request_obj):
        actor = request_obj.requestor.get_full_name() or request_obj.requestor.username
        priority_label = request_obj.get_priority_display()
        category_label = request_obj.get_product_category_display()
        due_display = request_obj.due_date.strftime("%b %d, %Y") if request_obj.due_date else "No due date"
        message = (
            f"New {priority_label} ticket {request_obj.reference_code} for {request_obj.account.name} "
            f"({category_label}) submitted by {actor}. Due {due_display}."
        )
        for admin in User.objects.filter(role__in=ADMIN_PANEL_ROLES):
            if admin.pk == request_obj.requestor_id:
                continue
            Notification.objects.create(
                recipient=admin,
                message=message,
                related_request=request_obj,
                actor=actor,
                source="Dashboard · New Request",
            )


def _refresh_sqr_request_account_map(form):
    import json
    form.fields["linked_request"].widget.attrs["data-account-map"] = json.dumps({
        str(req.pk): req.account.name
        for req in form.fields["linked_request"].queryset
        if getattr(req, "account", None) and req.account.name
    })


def _get_sqr_export_columns():
    def _date(d):
        return d.strftime("%Y-%m-%d") if d else ""

    def _name(u):
        return (u.get_full_name() or u.username) if u else ""

    return [
        ("SQR Date", lambda s: _date(s.created_at.date())),
        ("SQR ID", lambda s: s.reference_code or ""),
        ("Account Name", lambda s: s.customer_name or ""),
        ("Service Description", lambda s: s.project_title or ""),
        ("Scope of Services", lambda s: s.project_details or ""),
        ("RQ ID", lambda s: s.linked_request.reference_code if s.linked_request else ""),
        ("Group Name", lambda s: s.customer_company or ""),
        ("Account Manager", lambda s: s.customer_contact or ""),
        ("Requester Name", lambda s: _name(s.engineer)),
        ("Approver Name", lambda s: _name(s.pm_esg_reviewer)),
        ("SQR Doc. Ref. Link", lambda s: s.sqr_folder_link or ""),
        ("SSE Man-hrs", lambda s: float(s.sse_manhrs) if s.sse_manhrs is not None else ""),
        ("SSE Amount", lambda s: float(s.sse_amount) if s.sse_amount is not None else ""),
        ("PM Man-hrs", lambda s: float(s.pm_manhrs) if s.pm_manhrs is not None else ""),
        ("PM Amount", lambda s: float(s.pm_amount) if s.pm_amount is not None else ""),
        ("Managed Support Svc. Amt.", lambda s: float(s.managed_support_amount) if s.managed_support_amount is not None else ""),
        ("Discount Rate (%)", lambda s: float(s.discount_rate) if s.discount_rate is not None else ""),
        ("Discount Amount", lambda s: float(s.computed_discount_amount) if s.computed_discount_amount is not None else ""),
        ("Total Price", lambda s: float(s.computed_total_price) if s.computed_total_price is not None else ""),
        ("SQR Status", lambda s: s.get_status_display() if hasattr(s, "get_status_display") else s.status),
        ("Approval Date", lambda s: _date(s.reviewed_at.date()) if s.reviewed_at else ""),
        ("Validity Due Date", lambda s: _date(s.validity_due_date)),
        ("Proposal Status", lambda s: s.get_proposal_status_display() if hasattr(s, "get_proposal_status_display") else (s.proposal_status or "")),
        ("PO / PNL Date", lambda s: _date(s.po_pnl_date)),
        ("Assigned PM", lambda s: _name(s.assigned_pm) or _name(s.pm_esg_reviewer)),
        ("Assigned SSE", lambda s: _name(s.assigned_sse)),
        ("Start Date", lambda s: _date(s.delivery_start_date)),
        ("Target Finish Date", lambda s: _date(s.delivery_target_finish_date)),
        ("Overall Status", lambda s: s.overall_status or ""),
        ("Health Status", lambda s: s.delivery_health or ""),
        ("Overall Progress %", lambda s: s.delivery_progress if s.delivery_progress is not None else ""),
        ("Key Updates / Risks", lambda s: s.key_updates_risks or ""),
        ("Actual Finish Date", lambda s: _date(s.delivery_actual_finish_date)),
        ("Completion Signed Date", lambda s: _date(s.delivery_completion_signed_date)),
        ("Completion Warranty End Date", lambda s: _date(s.computed_post_svc_warranty_end_date)),
        ("Maintenance Start Date", lambda s: _date(s.computed_support_start_date)),
        ("Maintenance End Date", lambda s: _date(s.computed_managed_support_end_date)),
        ("PO Remarks", lambda s: s.revenue_remarks or ""),
        (
            "Billing Type",
            lambda s: (
                {"internal": "Internal", "invoiced": "Invoiced", "unbilled": "Unbilled"}.get(
                    s.revenue_source, s.revenue_source or ""
                )
            ),
        ),
        ("Billed Date", lambda s: _date(s.revenue_date)),
        ("Billing Reference", lambda s: (s.revenue_reference_no or "").upper()),
        (
            "Billing Status",
            lambda s: (
                s.get_revenue_status_display()
                if s.revenue_status
                else ("Billed" if s.revenue_date else "Not Yet Billed")
            ),
        ),
        ("Billing Remarks", lambda s: s.revenue_overview or ""),
    ]


def _get_sqr_esg_export_columns():
    def _name(u):
        return (u.get_full_name() or u.username) if u else ""

    def _date_short(d):
        return f"{d.month}/{d.day}/{str(d.year)[-2:]}" if d else ""

    def _date_medium_safe(d):
        return f"{d.day}-{d.strftime('%b')}-{str(d.year)[-2:]}" if d else ""

    def _project_name(s):
        if s.customer_name and s.project_title:
            return f"{s.customer_name} - {s.project_title}"
        return s.customer_name or s.project_title or ""

    def _sse1(s):
        return _name(s.assigned_sse) or _name(s.engineer)

    def _pm(s):
        return _name(s.assigned_pm) or _name(s.pm_esg_reviewer)

    def _project_flag(s):
        return "Yes" if s.assigned_pm or s.pm_esg_reviewer or s.delivery_start_date else "No"

    def _status(s):
        if s.overall_status == "on_hold":
            return "ON HOLD"
        if s.overall_status == "completed" or s.delivery_health == "completed":
            return "COMPLETED"
        if s.delivery_start_date or s.overall_status == "in_progress" or s.proposal_status == "closed_won":
            return "ONGOING"
        if s.status == "for_revision":
            return "FOR REVISION"
        return "NOT STARTED"

    def _start_finish_dates(s):
        if not (s.delivery_start_date or s.delivery_target_finish_date or s.delivery_actual_finish_date):
            return ""
        start = _date_short(s.delivery_start_date) or "TBD"
        finish = _date_short(s.delivery_actual_finish_date) or _date_short(s.delivery_target_finish_date) or "TBD"
        return f"{start}-{finish}"

    def _money(value):
        return float(value) if value is not None else ""

    def _total_revenue(s):
        return _money(s.computed_total_price if s.computed_total_price is not None else s.quotation_total_price)

    def _revenue_source(s):
        if s.revenue_source == "internal":
            return "Internal"
        if s.revenue_source == "invoiced":
            return "Invoiced"
        if s.revenue_source == "unbilled":
            return "Unbilled"
        return ""
    return [
        ("Account Name", lambda s: s.customer_name or ""),
        ("Service Description", lambda s: s.project_title or ""),
        ("Project Name", _project_name),
        ("SSE1", _sse1),
        ("SSE2", lambda s: ""),
        ("SSE3", lambda s: ""),
        ("Project?", _project_flag),
        ("PM", _pm),
        ("Status", _status),
        ("Start and Finish Dates", _start_finish_dates),
        ("% Complete", lambda s: s.delivery_progress if s.delivery_progress is not None else ""),
        ("Key Updates", lambda s: s.key_updates_risks or ""),
        ("Active Risks/Issues", lambda s: s.key_updates_risks or ""),
        ("Scope Status", lambda s: s.project_details or ""),
        ("Schedule Status", lambda s: _date_short(s.delivery_target_finish_date)),
        ("Budget Status", lambda s: ""),
        ("Folder", lambda s: s.sqr_folder_link or ""),
        ("Quoted SSE hrs", lambda s: float(s.sse_manhrs) if s.sse_manhrs is not None else ""),
        ("Billed SSE hrs", lambda s: ""),
        ("Quoted PM hrs", lambda s: float(s.effective_pm_manhrs) if s.effective_pm_manhrs is not None else ""),
        ("Billed PM hrs", lambda s: ""),
        ("Project Management", lambda s: _money(s.effective_pm_amount)),
        ("Deployment", lambda s: _money(s.sse_amount)),
        ("On-Call", lambda s: _money(s.managed_support_amount)),
        ("Total Revenue", _total_revenue),
        ("Revenue?", lambda s: "Yes" if s.revenue_date or s.revenue_declaration == "declared" else "No"),
        ("SI Type", _revenue_source),
        ("SI Reference", lambda s: (s.revenue_reference_no or "").upper()),
        ("SI Date", lambda s: _date_medium_safe(s.revenue_date)),
        ("Remarks", lambda s: s.revenue_remarks or s.remarks or ""),
        ("Action", lambda s: "History"),
    ]


def _normalize_import_text(value) -> str:
    return " ".join(
        str(value or "")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("_", " ")
        .split()
    ).casefold()


def _build_choice_import_map(choices) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for value, label in choices:
        mapping[_normalize_import_text(value)] = value
        mapping[_normalize_import_text(label)] = value
    return mapping


def _make_sqr_edit_form(user):
    """Return a SqrSubmissionForm for the edit modal scoped to engineer-accessible requests."""
    used_req_ids = SqrSubmission.objects.exclude(linked_request_id=None).values_list("linked_request_id", flat=True)
    form = SqrSubmissionForm(auto_id="edit_%s")
    form.fields["linked_request"].queryset = (
        Request.objects.filter(Q(engineer=user) | Q(backup_engineer=user))
        .exclude(id__in=used_req_ids)
        .select_related("account").only("id", "reference_code", "account__name").order_by("-id")
    )
    _refresh_sqr_request_account_map(form)
    return form


class SqrListView(LoginRequiredMixin, TemplateView):
    template_name = "hub/sqr.html"

    VALID_TABS = {"proposal", "delivery", "revenue-report", "reports"}
    PM_ONLY_TABS = {"delivery", "revenue-report", "reports"}

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if request.user.role not in SQR_ACCESS_ROLES:
            messages.error(request, "You are not allowed to access SQR.")
            return redirect("hub:dashboard")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = SqrSubmission.objects.select_related(
            "engineer",
            "pm_esg_reviewer",
            "reviewed_by",
            "assigned_pm",
            "assigned_sse",
            "linked_request",
        ).order_by("-created_at", "-pk")
        if self.request.user.role in ENGINEER_ACCESS_ROLES:
            return queryset.filter(engineer=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        can_create = user.role in ENGINEER_ACCESS_ROLES
        can_review = user.role in ADMIN_PANEL_ROLES
        is_pm = user.role in ADMIN_PANEL_ROLES
        is_admin = user.role == User.Roles.ADMIN

        def decimal_to_float(value):
            return float(value or Decimal("0"))

        def submission_discounted_value(submission):
            return (
                submission.computed_total_price
                or submission.discounted_price
                or submission.quotation_total_price
                or Decimal("0")
            )

        def submission_gross_value(submission):
            return (
                submission.computed_gross_total
                or submission.quotation_total_price
                or submission.computed_total_price
                or submission.discounted_price
                or Decimal("0")
            )

        def compact_currency(value: Decimal) -> str:
            amount = decimal_to_float(value)
            abs_amount = abs(amount)
            if abs_amount >= 1_000_000:
                return f"₱{amount / 1_000_000:.2f}M"
            if abs_amount >= 1_000:
                return f"₱{amount / 1_000:.1f}K"
            return f"₱{amount:,.2f}"

        active_tab = (self.request.GET.get("tab") or "proposal").strip().lower()
        if active_tab not in self.VALID_TABS:
            active_tab = "proposal"
        if active_tab in self.PM_ONLY_TABS and not is_pm:
            active_tab = "proposal"

        form = kwargs.get("form")
        if can_create and form is None:
            form = SqrSubmissionForm()
            # Scope RQ ID dropdown to requests assigned to this engineer only,
            # excluding requests that already have an SQR submission.
            _used_req_ids = SqrSubmission.objects.exclude(linked_request_id=None).values_list("linked_request_id", flat=True)
            form.fields["linked_request"].queryset = (
                Request.objects.filter(engineer=user)
                .exclude(id__in=_used_req_ids)
                .select_related("account").only("id", "reference_code", "account__name").order_by("-id")
            )
            _refresh_sqr_request_account_map(form)

        all_submissions = list(self.get_queryset())

        # ── Proposal Stage ──────────────────────────────────────────────────
        proposal_counts = {
            "total": len(all_submissions),
            "processing": sum(1 for s in all_submissions if s.status == SqrSubmission.Status.FOR_PROCESSING),
            "for_revision": sum(1 for s in all_submissions if s.status == SqrSubmission.Status.FOR_REVISION),
            "approved": sum(1 for s in all_submissions if s.status == SqrSubmission.Status.APPROVED),
            "submitted_pending": sum(
                1 for s in all_submissions
                if s.proposal_status == SqrSubmission.ProposalStatus.SUBMITTED_PENDING
            ),
            "negotiation": sum(
                1 for s in all_submissions
                if s.proposal_status == SqrSubmission.ProposalStatus.NEGOTIATION_REVIEW
            ),
            "closed_won": sum(
                1 for s in all_submissions
                if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_WON
            ),
            "closed_lost": sum(
                1 for s in all_submissions
                if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_LOST
            ),
            "closed_canceled": sum(
                1 for s in all_submissions
                if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_CANCELED
            ),
        }

        # ── Service Delivery Stage ───────────────────────────────────────────
        delivery_submissions = (
            [s for s in all_submissions if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_WON]
            if is_pm else []
        )
        delivery_health_counts = {}
        for choice_val, choice_label in SqrSubmission.DeliveryHealth.choices:
            delivery_health_counts[choice_val] = sum(
                1 for s in delivery_submissions if s.delivery_health == choice_val
            )

        # ── Revenue & Report Stage ──────────────────────────────────────────
        won_submissions = (
            [s for s in all_submissions if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_WON]
            if is_pm else []
        )
        lost_submissions = (
            [s for s in all_submissions if s.proposal_status == SqrSubmission.ProposalStatus.CLOSED_LOST]
            if is_pm else []
        )
        in_negotiation = (
            [s for s in all_submissions if s.proposal_status == SqrSubmission.ProposalStatus.NEGOTIATION_REVIEW]
            if is_pm else []
        )
        total_with_deal_status = sum(
            1 for s in all_submissions if s.proposal_status
        ) if is_pm else 0
        won_total_price = sum(submission_gross_value(s) for s in won_submissions)
        won_discounted_total = sum(submission_discounted_value(s) for s in won_submissions)
        lost_total_price = sum(submission_discounted_value(s) for s in lost_submissions)
        win_rate = (
            round(len(won_submissions) / total_with_deal_status * 100, 1)
            if total_with_deal_status > 0 else 0
        )
        top_won = sorted(won_submissions, key=submission_discounted_value, reverse=True)[:5]

        revenue_data = {
            "won_count": len(won_submissions),
            "lost_count": len(lost_submissions),
            "negotiation_count": len(in_negotiation),
            "won_total_price": won_total_price,
            "won_discounted_total": won_discounted_total,
            "lost_total_price": lost_total_price,
            "win_rate": win_rate,
            "top_won": top_won,
            "lost_submissions": lost_submissions,
            "revenue_submissions": [s for s in all_submissions if s.status == SqrSubmission.Status.APPROVED] if is_pm else [],
        }

        sqr_reports_summary = None
        sqr_reports_status_chart = {"labels": [], "totals": []}
        sqr_reports_deal_chart = {"labels": [], "totals": []}
        sqr_reports_delivery_chart = {"labels": [], "totals": []}
        sqr_reports_group_chart = {"labels": [], "totals": []}
        sqr_reports_scope_chart = {"labels": [], "totals": []}
        sqr_reports_funnel_chart = {"labels": [], "totals": []}
        sqr_reports_monthly_chart = {"labels": [], "totals": []}
        sqr_reports_revenue_chart = {"labels": [], "totals": []}
        sqr_reports_quick_overview_chart = {"labels": [], "totals": []}
        sqr_reports_top_accounts = []
        sqr_reports_top_won = []
        sqr_reports_pipeline = {}
        sqr_reports_focus_items = []

        if is_pm:
            approved_submissions = [
                submission for submission in all_submissions
                if submission.status == SqrSubmission.Status.APPROVED
            ]
            billed_submissions = [
                submission for submission in approved_submissions if submission.revenue_date
            ]
            for_billing_submissions = [
                submission for submission in approved_submissions if not submission.revenue_date
            ]
            at_risk_submissions = [
                submission for submission in delivery_submissions
                if submission.delivery_health == SqrSubmission.DeliveryHealth.AT_RISK
            ]
            off_track_submissions = [
                submission for submission in delivery_submissions
                if submission.delivery_health == SqrSubmission.DeliveryHealth.OFF_TRACK
            ]
            on_track_submissions = [
                submission for submission in delivery_submissions
                if submission.delivery_health == SqrSubmission.DeliveryHealth.ON_TRACK
            ]
            pending_approval_count = proposal_counts["processing"] + proposal_counts["for_revision"]

            status_rows = [
                ("For Processing", proposal_counts["processing"]),
                ("For Revision", proposal_counts["for_revision"]),
                ("Approved", proposal_counts["approved"]),
            ]
            sqr_reports_status_chart = {
                "labels": [label for label, total in status_rows if total],
                "totals": [total for _, total in status_rows if total],
            }

            deal_rows = [
                ("Submitted", proposal_counts["submitted_pending"]),
                ("On Review", proposal_counts["negotiation"]),
                ("Closed-Won", proposal_counts["closed_won"]),
                ("Closed-Lost", proposal_counts["closed_lost"]),
                ("Closed-Canceled", proposal_counts.get("closed_canceled", 0)),
            ]
            sqr_reports_deal_chart = {
                "labels": [label for label, total in deal_rows if total],
                "totals": [total for _, total in deal_rows if total],
            }

            sqr_reports_delivery_chart = {
                "labels": [label for value, label in SqrSubmission.DeliveryHealth.choices if delivery_health_counts.get(value)],
                "totals": [delivery_health_counts.get(value, 0) for value, _ in SqrSubmission.DeliveryHealth.choices if delivery_health_counts.get(value)],
            }

            group_counter = Counter(
                ((submission.customer_company or "Unspecified").strip() or "Unspecified")
                for submission in all_submissions
            )
            sorted_group_rows = sorted(group_counter.items(), key=lambda item: (-item[1], item[0]))[:8]
            sqr_reports_group_chart = {
                "labels": [label for label, _ in sorted_group_rows],
                "totals": [total for _, total in sorted_group_rows],
            }

            scope_counter = Counter(
                ((submission.project_details or "Unspecified").strip() or "Unspecified")
                for submission in all_submissions
            )
            sorted_scope_rows = sorted(scope_counter.items(), key=lambda item: (-item[1], item[0]))[:8]
            sqr_reports_scope_chart = {
                "labels": [label for label, _ in sorted_scope_rows],
                "totals": [total for _, total in sorted_scope_rows],
            }

            sqr_reports_funnel_chart = {
                "labels": ["Total SQR", "Approved", "Closed Won", "Billed"],
                "totals": [
                    proposal_counts["total"],
                    proposal_counts["approved"],
                    len(won_submissions),
                    len(billed_submissions),
                ],
            }
            sqr_reports_quick_overview_chart = {
                "labels": ["Pending", "Approved", "Won", "Billed"],
                "totals": [
                    pending_approval_count,
                    proposal_counts["approved"],
                    len(won_submissions),
                    len(billed_submissions),
                ],
            }

            monthly_counter = Counter()
            monthly_labels = {}
            for submission in all_submissions:
                localized_created = timezone.localtime(submission.created_at, MANILA_TZ)
                month_key = localized_created.strftime("%Y-%m")
                monthly_counter[month_key] += 1
                monthly_labels[month_key] = localized_created.strftime("%b %Y")

            sorted_month_keys = sorted(monthly_counter.keys())
            sqr_reports_monthly_chart = {
                "labels": [monthly_labels[key] for key in sorted_month_keys],
                "totals": [monthly_counter[key] for key in sorted_month_keys],
            }

            total_managed_support = sum(
                submission.managed_support_amount or Decimal("0")
                for submission in all_submissions
            )
            revenue_rows = [
                ("Won Gross", won_total_price),
                ("Won Discounted", won_discounted_total),
                ("Lost Value", lost_total_price),
                ("Managed Support", total_managed_support),
            ]
            sqr_reports_revenue_chart = {
                "labels": [label for label, value in revenue_rows if value and value > 0],
                "totals": [decimal_to_float(value) for _, value in revenue_rows if value and value > 0],
            }

            account_totals = defaultdict(lambda: {"count": 0, "value": Decimal("0")})
            for submission in won_submissions:
                account_name = (submission.customer_name or "Unspecified Account").strip() or "Unspecified Account"
                account_totals[account_name]["count"] += 1
                account_totals[account_name]["value"] += submission_discounted_value(submission)

            sqr_reports_top_accounts = [
                {
                    "name": account_name,
                    "count": metrics["count"],
                    "value": metrics["value"],
                }
                for account_name, metrics in sorted(
                    account_totals.items(),
                    key=lambda item: (-item[1]["value"], -item[1]["count"], item[0]),
                )[:6]
            ]

            sqr_reports_top_won = [
                {
                    "reference_code": submission.reference_code,
                    "customer_name": submission.customer_name,
                    "group_name": submission.customer_company or "—",
                    "discounted_price": submission_discounted_value(submission),
                    "status": submission.get_delivery_health_display() if submission.delivery_health else "Awaiting delivery setup",
                }
                for submission in top_won
            ]

            approval_lead_days = [
                max(
                    (timezone.localtime(submission.reviewed_at, MANILA_TZ).date() - timezone.localtime(submission.created_at, MANILA_TZ).date()).days,
                    0,
                )
                for submission in approved_submissions
                if submission.reviewed_at
            ]
            avg_approval_days = (
                round(sum(approval_lead_days) / len(approval_lead_days), 1)
                if approval_lead_days else 0
            )
            avg_won_value = (
                won_discounted_total / len(won_submissions)
                if won_submissions else Decimal("0")
            )

            approval_rate = (
                round((proposal_counts["approved"] / proposal_counts["total"]) * 100, 1)
                if proposal_counts["total"] else 0
            )
            billed_conversion_rate = (
                round((len(billed_submissions) / proposal_counts["approved"]) * 100, 1)
                if proposal_counts["approved"] else 0
            )
            win_from_approved_rate = (
                round((len(won_submissions) / proposal_counts["approved"]) * 100, 1)
                if proposal_counts["approved"] else 0
            )
            sqr_reports_summary = {
                "total_submissions": proposal_counts["total"],
                "approval_rate": approval_rate,
                "win_rate": win_rate,
                "active_delivery": len(
                    [
                        submission for submission in delivery_submissions
                        if submission.delivery_health in {
                            SqrSubmission.DeliveryHealth.ON_TRACK,
                            SqrSubmission.DeliveryHealth.OFF_TRACK,
                            SqrSubmission.DeliveryHealth.AT_RISK,
                        }
                    ]
                ),
                "won_discounted_total": won_discounted_total,
                "won_discounted_total_label": compact_currency(won_discounted_total),
                "lost_total_price": lost_total_price,
                "lost_total_price_label": compact_currency(lost_total_price),
                "managed_support_total": total_managed_support,
                "managed_support_total_label": compact_currency(total_managed_support),
                "billed_count": len(billed_submissions),
                "for_billing_count": len(for_billing_submissions),
                "at_risk_count": len(at_risk_submissions),
                "off_track_count": len(off_track_submissions),
                "on_track_count": len(on_track_submissions),
                "pending_approval_count": pending_approval_count,
                "avg_won_value_label": compact_currency(avg_won_value),
                "avg_approval_days": avg_approval_days,
                "billed_conversion_rate": billed_conversion_rate,
                "win_from_approved_rate": win_from_approved_rate,
            }

            sqr_reports_pipeline = {
                "pending_approval_count": pending_approval_count,
                "approved_count": proposal_counts["approved"],
                "negotiation_count": proposal_counts["negotiation"],
                "won_count": len(won_submissions),
                "lost_count": len(lost_submissions),
                "billed_count": len(billed_submissions),
                "for_billing_count": len(for_billing_submissions),
                "at_risk_count": len(at_risk_submissions),
                "off_track_count": len(off_track_submissions),
                "on_track_count": len(on_track_submissions),
                "avg_won_value_label": compact_currency(avg_won_value),
                "avg_approval_days": avg_approval_days,
            }

            sqr_reports_focus_items = [
                {
                    "title": "Pending internal action",
                    "value": pending_approval_count,
                    "description": "SQRs still waiting for processing or revision closure.",
                    "icon": "bi-hourglass-split",
                    "tone": "warning",
                },
                {
                    "title": "Revenue still pending",
                    "value": len(for_billing_submissions),
                    "description": "Approved quotations not yet billed in the revenue stage.",
                    "icon": "bi-receipt-cutoff",
                    "tone": "primary",
                },
                {
                    "title": "Delivery watchlist",
                    "value": len(at_risk_submissions) + len(off_track_submissions),
                    "description": "Won projects marked at risk or off track and needing attention.",
                    "icon": "bi-exclamation-diamond",
                    "tone": "danger",
                },
                {
                    "title": "Average won deal size",
                    "value": compact_currency(avg_won_value),
                    "description": "Typical discounted value of each successfully won quotation.",
                    "icon": "bi-cash-stack",
                    "tone": "success",
                },
            ]

        my_assigned_sqr_count = sum(
            1 for submission in all_submissions if submission.pm_esg_reviewer_id == user.id
        )
        # Approvers (PM-ESG) default to "My Assigned SQR"; Total still shows the full list.
        default_my_assigned_filter = user.role == PM_ESG_ROLE

        context.update(
            {
                "can_create_sqr": can_create,
                "can_review_sqr": can_review,
                "is_pm": is_pm,
                "is_admin": is_admin,
                "is_pm_esg": user.role == PM_ESG_ROLE,
                "my_assigned_sqr_count": my_assigned_sqr_count,
                "default_my_assigned_filter": default_my_assigned_filter,
                "sqr_current_user_id": user.pk,
                "sqr_form": form,
                "sqr_import_form": SqrImportForm() if is_admin else None,
                "sqr_edit_form": _make_sqr_edit_form(user) if can_create else None,
                "sqr_form_has_errors": bool(form and form.is_bound and form.errors),
                "active_sqr_tab": active_tab,
                "proposal_submissions": all_submissions,
                "proposal_counts": proposal_counts,
                "delivery_submissions": delivery_submissions,
                "delivery_health_counts": delivery_health_counts,
                "revenue_data": revenue_data,
                "sqr_reports_summary": sqr_reports_summary,
                "sqr_reports_status_chart": sqr_reports_status_chart,
                "sqr_reports_deal_chart": sqr_reports_deal_chart,
                "sqr_reports_delivery_chart": sqr_reports_delivery_chart,
                "sqr_reports_group_chart": sqr_reports_group_chart,
                "sqr_reports_scope_chart": sqr_reports_scope_chart,
                "sqr_reports_funnel_chart": sqr_reports_funnel_chart,
                "sqr_reports_monthly_chart": sqr_reports_monthly_chart,
                "sqr_reports_revenue_chart": sqr_reports_revenue_chart,
                "sqr_reports_quick_overview_chart": sqr_reports_quick_overview_chart,
                "sqr_reports_top_accounts": sqr_reports_top_accounts,
                "sqr_reports_top_won": sqr_reports_top_won,
                "sqr_reports_pipeline": sqr_reports_pipeline,
                "sqr_reports_focus_items": sqr_reports_focus_items,
                "sqr_pm_users_json": json.dumps(list(
                    User.objects.filter(role=User.Roles.PM_ESG)
                    .values("pk", "first_name", "last_name", "username")
                    .order_by("first_name", "last_name")
                )) if is_pm else "[]",
                "sqr_sse_users_json": json.dumps(list(
                    User.objects.filter(role__in=[User.Roles.ENGINEER, User.Roles.ON_HOLD])
                    .values("pk", "first_name", "last_name", "username")
                    .order_by("first_name", "last_name")
                )) if is_pm else "[]",
                "sqr_rq_options_json": json.dumps(list(
                    Request.objects.order_by("-id")[:300]
                    .values("pk", "reference_code")
                )) if is_pm else "[]",
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        if request.user.role not in ENGINEER_ACCESS_ROLES:
            messages.error(request, "Only engineers can submit SQR entries.")
            return redirect("hub:sqr")

        form = SqrSubmissionForm(request.POST)
        form.fields["linked_request"].queryset = (
            Request.objects.filter(engineer=request.user).select_related("account").only("id", "reference_code", "account__name").order_by("-id")
        )
        _refresh_sqr_request_account_map(form)
        if form.is_valid():
            linked_request = form.cleaned_data.get("linked_request")
            if linked_request and SqrSubmission.objects.filter(linked_request=linked_request).exists():
                form.add_error("linked_request", "An SQR already exists for the selected Request ID.")
                context = self.get_context_data(form=form)
                return self.render_to_response(context)

            try:
                with transaction.atomic():
                    if linked_request:
                        Request.objects.select_for_update().filter(pk=linked_request.pk).first()
                        if SqrSubmission.objects.select_for_update().filter(linked_request=linked_request).exists():
                            form.add_error("linked_request", "An SQR already exists for the selected Request ID.")
                            context = self.get_context_data(form=form)
                            return self.render_to_response(context)

                    submission = form.save(commit=False)
                    submission.engineer = request.user
                    submission.status = SqrSubmission.Status.FOR_PROCESSING
                    # Auto-compute Managed Support Service Amount (col P) per business rules
                    _MANAGED_SCOPES = frozenset([
                        "Deployment Only", "Deployment and Project Management", "Maintenance", "On-Call Services", "Implementation", "Implementation and Project Management", "Managed Support and Maintenance Service",
                    ])
                    scope = form.cleaned_data.get("project_details", "") or ""
                    group = form.cleaned_data.get("customer_company", "") or ""
                    if scope in _MANAGED_SCOPES:
                        submission.managed_support_amount = (
                            Decimal("149000.00") if group == "ESS" else Decimal("192000.00")
                        )
                    else:
                        submission.managed_support_amount = None
                    submission.save()
            except IntegrityError:
                form.add_error("linked_request", "An SQR already exists for the selected Request ID.")
                context = self.get_context_data(form=form)
                return self.render_to_response(context)
            # Reload with reviewer to ensure email field is populated
            submission = (
                SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer")
                .get(pk=submission.pk)
            )
            _send_sqr_new_submission_email(submission, http_request=request)
            SqrReportsDataView._notify_sqr_submission(submission)
            messages.success(request, f"SQR submitted successfully ({submission.reference_code}).")
            return redirect("hub:sqr")

        context = self.get_context_data(form=form)
        return self.render_to_response(context)


class SqrReportsDataView(LoginRequiredMixin, View):
    """AJAX endpoint that returns the latest SQR reports fragment for live refresh."""

    def get(self, request, *args, **kwargs):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

        report_view = SqrListView()
        report_view.request = request
        report_view.args = args
        report_view.kwargs = kwargs
        context = report_view.get_context_data()
        html = render_to_string("hub/partials/sqr_reports.html", context, request=request)
        return JsonResponse({"ok": True, "html": html})

    @staticmethod
    def _notify_sqr_submission(submission: SqrSubmission) -> None:
        actor_name = submission.engineer.get_full_name() or submission.engineer.username
        message = (
            f"{actor_name} submitted {submission.reference_code} for {submission.customer_name}."
        )
        recipients: dict[int, User] = {submission.pm_esg_reviewer_id: submission.pm_esg_reviewer}
        for admin in User.objects.filter(role=User.Roles.ADMIN):
            recipients[admin.pk] = admin
        recipients.pop(submission.engineer_id, None)
        for recipient in recipients.values():
            Notification.objects.create(
                recipient=recipient,
                message=message,
                actor=actor_name,
                source="SQR · New Submission",
            )


class SqrEngineerUpdateView(LoginRequiredMixin, UpdateView):
    model = SqrSubmission
    form_class = SqrSubmissionForm
    template_name = "hub/sqr_submission_form.html"
    success_url = reverse_lazy("hub:sqr")

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ENGINEER_ACCESS_ROLES and request.user.role != User.Roles.ADMIN:
            messages.error(request, "Only engineers or admins can edit SQR submissions.")
            return redirect("hub:sqr")
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, "That SQR submission no longer exists.")
            return redirect("hub:sqr")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        obj = self.object
        # Exclude requests already linked to OTHER SQR submissions;
        # this keeps the current linked_request in the queryset so it stays valid.
        others_used_ids = (
            SqrSubmission.objects.exclude(pk=obj.pk)
            .exclude(linked_request_id=None)
            .values_list("linked_request_id", flat=True)
        )
        if self.request.user.role == User.Roles.ADMIN:
            qs = (
                Request.objects
                .exclude(id__in=others_used_ids)
                .select_related("account").only("id", "reference_code", "account__name").order_by("-id")
            )
        else:
            qs = (
                Request.objects
                .filter(
                    Q(engineer=self.request.user)
                    | Q(backup_engineer=self.request.user)
                    | Q(pk=obj.linked_request_id)
                )
                .exclude(id__in=others_used_ids)
                .select_related("account").only("id", "reference_code", "account__name").order_by("-id")
            )
        form.fields["linked_request"].queryset = qs
        return form

    def get_queryset(self):
        if self.request.user.role == User.Roles.ADMIN:
            return SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer")
        return SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer").filter(engineer=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["back_url"] = reverse("hub:sqr")
        context["form_title"] = f"Edit Submission {self.object.reference_code}"
        context["submit_label"] = "Save changes"
        return context

    def form_valid(self, form):
        original = SqrSubmission.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        changed_fields = []
        for field in form.changed_data:
            if hasattr(self.object, field):
                changed_fields.append(field)
        if changed_fields:
            old_values = SqrInlineFieldUpdateView._snapshot_values(original, changed_fields)
            new_values = SqrInlineFieldUpdateView._snapshot_values(self.object, changed_fields)
            record_sqr_history(
                self.object,
                self.request.user,
                field=changed_fields[0] if len(changed_fields) == 1 else "",
                old_values=old_values,
                new_values=new_values,
                source="engineer_edit",
            )
        messages.success(self.request, f"SQR {self.object.reference_code} updated.")
        return response


class SqrEngineerDeleteView(LoginRequiredMixin, DeleteView):
    model = SqrSubmission
    template_name = "hub/sqr_confirm_delete.html"
    success_url = reverse_lazy("hub:sqr")

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ENGINEER_ACCESS_ROLES and request.user.role != User.Roles.ADMIN:
            messages.error(request, "Only engineers or admins can delete SQR submissions.")
            return redirect("hub:sqr")
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, "That SQR submission no longer exists.")
            return redirect("hub:sqr")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        if self.request.user.role == User.Roles.ADMIN:
            return SqrSubmission.objects.all()
        return SqrSubmission.objects.filter(engineer=self.request.user)

    def form_valid(self, form):
        reference_code = self.object.reference_code
        response = super().form_valid(form)
        messages.success(self.request, f"SQR {reference_code} deleted.")
        return response


SQR_HISTORY_FIELD_LABELS = {
    "customer_name": "Account Name",
    "customer_company": "Group Name",
    "customer_contact": "Account Manager",
    "project_title": "Service Description",
    "project_details": "Scope of Services",
    "linked_request": "RQ ID",
    "sse_manhrs": "SSE Man-hrs",
    "sse_amount": "SSE Amount",
    "pm_manhrs": "PM Man-hrs",
    "pm_amount": "PM Amount",
    "managed_support_amount": "Managed Support Svc. Amt.",
    "discount_rate": "Discount Rate",
    "status": "SQR Status",
    "reviewed_at": "Approval Date",
    "reviewed_by": "Reviewed By",
    "review_notes": "Review Notes",
    "validity_due_date": "Validity Due Date",
    "proposal_status": "Proposal Status",
    "po_pnl_date": "PO / PNL Date",
    "assigned_pm": "Assigned PM",
    "assigned_sse": "Assigned SSE",
    "delivery_start_date": "Start Date",
    "delivery_target_finish_date": "Target Finish Date",
    "overall_status": "Overall Status",
    "delivery_health": "Health Status",
    "delivery_progress": "Overall Progress %",
    "key_updates_risks": "Key Updates / Risks / Issues",
    "delivery_actual_finish_date": "Actual Finish Date",
    "delivery_completion_signed_date": "Completion Signed Date",
    "warranty_end_date": "Completion Warranty End Date",
    "managed_support_start_date": "Maintenance Start Date",
    "revenue_date": "Billed Date",
    "revenue_source": "Billing Type",
    "revenue_reference_no": "Billing Reference",
    "revenue_status": "Billing Status",
    "revenue_remarks": "PO Remarks",
    "revenue_overview": "Billing Remarks",
    "revenue_declaration": "Revenue Declaration",
}

SQR_HISTORY_CHOICE_LABELS = {
    "status": dict(SqrSubmission.Status.choices),
    "proposal_status": dict(SqrSubmission.ProposalStatus.choices),
    "delivery_health": dict(SqrSubmission.DeliveryHealth.choices),
    "overall_status": dict(SqrSubmission.OverallStatus.choices),
    "revenue_declaration": dict(SqrSubmission.RevenueDeclaration.choices),
    "revenue_status": dict(SqrSubmission.RevenueStatus.choices),
    "revenue_source": {"internal": "Internal", "invoiced": "Invoiced", "unbilled": "Unbilled"},
    "project_details": {
        "Deployment Only": "Deployment Only",
        "On-Call Services": "On-Call Services",
        "Maintenance": "Maintenance",
        "Project Management": "Project Management",
        "Deployment and Project Management": "Deployment and Project Management",
    },
}


def get_sqr_history_field_label(field: str) -> str:
    return SQR_HISTORY_FIELD_LABELS.get(field, (field or "SQR").replace("_", " ").title())


def format_sqr_history_value(field: str, value) -> str:
    if value in (None, ""):
        return "—"
    if field in SQR_HISTORY_CHOICE_LABELS:
        return SQR_HISTORY_CHOICE_LABELS[field].get(value, str(value))
    if field in {"assigned_pm", "assigned_sse", "reviewed_by"}:
        user = User.objects.filter(pk=value).first()
        return (user.get_full_name() or user.username) if user else str(value)
    if field == "linked_request":
        request_obj = Request.objects.filter(pk=value).first()
        return request_obj.reference_code if request_obj else str(value)
    if field == "discount_rate":
        return f"{value}%" if str(value) != "0" else "No Discount"
    if field.endswith("_amount") or field in {"quotation_total_price", "computed_total_price", "computed_discount_amount"}:
        try:
            amount = Decimal(str(value))
            return f"₱{amount:,.2f}"
        except Exception:
            return str(value)
    return str(value)


def build_sqr_history_summary(field: str, old_values: dict, new_values: dict, action: str = "updated") -> str:
    changed = [key for key in new_values.keys() if old_values.get(key) != new_values.get(key)]
    primary = field if field in changed else (changed[0] if changed else field)
    label = get_sqr_history_field_label(primary)
    if action == SqrSubmissionHistory.Action.RESTORED:
        return f"Restored {label}."
    if not changed:
        return f"Saved {label}."
    old_display = format_sqr_history_value(primary, old_values.get(primary))
    new_display = format_sqr_history_value(primary, new_values.get(primary))
    extra = len(changed) - 1
    suffix = f" and {extra} related value{'s' if extra != 1 else ''}" if extra > 0 else ""
    return f"Changed {label} from {old_display} to {new_display}{suffix}."


def record_sqr_history(
    submission: SqrSubmission,
    actor: User,
    *,
    field: str,
    old_values: dict,
    new_values: dict,
    action: str = SqrSubmissionHistory.Action.UPDATED,
    source: str = "inline",
    metadata: dict | None = None,
) -> SqrSubmissionHistory | None:
    if old_values == new_values:
        return None
    return SqrSubmissionHistory.objects.create(
        submission=submission,
        actor=actor,
        action=action,
        field=field or "",
        old_values=old_values or {},
        new_values=new_values or {},
        summary=build_sqr_history_summary(field, old_values or {}, new_values or {}, action),
        source=source,
        metadata=metadata or {},
    )


class SqrInlineFieldUpdateView(LoginRequiredMixin, View):
    """AJAX endpoint for inline cell editing of SQR submissions (Admin / PM-ESG)."""

    _ADMIN_ALLOWED = frozenset([
        "customer_name", "customer_company", "customer_contact",
        "project_title", "project_details", "sse_manhrs",
        "pm_manhrs", "discount_rate", "status", "proposal_status",
        "po_pnl_date", "delivery_start_date", "overall_status", "delivery_health",
        "delivery_progress", "key_updates_risks", "delivery_target_finish_date",
        "delivery_actual_finish_date", "delivery_completion_signed_date",
        "warranty_end_date", "revenue_date", "revenue_source", "revenue_reference_no",
        "revenue_status", "revenue_remarks", "revenue_overview", "revenue_declaration",
        "managed_support_amount", "assigned_pm", "assigned_sse", "linked_request",
    ])
    _PM_ESG_ALLOWED = frozenset([
        "project_details", "sse_manhrs", "pm_manhrs", "discount_rate", "status", "proposal_status",
        "po_pnl_date", "delivery_start_date", "overall_status", "delivery_health",
        "delivery_progress", "key_updates_risks", "delivery_target_finish_date",
        "delivery_actual_finish_date", "delivery_completion_signed_date",
        "warranty_end_date", "managed_support_start_date", "revenue_date", "revenue_source",
        "revenue_reference_no", "revenue_status", "revenue_remarks", "revenue_overview",
        "revenue_declaration", "managed_support_amount", "assigned_pm", "assigned_sse", "linked_request",
    ])
    _DATE_FIELDS = frozenset([
        "po_pnl_date", "delivery_start_date", "delivery_target_finish_date",
        "delivery_actual_finish_date", "delivery_completion_signed_date", "warranty_end_date",
        "managed_support_start_date", "revenue_date",
    ])
    _INT_FIELDS = frozenset(["discount_rate", "delivery_progress"])
    _DECIMAL_FIELDS = frozenset(["sse_manhrs", "pm_manhrs", "managed_support_amount"])
    _FK_FIELDS = frozenset(["assigned_pm", "assigned_sse", "linked_request"])
    _CLOSED_LOST_LOCKED_FIELDS = frozenset([
        "po_pnl_date", "assigned_pm", "assigned_sse", "delivery_start_date",
        "delivery_target_finish_date", "overall_status", "delivery_health",
        "delivery_progress", "key_updates_risks", "delivery_actual_finish_date",
        "delivery_completion_signed_date", "managed_support_start_date",
        "revenue_date", "revenue_source", "revenue_reference_no", "revenue_status",
        "revenue_remarks", "revenue_overview", "revenue_declaration",
    ])
    _SERIALIZED_FK_FIELDS = _FK_FIELDS | frozenset(["reviewed_by"])
    _UNDO_RETENTION = timedelta(hours=24)
    _PRICE_TRIGGERS = frozenset(["pm_amount", "sse_amount", "managed_support_amount", "discount_rate"])

    @classmethod
    def _allowed_fields_for_user(cls, user):
        if user.role == User.Roles.ADMIN:
            return cls._ADMIN_ALLOWED
        if user.role == User.Roles.PM_ESG:
            return cls._PM_ESG_ALLOWED
        return None

    @classmethod
    def _tracked_fields_for_change(cls, field, save_fields=None):
        tracked = set(save_fields or []) | {field}
        if field == "sse_manhrs":
            tracked.update(["sse_amount", "pm_manhrs", "pm_amount"])
        if field == "pm_manhrs":
            tracked.add("pm_amount")
        if field in ("project_details", "customer_company"):
            tracked.add("managed_support_amount")
        if field == "status":
            tracked.update(["reviewed_at", "reviewed_by", "validity_due_date", "review_notes", "assigned_pm"])
        return sorted(tracked)

    @classmethod
    def _serialize_field_value(cls, submission, field):
        if field in cls._SERIALIZED_FK_FIELDS:
            value = getattr(submission, f"{field}_id", None)
        else:
            value = getattr(submission, field)
        if isinstance(value, Decimal):
            normalized = value.normalize()
            return format(normalized, "f")
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return value

    @classmethod
    def _snapshot_values(cls, submission, fields):
        return {field: cls._serialize_field_value(submission, field) for field in fields}

    @classmethod
    def _apply_serialized_field_value(cls, submission, field, value):
        if field in cls._DATE_FIELDS or field == "validity_due_date":
            restored = date.fromisoformat(value) if value else None
        elif field == "reviewed_at":
            restored = datetime.fromisoformat(value) if value else None
        elif field in cls._INT_FIELDS:
            restored = int(value) if value not in ("", None) else None
            if field == "discount_rate" and restored is None:
                restored = 0
        elif field in cls._DECIMAL_FIELDS or field in ("sse_amount", "pm_amount"):
            restored = Decimal(str(value)) if value not in ("", None) else None
        else:
            restored = value

        if field in cls._SERIALIZED_FK_FIELDS:
            setattr(submission, f"{field}_id", restored)
        else:
            setattr(submission, field, restored)

    @staticmethod
    def _fmt_date_display(value):
        return value.strftime("%b %d, %Y") if value else ""

    @classmethod
    def _build_inline_response(cls, submission, field, save_fields, request=None, change=None, undone=False, approval_urls=None):
        response_data = {
            "ok": True,
            "field": field,
            "field_value": cls._serialize_field_value(submission, field),
        }
        approval_urls = approval_urls or {}
        if approval_urls.get("draft"):
            response_data["approval_email_draft_url"] = approval_urls["draft"]
        if approval_urls.get("eml"):
            response_data["approval_email_eml_url"] = approval_urls["eml"]
        if change is not None:
            response_data["change_id"] = change.pk
            response_data["undo_url"] = reverse("hub:sqr-inline-undo", args=[submission.pk, change.pk])
            response_data["undo_expires_at"] = change.expires_at.isoformat()
        if undone:
            response_data["undone"] = True
        if field == "status":
            response_data["workflow_side_effect_warning"] = True
            response_data["workflow_side_effect_message"] = (
                "Status was reverted in Request Hub. Emails, Teams chats, or Outlook drafts already opened cannot be undone."
                if undone else
                "This status change can be reverted in Request Hub, but emails, Teams chats, or Outlook drafts cannot be undone."
            )

        save_fields = set(save_fields or [])
        if "pm_amount" in save_fields:
            response_data["pm_amount"] = str(submission.pm_amount) if submission.pm_amount is not None else ""
        if "pm_manhrs" in save_fields:
            response_data["pm_manhrs"] = str(int(submission.pm_manhrs)) if submission.pm_manhrs is not None else ""
        if "sse_amount" in save_fields:
            response_data["sse_amount"] = str(submission.sse_amount) if submission.sse_amount is not None else ""
        if any(f in save_fields for f in cls._PRICE_TRIGGERS) or field in cls._PRICE_TRIGGERS:
            da = submission.computed_discount_amount
            tp = submission.computed_total_price
            response_data["computed_discount_amount"] = str(da) if da is not None else ""
            response_data["computed_total_price"] = str(tp) if tp is not None else ""
        if "managed_support_amount" in save_fields:
            msa = submission.managed_support_amount
            response_data["managed_support_amount"] = str(msa) if msa is not None else ""
        if "reviewed_at" in save_fields:
            rat = submission.reviewed_at
            rat_manila = rat.astimezone(MANILA_TZ) if rat else None
            response_data["reviewed_at"] = rat_manila.strftime("%b %d, %Y") if rat_manila else ""
        if "validity_due_date" in save_fields:
            vdd = submission.validity_due_date
            response_data["validity_due_date"] = vdd.strftime("%b %d, %Y") if vdd else ""
        if "assigned_pm" in save_fields:
            pm = submission.assigned_pm
            response_data["assigned_pm_pk"] = pm.pk if pm else ""
            response_data["assigned_pm_name"] = (pm.get_full_name() or pm.username) if pm else ""
        if field == "assigned_sse" or "assigned_sse" in save_fields:
            sse = submission.assigned_sse
            response_data["assigned_sse_name"] = (sse.get_full_name() or sse.username) if sse else ""
        if field == "linked_request" or "linked_request" in save_fields:
            req = submission.linked_request
            response_data["rq_pk"] = req.pk if req else ""
            response_data["rq_code"] = req.reference_code if req else ""

        if field == "delivery_completion_signed_date" or "delivery_completion_signed_date" in save_fields:
            response_data["post_svc_warranty_end_date"] = cls._fmt_date_display(submission.computed_post_svc_warranty_end_date)
            response_data["support_start_date"] = cls._fmt_date_display(submission.computed_support_start_date)
            response_data["managed_support_end_date"] = cls._fmt_date_display(submission.computed_managed_support_end_date)
        if field in ("project_details", "customer_company", "managed_support_start_date", "managed_support_amount") or save_fields.intersection({"project_details", "customer_company", "managed_support_start_date", "managed_support_amount"}):
            response_data["support_start_date"] = cls._fmt_date_display(submission.computed_support_start_date)
            response_data["managed_support_end_date"] = cls._fmt_date_display(submission.computed_managed_support_end_date)
        return response_data

    def post(self, request, pk):
        allowed = self._allowed_fields_for_user(request.user)
        if allowed is None:
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

        try:
            data = json.loads(request.body)
        except (ValueError, KeyError):
            return JsonResponse({"ok": False, "error": "Invalid request"}, status=400)

        field = data.get("field", "")
        value = data.get("value", "")
        review_notes_override = data.get("review_notes", None)  # supplied when status → for_revision

        if not field or field not in allowed:
            return JsonResponse({"ok": False, "error": "Field not allowed"}, status=400)

        submission = get_object_or_404(
            SqrSubmission.objects.select_related("engineer"), pk=pk
        )
        old_status = submission.status
        original_submission = SqrSubmission.objects.get(pk=pk)

        if (
            submission.proposal_status in (
                SqrSubmission.ProposalStatus.CLOSED_LOST,
                SqrSubmission.ProposalStatus.CLOSED_CANCELED,
            )
            and field in self._CLOSED_LOST_LOCKED_FIELDS
        ):
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Delivery and billing fields are locked when Proposal Status is Closed-Lost or Closed-Canceled",
                },
                status=400,
            )

        try:
            if field in self._DATE_FIELDS:
                coerced = date.fromisoformat(value) if value else None
            elif field in self._INT_FIELDS:
                coerced = int(value) if value not in ("", None) else None
                if field == "discount_rate" and coerced is None:
                    coerced = 0
                if field == "delivery_progress" and coerced is not None:
                    coerced = max(0, min(100, coerced))
            elif field in self._DECIMAL_FIELDS:
                coerced = Decimal(str(value)) if value not in ("", None) else None
            elif field in self._FK_FIELDS:
                coerced = int(value) if value not in ("", None) else None
            else:
                coerced = value  # str fields — allow empty string
        except (ValueError, TypeError):
            return JsonResponse({"ok": False, "error": f"Invalid value for {field}"}, status=400)

        # For FK fields, set the _id attribute directly
        if field in self._FK_FIELDS:
            setattr(submission, field + "_id", coerced)
        else:
            setattr(submission, field, coerced)
        save_fields = [field]

        # Auto-recompute PM Amount (col O = col N × 3000) when pm_manhrs changes
        if field == "pm_manhrs":
            submission.pm_amount = (
                (Decimal(str(coerced)) * Decimal("3000")).quantize(Decimal("0.01"))
                if coerced is not None else None
            )
            save_fields.append("pm_amount")

        # Auto-recompute SSE Amount (col M = col L × 2000) when sse_manhrs changes
        if field == "sse_manhrs":
            submission.sse_amount = (
                (Decimal(str(coerced)) * Decimal("2000")).quantize(Decimal("0.01"))
                if coerced is not None else None
            )
            submission.pm_manhrs = submission.recommended_pm_manhrs_for_sse(coerced)
            submission.pm_amount = (
                (Decimal(str(submission.pm_manhrs)) * Decimal("3000")).quantize(Decimal("0.01"))
                if submission.pm_manhrs is not None else None
            )
            save_fields.extend(["sse_amount", "pm_manhrs", "pm_amount"])

        # Auto-recompute Managed Support Svc. Amt. (col P) when scope or group changes
        _MANAGED_SCOPES = frozenset([
            "Deployment Only", "Deployment and Project Management", "Maintenance", "On-Call Services", "Implementation", "Implementation and Project Management", "Managed Support and Maintenance Service",
        ])
        if field in ("project_details", "customer_company"):
            scope = coerced if field == "project_details" else submission.project_details
            group = coerced if field == "customer_company" else submission.customer_company
            new_msa = (
                Decimal("149000.00") if group == "ESS" else Decimal("192000.00")
            ) if scope in _MANAGED_SCOPES else None
            submission.managed_support_amount = new_msa
            save_fields.append("managed_support_amount")

        # When status changes: auto-set reviewed_at, validity_due_date, reviewed_by, assigned_pm
        if field == "status":
            if coerced == SqrSubmission.Status.APPROVED:
                now = timezone.now()
                if not submission.reviewed_at:
                    submission.reviewed_at = now
                submission.validity_due_date = submission.reviewed_at.date() + timedelta(days=45)
                submission.reviewed_by = request.user
                save_fields += ["reviewed_at", "reviewed_by", "validity_due_date"]
                if not submission.assigned_pm_id:
                    submission.assigned_pm_id = submission.pm_esg_reviewer_id
                    save_fields.append("assigned_pm")
            else:
                submission.reviewed_at = None
                submission.validity_due_date = None
                save_fields += ["reviewed_at", "validity_due_date"]
            # Save review_notes when status → for_revision (provided by the comments modal)
            if coerced == SqrSubmission.Status.FOR_REVISION and review_notes_override is not None:
                submission.review_notes = review_notes_override
                submission.reviewed_by = request.user
                if "review_notes" not in save_fields:
                    save_fields.append("review_notes")
                if "reviewed_by" not in save_fields:
                    save_fields.append("reviewed_by")

        tracked_fields = self._tracked_fields_for_change(field, save_fields)
        old_values = self._snapshot_values(original_submission, tracked_fields)
        with transaction.atomic():
            submission.save(update_fields=save_fields)
            new_values = self._snapshot_values(submission, tracked_fields)
            change = SqrSubmissionChange.objects.create(
                submission=submission,
                changed_by=request.user,
                field=field,
                old_values=old_values,
                new_values=new_values,
                expires_at=timezone.now() + self._UNDO_RETENTION,
            )
            record_sqr_history(
                submission,
                request.user,
                field=field,
                old_values=old_values,
                new_values=new_values,
                source="inline",
                metadata={"change_id": change.pk},
            )

        # Email engineer when status changed to For Revision via inline edit
        if (
            field == "status"
            and coerced == SqrSubmission.Status.FOR_REVISION
            and old_status != SqrSubmission.Status.FOR_REVISION
        ):
            reviewer_name = request.user.get_full_name() or request.user.username
            _send_sqr_for_revision_email(
                submission,
                reviewer_name,
                http_request=request,
            )

        # When status → Approved: launch the user's default email app with the draft.
        approval_urls = {}
        if (
            field == "status"
            and coerced == SqrSubmission.Status.APPROVED
            and old_status != SqrSubmission.Status.APPROVED
        ):
            approval_urls["draft"] = _build_sqr_approval_mailto_url(_get_sqr_approval_email_context(submission))
            approval_urls["eml"] = reverse("hub:sqr-approval-email-eml", args=[submission.pk])

        response_data = self._build_inline_response(
            submission,
            field,
            tracked_fields,
            request=request,
            change=change,
            approval_urls=approval_urls,
        )
        return JsonResponse(response_data)


class SqrInlineFieldUndoView(SqrInlineFieldUpdateView):
    """AJAX endpoint that reverts a previously saved inline SQR field change."""

    def post(self, request, pk, change_id):
        allowed = self._allowed_fields_for_user(request.user)
        if allowed is None:
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

        with transaction.atomic():
            change = get_object_or_404(
                SqrSubmissionChange.objects.select_for_update().select_related("submission"),
                pk=change_id,
                submission_id=pk,
            )
            if change.field not in allowed:
                return JsonResponse({"ok": False, "error": "Field not allowed"}, status=400)
            if change.is_undone:
                return JsonResponse({"ok": False, "error": "This change was already undone."}, status=400)
            if change.is_expired:
                return JsonResponse({"ok": False, "error": "Undo expired for this change."}, status=400)

            submission = SqrSubmission.objects.select_for_update().get(pk=pk)
            if (
                submission.proposal_status in (
                    SqrSubmission.ProposalStatus.CLOSED_LOST,
                    SqrSubmission.ProposalStatus.CLOSED_CANCELED,
                )
                and change.field != "proposal_status"
                and any(field in self._CLOSED_LOST_LOCKED_FIELDS for field in change.old_values)
            ):
                return JsonResponse(
                    {
                        "ok": False,
                        "error": "Delivery and billing fields are locked when Proposal Status is Closed-Lost or Closed-Canceled",
                    },
                    status=400,
                )

            current_values = self._snapshot_values(submission, change.new_values.keys())
            if current_values != change.new_values:
                return JsonResponse(
                    {
                        "ok": False,
                        "error": "Undo skipped because this SQR was changed again after the saved edit.",
                    },
                    status=409,
                )

            restore_fields = list(change.old_values.keys())
            for restore_field, restore_value in change.old_values.items():
                self._apply_serialized_field_value(submission, restore_field, restore_value)
            update_values = {}
            for restore_field in restore_fields:
                if restore_field in self._SERIALIZED_FK_FIELDS:
                    update_values[f"{restore_field}_id"] = getattr(submission, f"{restore_field}_id")
                else:
                    update_values[restore_field] = getattr(submission, restore_field)
            SqrSubmission.objects.filter(pk=submission.pk).update(**update_values)
            change.undone_at = timezone.now()
            change.undone_by = request.user
            change.save(update_fields=["undone_at", "undone_by"])
            submission = SqrSubmission.objects.select_related(
                "assigned_pm", "assigned_sse", "linked_request"
            ).get(pk=pk)

        response_data = self._build_inline_response(
            submission,
            change.field,
            restore_fields,
            request=request,
            undone=True,
        )
        return JsonResponse(response_data)


class SqrHistoryView(LoginRequiredMixin, View):
    """Return PM/Admin-visible permanent history for a single SQR."""

    def get(self, request, pk):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

        submission = get_object_or_404(SqrSubmission, pk=pk)
        entries = []
        for entry in submission.history_entries.select_related("actor", "restored_by").all()[:200]:
            changed_fields = [
                key for key in entry.new_values.keys()
                if entry.old_values.get(key) != entry.new_values.get(key)
            ]
            fields = changed_fields or list(entry.new_values.keys()) or ([entry.field] if entry.field else [])
            entries.append({
                "id": entry.pk,
                "action": entry.action,
                "action_label": entry.get_action_display(),
                "field": entry.field,
                "field_label": get_sqr_history_field_label(entry.field),
                "summary": entry.summary,
                "actor": entry.actor.get_full_name() or entry.actor.username,
                "created_at": timezone.localtime(entry.created_at, MANILA_TZ).strftime("%b %d, %Y · %I:%M %p"),
                "source": entry.source,
                "is_restored": entry.is_restored,
                "restored_by": (entry.restored_by.get_full_name() or entry.restored_by.username) if entry.restored_by else "",
                "restored_at": timezone.localtime(entry.restored_at, MANILA_TZ).strftime("%b %d, %Y · %I:%M %p") if entry.restored_at else "",
                "restore_url": reverse("hub:sqr-history-restore", args=[submission.pk, entry.pk]),
                "restorable": entry.action == SqrSubmissionHistory.Action.UPDATED and not entry.is_restored and bool(entry.old_values),
                "changes": [
                    {
                        "field": changed_field,
                        "field_label": get_sqr_history_field_label(changed_field),
                        "old": format_sqr_history_value(changed_field, entry.old_values.get(changed_field)),
                        "new": format_sqr_history_value(changed_field, entry.new_values.get(changed_field)),
                    }
                    for changed_field in fields
                ],
            })
        return JsonResponse({
            "ok": True,
            "submission": {
                "id": submission.pk,
                "reference_code": submission.reference_code or str(submission),
            },
            "entries": entries,
        })


class SqrHistoryRestoreView(LoginRequiredMixin, View):
    """Restore a single permanent SQR history entry when no later edit touched it."""

    def post(self, request, pk, history_id):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

        allowed = SqrInlineFieldUpdateView._allowed_fields_for_user(request.user) or frozenset()
        with transaction.atomic():
            history = get_object_or_404(
                SqrSubmissionHistory.objects.select_for_update().select_related("submission"),
                pk=history_id,
                submission_id=pk,
            )
            if history.action != SqrSubmissionHistory.Action.UPDATED or history.is_restored:
                return JsonResponse({"ok": False, "error": "This history entry cannot be restored."}, status=400)
            restore_fields = list(history.old_values.keys())
            disallowed = [field for field in restore_fields if field not in allowed and field not in {"reviewed_at", "reviewed_by", "validity_due_date", "review_notes", "sse_amount", "pm_amount", "managed_support_amount"}]
            if disallowed:
                return JsonResponse({"ok": False, "error": "You cannot restore one or more fields in this history entry."}, status=403)

            submission = SqrSubmission.objects.select_for_update().get(pk=pk)
            current_values = SqrInlineFieldUpdateView._snapshot_values(submission, history.new_values.keys())
            if current_values != history.new_values:
                return JsonResponse(
                    {
                        "ok": False,
                        "error": "Restore skipped because this SQR changed after the selected history entry.",
                    },
                    status=409,
                )

            for restore_field, restore_value in history.old_values.items():
                SqrInlineFieldUpdateView._apply_serialized_field_value(submission, restore_field, restore_value)
            update_values = {}
            for restore_field in restore_fields:
                if restore_field in SqrInlineFieldUpdateView._SERIALIZED_FK_FIELDS:
                    update_values[f"{restore_field}_id"] = getattr(submission, f"{restore_field}_id")
                else:
                    update_values[restore_field] = getattr(submission, restore_field)
            SqrSubmission.objects.filter(pk=submission.pk).update(**update_values)
            history.restored_at = timezone.now()
            history.restored_by = request.user
            history.save(update_fields=["restored_at", "restored_by"])
            restore_entry = record_sqr_history(
                submission,
                request.user,
                field=history.field,
                old_values=history.new_values,
                new_values=history.old_values,
                action=SqrSubmissionHistory.Action.RESTORED,
                source="history_restore",
                metadata={"restored_history_id": history.pk},
            )

        return JsonResponse({
            "ok": True,
            "message": "History entry restored.",
            "restored_history_id": history.pk,
            "restore_entry_id": restore_entry.pk if restore_entry else None,
            "reload_required": True,
        })


class SqrReviewUpdateView(LoginRequiredMixin, View):
    """Legacy SQR review page removed; status review now happens inline on the SQR tracker."""

    def get(self, request, *args, **kwargs):
        messages.info(request, "The SQR review page has been removed. Update SQR status directly from the tracker.")
        return redirect("hub:sqr")

    def post(self, request, *args, **kwargs):
        messages.info(request, "The SQR review page has been removed. Update SQR status directly from the tracker.")
        return redirect("hub:sqr")


class SqrRevenueTrackerUpdateView(LoginRequiredMixin, View):
    @staticmethod
    def _flash_form_errors(request, form):
        for field, errors in form.errors.items():
            if field == "__all__":
                for error in errors:
                    messages.error(request, error)
                continue
            label = form.fields.get(field).label if field in form.fields else field
            for error in errors:
                messages.error(request, f"{label}: {error}")

    @staticmethod
    def _tracker_redirect_url():
        return f"{reverse('hub:sqr')}?{urlencode({'tab': 'revenue-tracker'})}"

    def post(self, request, pk):
        if request.user.role != PM_ESG_ROLE:
            messages.error(request, "Only PM-ESG can update Revenue Tracker details.")
            return redirect(self._tracker_redirect_url())

        submission = get_object_or_404(
            SqrSubmission.objects.select_related("pm_esg_reviewer"),
            pk=pk,
            pm_esg_reviewer=request.user,
        )

        stage_action = (request.POST.get("stage_action") or "").strip().lower()
        if stage_action == "quotation":
            if submission.revenue_stage_key != "quotation":
                messages.error(request, "Quotation pricing can only be updated in Quotation stage.")
                return redirect(self._tracker_redirect_url())

            form = SqrRevenueQuotationForm(request.POST, instance=submission, prefix=f"quote-{submission.pk}")
            if form.is_valid():
                form.save()
                messages.success(request, f"Quotation pricing saved for {submission.reference_code}.")
            else:
                self._flash_form_errors(request, form)
            return redirect(self._tracker_redirect_url())

        if stage_action == "order":
            if submission.revenue_stage_key != "order":
                messages.error(request, "Only approved SQR entries in Order stage can attach a P.O.")
                return redirect(self._tracker_redirect_url())

            form = SqrRevenueOrderForm(request.POST, instance=submission, prefix=f"order-{submission.pk}")
            if form.is_valid():
                updated_submission = form.save(commit=False)
                if not updated_submission.quotation_total_price:
                    messages.error(request, "Set the quotation pricing before attaching a P.O.")
                    return redirect(self._tracker_redirect_url())

                if not submission.po_attachment_link:
                    updated_submission.po_attached_at = timezone.now()
                updated_submission.save()
                messages.success(request, f"P.O attached for {submission.reference_code}. Moved to Revenue stage.")
            else:
                self._flash_form_errors(request, form)
            return redirect(self._tracker_redirect_url())

        messages.error(request, "Invalid revenue tracker action.")
        return redirect(self._tracker_redirect_url())


class SqrProposalUpdateView(LoginRequiredMixin, View):
    """PM-ESG / Admin: update pricing and deal/proposal status on an approved SQR."""

    def post(self, request, pk):
        if request.user.role not in ADMIN_PANEL_ROLES:
            messages.error(request, "Only PM-ESG or Admin can update Proposal details.")
            return redirect("hub:sqr")

        submission = get_object_or_404(
            SqrSubmission.objects.select_related("pm_esg_reviewer"),
            pk=pk,
        )

        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            messages.error(request, "Only the assigned PM-ESG can update this proposal.")
            return redirect("hub:sqr")

        if submission.status != SqrSubmission.Status.APPROVED:
            messages.error(request, "Proposal details can only be updated after the SQR is Approved.")
            return redirect("hub:sqr")

        form = SqrProposalStatusForm(request.POST, instance=submission)
        if form.is_valid():
            form.save()
            messages.success(request, f"Proposal details updated for {submission.reference_code}.")
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f"{label}: {error}")

        return redirect("hub:sqr")


class SqrToRevenueView(LoginRequiredMixin, View):
    """PM-ESG / Admin: unlock Step 4 Revenue Stage by clicking 'To Revenue'."""

    def post(self, request, pk):
        if request.user.role not in ADMIN_PANEL_ROLES:
            messages.error(request, "Only PM-ESG or Admin can unlock the Revenue Stage.")
            return redirect("hub:sqr")

        submission = get_object_or_404(SqrSubmission, pk=pk)

        if not submission.revenue_unlocked:
            submission.revenue_unlocked = True
            submission.save(update_fields=["revenue_unlocked", "updated_at"])
            messages.success(request, f"Revenue Stage unlocked for {submission.reference_code}.")
        return redirect("hub:sqr")


class SqrDeliveryUpdateView(LoginRequiredMixin, View):
    """PM-ESG / Admin: update Service Delivery tracking on a Closed Won SQR."""

    def post(self, request, pk):
        if request.user.role not in ADMIN_PANEL_ROLES:
            messages.error(request, "Only PM-ESG or Admin can update Service Delivery details.")
            return redirect("hub:sqr")

        submission = get_object_or_404(
            SqrSubmission.objects.select_related("pm_esg_reviewer"),
            pk=pk,
        )

        form = SqrDeliveryForm(request.POST, instance=submission)
        if form.is_valid():
            form.save()
            messages.success(request, f"Service delivery updated for {submission.reference_code}.")
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f"{label}: {error}")

        return redirect("hub:sqr")


class SqrRevenueUpdateView(LoginRequiredMixin, View):
    """PM-ESG / Admin: record revenue recognition details on an approved SQR."""

    def post(self, request, pk):
        if request.user.role not in ADMIN_PANEL_ROLES:
            messages.error(request, "Only PM-ESG or Admin can update Revenue details.")
            return redirect("hub:sqr")

        submission = get_object_or_404(
            SqrSubmission.objects.select_related("pm_esg_reviewer"),
            pk=pk,
        )

        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            messages.error(request, "Only the assigned PM-ESG can update this revenue record.")
            return redirect("hub:sqr")

        if submission.status != SqrSubmission.Status.APPROVED:
            messages.error(request, "Revenue details can only be updated after the SQR is Approved.")
            return redirect("hub:sqr")

        form = SqrRevenueForm(request.POST, instance=submission)
        if form.is_valid():
            form.save()
            messages.success(request, f"Revenue details updated for {submission.reference_code}.")
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f"{label}: {error}")

        return redirect("hub:sqr")


class SqrTeamsRedirectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        submission = get_object_or_404(
            SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer"),
            pk=pk,
        )

        redirect_target = request.META.get("HTTP_REFERER") or reverse("hub:sqr")

        if request.user.role not in ADMIN_PANEL_ROLES:
            return JsonResponse({"error": "Only PM-ESG or Admin can create SQR revision Teams groups."}, status=403)

        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            return JsonResponse({"error": "Only the assigned PM-ESG approver can create the revision Teams group."}, status=403)

        if submission.status != SqrSubmission.Status.FOR_REVISION:
            return JsonResponse({"error": "Set SQR status to For Revision before creating a Teams group."}, status=400)

        approver_email = submission.pm_esg_reviewer.email if submission.pm_esg_reviewer and submission.pm_esg_reviewer.email else None
        requestor_email = submission.engineer.email if submission.engineer and submission.engineer.email else None
        if not approver_email or not requestor_email:
            return JsonResponse({"error": "Unable to create Teams group. Ensure both approver and requestor emails are configured."}, status=400)

        requestor_name = submission.engineer.get_full_name() or submission.engineer.username
        participants = ",".join(sorted({approver_email, requestor_email}))
        topic = f"SQR {submission.reference_code} {submission.customer_name}"
        raw_notes = (submission.review_notes or "").strip()
        if raw_notes:
            lines = [l.strip() for l in raw_notes.splitlines() if l.strip()]
            numbered = "\n".join(f"{i + 1}. {line}" for i, line in enumerate(lines))
        else:
            numbered = "1.\n2.\n3.\n4.\n5."
        message_body = (
            f"Hi @{requestor_name}\n"
            "Submitted SQR is for revision, please refer to the ff. comments below.\n\n"
            f"Comments\n{numbered}\n\n"
            "Thanks"
        )
        teams_url = (
            "https://teams.microsoft.com/l/chat/0/0?users="
            f"{quote(participants)}&topicName={quote(topic)}&message={quote(message_body)}"
        )

        return JsonResponse({"teams_url": teams_url})


class SqrApprovalOutlookRedirectView(LoginRequiredMixin, View):
    """Launch the default mail app compose window for the approved SQR draft."""

    def _get_submission(self, pk):
        return get_object_or_404(
            SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer", "linked_request"),
            pk=pk,
        )

    def _validate_access(self, request, submission):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return "Only PM-ESG or Admin can draft SQR approval advisory emails."
        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            return "Only the assigned PM-ESG approver can draft this approval advisory email."
        if submission.status != SqrSubmission.Status.APPROVED:
            return "Set SQR status to Approved before drafting the advisory email."
        requestor_email = submission.engineer.email if submission.engineer and submission.engineer.email else None
        if not requestor_email:
            return "Unable to draft an email. Ensure the requestor (engineer) has an email configured."
        return ""

    def get(self, request, pk):
        submission = self._get_submission(pk)
        redirect_target = reverse("hub:sqr")
        validation_error = self._validate_access(request, submission)
        if validation_error:
            messages.error(request, validation_error)
            return redirect(redirect_target)

        context = _get_sqr_approval_email_context(submission)
        return HttpResponseRedirect(_build_sqr_approval_mailto_url(context))

    def post(self, request, pk):
        return self.get(request, pk)


class SqrApprovalEmlDownloadView(LoginRequiredMixin, View):
    """Download a formatted Outlook-compatible EML fallback for approved SQR quotations."""

    def _get_submission(self, pk):
        return get_object_or_404(
            SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer", "linked_request"),
            pk=pk,
        )

    def _validate_access(self, request, submission):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return "Only PM-ESG or Admin can draft SQR approval advisory emails."
        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            return "Only the assigned PM-ESG approver can draft this approval advisory email."
        if submission.status != SqrSubmission.Status.APPROVED:
            return "Set SQR status to Approved before drafting the advisory email."
        requestor_email = submission.engineer.email if submission.engineer and submission.engineer.email else None
        if not requestor_email:
            return "Unable to draft an email. Ensure the requestor (engineer) has an email configured."
        return ""

    def get(self, request, pk):
        submission = self._get_submission(pk)
        validation_error = self._validate_access(request, submission)
        if validation_error:
            messages.error(request, validation_error)
            return redirect("hub:sqr")

        sender_email = (getattr(request.user, "email", "") or "").strip() or settings.DEFAULT_FROM_EMAIL
        context = _get_sqr_approval_email_context(submission)
        mime_payload = _build_sqr_approval_mime_message(context, sender_email, mark_as_unsent=True)
        eml_bytes = base64.b64decode(mime_payload)

        filename_base = re.sub(r"[^A-Za-z0-9._-]+", "_", context["subject"]).strip("_") or submission.reference_code or "sqr-approval-draft"
        response = HttpResponse(eml_bytes, content_type="message/rfc822")
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.eml"'
        return response


class SqrApprovalFormattedDraftView(LoginRequiredMixin, View):
    """Open the default mail app with an editable draft for the approved SQR email."""

    def _get_submission(self, pk):
        return get_object_or_404(
            SqrSubmission.objects.select_related("engineer", "pm_esg_reviewer", "linked_request"),
            pk=pk,
        )

    def _validate_access(self, request, submission):
        if request.user.role not in ADMIN_PANEL_ROLES:
            return "Only PM-ESG or Admin can draft SQR approval advisory emails."
        if request.user.role == PM_ESG_ROLE and submission.pm_esg_reviewer_id != request.user.id:
            return "Only the assigned PM-ESG approver can draft this approval advisory email."
        if submission.status != SqrSubmission.Status.APPROVED:
            return "Set SQR status to Approved before drafting the advisory email."
        requestor_email = submission.engineer.email if submission.engineer and submission.engineer.email else None
        if not requestor_email:
            return "Unable to draft an email. Ensure the requestor (engineer) has an email configured."
        return ""

    def get(self, request, pk):
        submission = self._get_submission(pk)
        validation_error = self._validate_access(request, submission)
        if validation_error:
            messages.error(request, validation_error)
            return redirect("hub:sqr")

        context = _get_sqr_approval_email_context(submission)
        return HttpResponseRedirect(_build_sqr_approval_mailto_url(context))


class RequestDetailView(LoginRequiredMixin, DetailView):
    model = Request
    template_name = "hub/request_detail.html"
    context_object_name = "request_obj"

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        if user.role == PM_ESS_ROLE:
            return qs.filter(Q(requestor=user) | Q(requestor__role=User.Roles.REQUESTOR_ESS))
        if user.role in REQUESTOR_ROLES:
            return qs.filter(requestor=user)
        if user.role in ENGINEER_ACCESS_ROLES:
            return qs.filter(engineer=user)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request_obj = context["request_obj"]
        context["status_logs"] = request_obj.status_logs.select_related("author")
        context.update(get_request_activity_log_context(request_obj))
        can_comment = self._user_can_comment(self.request.user, request_obj)
        context["can_comment"] = can_comment
        if can_comment:
            context["log_form"] = kwargs.get("log_form") or StatusLogForm()
        if (
            self.request.user.role in ENGINEER_ACCESS_ROLES
            and request_obj.engineer_id == self.request.user.id
        ):
            context["status_form"] = kwargs.get("status_form") or RequestStatusForm(instance=request_obj)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not self._user_can_comment(request.user, self.object):
            return redirect("hub:request-manage-collab", pk=self.object.pk)
        form = StatusLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.request = self.object
            log.author = request.user
            log.save()
            notify_status_update(log, "Request Detail · Status Update")
            messages.success(request, "Status log saved.")
            return redirect("hub:request-manage-collab", pk=self.object.pk)
        context = self.get_context_data(log_form=form)
        return self.render_to_response(context)

    @staticmethod
    def _user_can_comment(user, request_obj):
        if not user.is_authenticated:
            return False
        if user.role in ADMIN_PANEL_ROLES:
            return True
        if user.role in ENGINEER_ACCESS_ROLES and request_obj.engineer_id == user.id:
            return True
        if user.role in REQUEST_CREATOR_ROLES and request_obj.requestor_id == user.id:
            return True
        return False


class RequestAdminUpdateView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Request
    form_class = RequestAdminForm
    template_name = "hub/request_manager_form.html"
    context_object_name = "service_request"
    success_url = reverse_lazy("hub:dashboard")

    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, "That request no longer exists.")
            return redirect("hub:dashboard")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["allow_capacity_override"] = self.request.method == "POST" and self.request.POST.get("override_capacity") == "1"
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.POST.get("form_type") == "status_log":
            return self._handle_status_log_post(request)
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance._actor_user = self.request.user
        form.instance._actor_source = "Admin · Manage Request"
        original = Request.objects.get(pk=form.instance.pk)
        previous_engineer_id = original.engineer_id
        previous_backup_id = original.backup_engineer_id
        changed_fields = list(form.changed_data)
        response = super().form_valid(form)
        clear_engineer_outlook_lock_on_reassignment(
            self.object,
            previous_engineer_id=previous_engineer_id,
        )
        if changed_fields:
            self._notify_request_update(original, self.object, changed_fields)
        assignment_email_result = notify_engineer_assignment_email(
            self.object,
            actor_user=self.request.user,
            request=self.request,
            previous_engineer_id=previous_engineer_id,
            previous_backup_id=previous_backup_id,
        )
        notify_engineer_assignment_notification(
            self.object,
            actor_user=self.request.user,
            previous_engineer_id=previous_engineer_id,
            previous_backup_id=previous_backup_id,
        )
        messages.success(self.request, "Request updated.")
        flash_assignment_email_feedback(self.request, assignment_email_result, action_label="updated")
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        referer = self.request.META.get("HTTP_REFERER")
        fallback = reverse("hub:dashboard")
        if not referer or referer == self.request.build_absolute_uri():
            context["back_url"] = fallback
        else:
            context["back_url"] = referer
        context["hide_sign_in_nav"] = True
        context["delete_url"] = reverse("hub:request-delete", args=[self.object.pk])
        context["status_logs"] = (
            StatusLog.objects.filter(request=self.object)
            .select_related("author")
            .order_by("-created_at")
        )
        context.setdefault("log_form", StatusLogForm())
        context["engineer_capacity_map"] = self._build_engineer_capacity_map()
        context["status_form"] = None
        context["status_allowed"] = False
        context["account_name_choices"] = []
        context["is_admin_form"] = True
        context.update(get_request_activity_log_context(self.object))
        return context

    def _handle_status_log_post(self, request):
        form = StatusLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.request = self.object
            log.author = request.user
            log.save()
            notify_status_update(log, "Admin · Manage Request · Status Update")
            messages.success(request, "Status update posted.")
            return HttpResponseRedirect(self.request.path)
        if form.errors:
            message_field = form.fields.get("message")
            if message_field:
                existing_classes = message_field.widget.attrs.get("class", "")
                if "is-invalid" not in existing_classes.split():
                    message_field.widget.attrs["class"] = (existing_classes + " is-invalid").strip()
        context = self.get_context_data(log_form=form)
        return self.render_to_response(context)

    def _notify_request_update(self, original, updated, changed_fields):
        actor = self.request.user.get_full_name() or self.request.user.username
        summary_text = summarize_request_changes(original, updated, changed_fields)

        if not summary_text:
            return

        create_change_status_log(updated, self.request.user, "Admin · Manage Request", summary_text)
        recipients = {}
        if updated.requestor:
            recipients[updated.requestor.pk] = updated.requestor
        if updated.engineer:
            recipients[updated.engineer.pk] = updated.engineer
        if updated.backup_engineer:
            recipients[updated.backup_engineer.pk] = updated.backup_engineer

        recipients.pop(self.request.user.pk, None)

        for recipient in recipients.values():
            Notification.objects.create(
                recipient=recipient,
                message=f"{actor} updated {updated.reference_code}: {summary_text}",
                related_request=updated,
                actor=actor,
                source="Admin · Manage Request",
            )

    def _build_engineer_capacity_map(self):
        data = {}
        engineers = User.objects.filter(role__in=ASSIGNABLE_ENGINEER_ROLES)
        for engineer in engineers:
            assigned = Request.objects.filter(engineer=engineer, status=Request.Status.ONGOING)
            if self.object.pk:
                assigned = assigned.exclude(pk=self.object.pk)
            has_deployment = assigned.filter(engagement_type__in=[
                Request.Engagement.DEPLOYMENT, Request.Engagement.CERTIFICATION
            ]).exists()
            capacity = 3 if has_deployment else 5
            data[str(engineer.pk)] = {
                "name": engineer.get_full_name() or engineer.username or "Engineer",
                "load": assigned.count(),
                "capacity": capacity,
            }
        return data


class RequestUpdateView(LoginRequiredMixin, UpdateView):
    model = Request
    form_class = RequestForm
    template_name = "hub/request_update.html"
    success_url = reverse_lazy("hub:dashboard")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.setdefault("actor_role", self.request.user.role)
        return kwargs

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in REQUEST_CREATOR_ROLES:
            return redirect("hub:dashboard")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(requestor=self.request.user)
            .select_related("account", "engineer")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["account_name_choices"] = context["form"].account_name_suggestions
        context["is_edit"] = True
        return context

    def form_valid(self, form):
        form.instance._actor_user = self.request.user
        form.instance._actor_source = "Requestor · Edit Request"
        original = Request.objects.get(pk=form.instance.pk)
        previous_engineer_id = original.engineer_id
        changed_fields = normalize_request_form_changed_fields(form.changed_data)
        response = super().form_valid(form)
        clear_engineer_outlook_lock_on_reassignment(
            self.object,
            previous_engineer_id=previous_engineer_id,
        )
        if original.account_id != self.object.account_id and "account" not in changed_fields:
            changed_fields.append("account")
        if changed_fields:
            notify_account_manager_request_update(
                self.request.user,
                original,
                self.object,
                changed_fields,
                "Requestor · Edit Request",
            )
        messages.success(self.request, "Request updated.")
        return response


class RequestCollaborativeManageView(LoginRequiredMixin, View):
    template_name = "hub/request_manager_form.html"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())

        allowed = ADMIN_PANEL_ROLES | REQUEST_CREATOR_ROLES | ENGINEER_ACCESS_ROLES | {PM_ESS_ROLE}
        if request.user.role not in allowed:
            messages.error(request, "You are not allowed to manage this request.")
            return redirect("hub:dashboard")
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        queryset = Request.objects.select_related("account", "engineer", "backup_engineer", "requestor")
        pk = self.kwargs["pk"]
        user = self.request.user

        # Admins and PM-ESG can view any request
        if user.role in ADMIN_PANEL_ROLES:
            return get_object_or_404(queryset, pk=pk)

        if user.role == PM_ESS_ROLE:
            queryset = queryset.filter(pk=pk).filter(Q(requestor__role=User.Roles.REQUESTOR_ESS) | Q(requestor=user))
        elif user.role in REQUEST_CREATOR_ROLES:
            # Requestors can only see their own requests
            queryset = queryset.filter(pk=pk, requestor=user)
        elif user.role in ENGINEER_ACCESS_ROLES:
            # Engineers can see requests they are assigned to directly, or requests
            # linked to SQR submissions they own. The linked-SQR case handles
            # historical rows where the SQR requester and request assignee differ.
            queryset = queryset.filter(pk=pk).filter(
                Q(engineer=user) | Q(backup_engineer=user) | Q(sqr_links__engineer=user)
            ).distinct()
        else:
            raise Http404

        return get_object_or_404(queryset)

    def get(self, request, *args, **kwargs):
        try:
            request_obj = self.get_object()
        except Http404:
            messages.error(request, "You are not allowed to manage this request.")
            return redirect("hub:dashboard")
        context = self.get_context_data(request_obj)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        try:
            request_obj = self.get_object()
        except Http404:
            messages.error(request, "You are not allowed to manage this request.")
            return redirect("hub:dashboard")
        form_type = request.POST.get("form_type", "details")
        if form_type == "status_log":
            return self._handle_status_log_post(request, request_obj)
        if form_type == "status":
            return self._handle_status_update(request, request_obj)
        return self._handle_details_update(request, request_obj)

    def _actor_prefix(self) -> str:
        return "Requestor" if self.request.user.role in REQUEST_CREATOR_ROLES or self.request.user.role == PM_ESS_ROLE else "Engineer"

    def _source_label(self, suffix: str) -> str:
        return f"{self._actor_prefix()} · {suffix}"

    def get_context_data(self, request_obj, form=None, status_form=None, log_form=None):
        if form is None:
            form = RequestForm(instance=request_obj, actor_role=self.request.user.role, actor_user=self.request.user)
        status_allowed = self.request.user.role in ENGINEER_ACCESS_ROLES
        if status_allowed and status_form is None:
            status_form = RequestStatusForm(instance=request_obj)
        elif not status_allowed:
            status_form = None
        if log_form is None:
            log_form = StatusLogForm()

        referer = self.request.META.get("HTTP_REFERER")
        fallback = reverse("hub:dashboard")
        if not referer or referer == self.request.build_absolute_uri():
            back_url = fallback
        else:
            back_url = referer

        linked_sqr = SqrSubmission.objects.filter(linked_request=request_obj).first()

        return {
            "object": request_obj,
            "form": form,
            "status_form": status_form,
            "log_form": log_form,
            "status_logs": request_obj.status_logs.select_related("author").order_by("-created_at"),
            "account_name_choices": getattr(form, "account_name_suggestions", ()),
            "back_url": back_url,
            "status_allowed": status_allowed,
            "linked_sqr": linked_sqr,
            **get_request_activity_log_context(request_obj),
        }

    def _handle_details_update(self, request, request_obj):
        form = RequestForm(request.POST, instance=request_obj, actor_role=request.user.role, actor_user=request.user)
        if form.is_valid():
            source_label = self._source_label("Manage Request")
            form.instance._actor_user = request.user
            form.instance._actor_source = source_label
            original = Request.objects.get(pk=request_obj.pk)
            previous_engineer_id = original.engineer_id
            previous_backup_id = original.backup_engineer_id
            changed_fields = normalize_request_form_changed_fields(form.changed_data)
            form.save()
            clear_engineer_outlook_lock_on_reassignment(
                request_obj,
                previous_engineer_id=previous_engineer_id,
            )
            if original.account_id != request_obj.account_id and "account" not in changed_fields:
                changed_fields.append("account")
            if changed_fields:
                notify_account_manager_request_update(
                    request.user,
                    original,
                    request_obj,
                    changed_fields,
                    source_label,
                )
            assignment_email_result = notify_engineer_assignment_email(
                request_obj,
                actor_user=request.user,
                request=request,
                previous_engineer_id=previous_engineer_id,
                previous_backup_id=previous_backup_id,
            )
            notify_engineer_assignment_notification(
                request_obj,
                actor_user=request.user,
                previous_engineer_id=previous_engineer_id,
                previous_backup_id=previous_backup_id,
            )
            messages.success(request, "Request details updated.")
            flash_assignment_email_feedback(request, assignment_email_result, action_label="updated")
            return HttpResponseRedirect(request.path)
        context = self.get_context_data(request_obj, form=form)
        return render(request, self.template_name, context)

    def _handle_status_update(self, request, request_obj):
        if request.user.role not in ENGINEER_ACCESS_ROLES:
            messages.error(request, "Only the assigned engineer can update the status.")
            return HttpResponseRedirect(request.path)
        status_form = RequestStatusForm(request.POST, instance=request_obj)
        if status_form.is_valid():
            send_closing_email = (request.POST.get("send_closing_email") or "").strip() in {"1", "true", "True", "on"}
            source_label = self._source_label("Manage Request · Status")
            original = Request.objects.get(pk=request_obj.pk)
            request_obj._actor_user = request.user
            request_obj._actor_source = source_label
            status_form.save()
            changed_fields = []
            if original.status != request_obj.status:
                changed_fields.append("status")
            if original.end_date != request_obj.end_date:
                changed_fields.append("end_date")
            if changed_fields:
                notify_account_manager_request_update(
                    request.user,
                    original,
                    request_obj,
                    changed_fields,
                    source_label,
                )
            if (
                send_closing_email
                and original.status != Request.Status.COMPLETED
                and request_obj.status == Request.Status.COMPLETED
            ):
                return RequestClosingOutlookRedirectView().post(request, request_obj.pk)
            messages.success(request, "Request status updated.")
            return HttpResponseRedirect(request.path)
        if status_form.errors:
            for field in status_form:
                if field.errors:
                    existing_classes = field.field.widget.attrs.get("class", "")
                    if "is-invalid" not in existing_classes.split():
                        field.field.widget.attrs["class"] = (existing_classes + " is-invalid").strip()
        context = self.get_context_data(request_obj, status_form=status_form)
        return render(request, self.template_name, context)

    def _handle_status_log_post(self, request, request_obj):
        log_form = StatusLogForm(request.POST)
        if log_form.is_valid():
            log = log_form.save(commit=False)
            log.request = request_obj
            log.author = request.user
            log.save()
            source_label = self._source_label("Manage Request · Status Update")
            notify_status_update(log, source_label)
            messages.success(request, "Status update posted.")
            return HttpResponseRedirect(request.path)
        if log_form.errors:
            message_field = log_form.fields.get("message")
            if message_field:
                existing_classes = message_field.widget.attrs.get("class", "")
                if "is-invalid" not in existing_classes.split():
                    message_field.widget.attrs["class"] = (existing_classes + " is-invalid").strip()
        context = self.get_context_data(request_obj, log_form=log_form)
        return render(request, self.template_name, context)


class StatusLogUpdateView(LoginRequiredMixin, UpdateView):
    model = StatusLog
    form_class = StatusLogForm
    template_name = "hub/status_log_form.html"

    def get_queryset(self):
        # Only allow the author to edit their own logs
        return super().get_queryset().filter(author=self.request.user)

    def get_success_url(self):
        request_obj = self.object.request
        user = self.request.user
        if user.role in ADMIN_PANEL_ROLES:
            return reverse("hub:request-manage", args=[request_obj.pk])
        # For engineers, PMs, and requestors
        return reverse("hub:request-manage-collab", args=[request_obj.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Determine back URL similarly to get_success_url
        request_obj = self.object.request
        user = self.request.user
        if user.role in ADMIN_PANEL_ROLES:
            context["back_url"] = reverse("hub:request-manage", args=[request_obj.pk])
        else:
            context["back_url"] = reverse("hub:request-manage-collab", args=[request_obj.pk])
        return context


class RequestDeleteView(LoginRequiredMixin, DeleteView):
    model = Request
    success_url = reverse_lazy("hub:dashboard")
    template_name = "hub/request_confirm_delete.html"

    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, "That request no longer exists.")
            return redirect("hub:dashboard")

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role in ADMIN_PANEL_ROLES:
            return qs
        if user.role in REQUEST_CREATOR_ROLES:
            return qs.filter(requestor=user)
        return qs.none()

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        Request.objects.filter(pk=obj.pk).update(is_deleted=True, deleted_at=timezone.now())
        messages.success(request, f"Request {obj.reference_code} deleted. You can restore it from Profile → Backup &amp; Restore.")
        return redirect(self.success_url)


class RequestRestoreView(LoginRequiredMixin, View):
    """POST → restore a soft-deleted request (admin only)."""

    def post(self, request, pk):
        if request.user.role != User.Roles.ADMIN:
            messages.error(request, "Access denied.")
            return redirect("accounts:update")
        updated = Request.all_objects.filter(pk=pk, is_deleted=True).update(
            is_deleted=False, deleted_at=None
        )
        if updated:
            ref = Request.all_objects.filter(pk=pk).values_list("reference_code", flat=True).first()
            messages.success(request, f"Request {ref} has been restored.")
        else:
            messages.error(request, "Request not found or already active.")
        return redirect(str(reverse_lazy("accounts:update")) + "#backup")


class RequestTeamsRedirectView(AdminOrEngineerRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(Request.objects.select_related("engineer", "requestor"), pk=pk)
        teams_url = request_obj.teams_chat_url

        if not teams_url:
            messages.error(
                request,
                "Unable to start a Teams chat. Ensure the engineer and requestor both have emails configured.",
            )
            return redirect("hub:dashboard")

        messages.info(request, "Opening Microsoft Teams in a new tab…")
        return redirect(teams_url)


class RequestOutlookRedirectView(AdminOrEngineerRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(
            Request.objects.select_related("engineer", "backup_engineer", "requestor", "account"),
            pk=pk,
        )

        redirect_target = request.META.get("HTTP_REFERER") or reverse("hub:dashboard")

        if request.user.role in ENGINEER_ACCESS_ROLES:
            if request.user != request_obj.engineer and request.user != request_obj.backup_engineer:
                messages.error(request, "You are not allowed to draft emails for this request.")
                return redirect(redirect_target)
            already_launched = RequestCommunication.objects.filter(
                request=request_obj,
                user__role__in=ENGINEER_ACCESS_ROLES,
                channel=RequestCommunication.Channel.OUTLOOK,
            ).exists()
            if already_launched:
                messages.warning(request, "You already launched the Outlook draft for this request.")
                return redirect(redirect_target)

        engineer_email = (
            request_obj.engineer.email if request_obj.engineer and request_obj.engineer.email else None
        )
        manager_email = (
            request_obj.requestor.email if request_obj.requestor and request_obj.requestor.email else None
        )
        backup_email = (
            request_obj.backup_engineer.email
            if request_obj.backup_engineer and request_obj.backup_engineer.email
            else None
        )

        if not engineer_email or not manager_email:
            messages.error(
                request,
                "Unable to draft an email. Ensure the engineer and requestor both have emails configured.",
            )
            return redirect(redirect_target)

        to_addresses = {engineer_email, manager_email}
        cc_addresses = {"ESGRequestHub@phildata.com"}
        if request_obj.requestor and request_obj.requestor.role == User.Roles.REQUESTOR_ESS:
            cc_addresses.add("JoanI@phildata.com")
        if backup_email:
            cc_addresses.add(backup_email)

        recipients = ",".join(sorted(to_addresses))
        cc_field = ",".join(sorted(cc_addresses))
        account_name = request_obj.account.name if request_obj.account else "Request"
        subject = quote(f"Re: {request_obj.reference_code} · {account_name}")

        requestor = request_obj.requestor
        if requestor:
            requestor_name = requestor.get_full_name() or requestor.username
        else:
            requestor_name = request_obj.account_manager or "Requestor"

        engagement_display = request_obj.get_engagement_type_display()
        product_display = request_obj.get_product_category_display()

        description_clean = (request_obj.description or "").strip()
        if description_clean:
            description_line = description_clean.replace("\r", " ").replace("\n", " ")
        else:
            description_line = "No description provided."

        sender_name = request.user.get_full_name() or request.user.username
        assigned_engineer_name = ""
        if request_obj.engineer:
            assigned_engineer_name = request_obj.engineer.get_full_name() or request_obj.engineer.username or ""
        elif request_obj.backup_engineer:
            assigned_engineer_name = request_obj.backup_engineer.get_full_name() or request_obj.backup_engineer.username or ""
        if not assigned_engineer_name:
            assigned_engineer_name = "our engineering team"

        if request.user.role in ENGINEER_ACCESS_ROLES:
            body_template = (
                "Hello {requestor_name},\n\n"
                "This is to acknowledge your request in Request Hub. We've logged the details below and started processing it.\n\n"
                "Reference: {reference}\n"
                "Request Type: {request_type}\n"
                "Product: {product}\n"
                "Description: {description}\n\n"
                "I will reach out with updates or any follow-up questions.\n"
                "If you have additional information or questions, simply reply to this email and we'll continue the thread.\n"
            )
        else:
            body_template = (
                "Hello {requestor_name},\n\n"
                "This is to acknowledge your request in Request Hub. We've logged the details below and started processing it\n\n"
                "Reference: {reference}\n"
                "Request Type: {request_type}\n"
                "Product: {product}\n"
                "Description: {description}\n\n"
                "{assigned_engineer} is looped in and will reach out with updates or any follow-up questions.\n"
                "If you have additional information or questions, simply reply to this email and we'll continue the thread.\n"
            )

        body = quote(
            body_template.format(
                requestor_name=requestor_name,
                reference=request_obj.reference_code,
                request_type=engagement_display,
                product=product_display,
                description=description_line,
                sender_name=sender_name,
                assigned_engineer=assigned_engineer_name,
            )
        )

        RequestCommunication.objects.create(
            request=request_obj,
            user=request.user,
            channel=RequestCommunication.Channel.OUTLOOK,
        )

        mailto_parts = [f"mailto:{recipients}", f"subject={subject}", f"body={body}"]
        if cc_field:
            mailto_parts.insert(1, f"cc={quote(cc_field)}")

        outlook_url = "?".join([mailto_parts[0], "&".join(mailto_parts[1:])])
        messages.info(request, "Drafting email in your default mail client…")
        return render(
            request,
            "hub/outlook_redirect.html",
            {
                "mailto_url": outlook_url,
                "launch_url": outlook_url,
                "formatted_draft_url": "",
                "return_url": redirect_target,
            },
        )


class RequestClosingOutlookRedirectView(AdminOrEngineerRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(
            Request.objects.select_related("engineer", "backup_engineer", "requestor", "account"),
            pk=pk,
        )

        redirect_target = request.META.get("HTTP_REFERER") or reverse("hub:dashboard")

        if request_obj.status != Request.Status.COMPLETED:
            messages.error(request, "Mark the request as completed before sending a closing email.")
            return redirect(redirect_target)

        if request.user.role in ENGINEER_ACCESS_ROLES:
            if request.user != request_obj.engineer and request.user != request_obj.backup_engineer:
                messages.error(request, "You are not allowed to close out this request.")
                return redirect(redirect_target)

        engineer_email = None
        if request_obj.engineer and request_obj.engineer.email:
            engineer_email = request_obj.engineer.email
        elif request.user.email:
            engineer_email = request.user.email

        manager_email = request_obj.requestor.email if request_obj.requestor and request_obj.requestor.email else None
        backup_email = (
            request_obj.backup_engineer.email
            if request_obj.backup_engineer and request_obj.backup_engineer.email
            else None
        )

        if not manager_email:
            messages.error(request, "Unable to draft a closing email. Ensure the requestor has an email configured.")
            return redirect(redirect_target)

        if not engineer_email:
            messages.error(request, "Unable to draft a closing email. Ensure the engineer has an email configured.")
            return redirect(redirect_target)

        to_addresses = {manager_email, engineer_email}
        cc_addresses = {"ESGRequestHub@phildata.com"}
        if request_obj.requestor and request_obj.requestor.role == User.Roles.REQUESTOR_ESS:
            cc_addresses.add("JoanI@phildata.com")
        if backup_email:
            cc_addresses.add(backup_email)

        recipients = ",".join(sorted(addr for addr in to_addresses if addr))
        cc_field = ",".join(sorted(cc_addresses))

        # Match the acknowledgement subject exactly so Outlook can group the request emails together.
        account_name = request_obj.account.name if request_obj.account else "Request"
        ack_subject = f"Re: {request_obj.reference_code} · {account_name}"
        subject = quote(ack_subject)

        requestor = request_obj.requestor
        if requestor:
            requestor_name = requestor.get_full_name() or requestor.username
        else:
            requestor_name = request_obj.account_manager or "Requestor"

        description_clean = (request_obj.description or "").strip()
        if description_clean:
            description_line = description_clean.replace("\r", " ").replace("\n", " ")
        else:
            description_line = "No description provided."

        detail_url = request.build_absolute_uri(request_obj.get_absolute_url())

        body_template = (
            "Hello {requestor_name},\n\n"
            "Following up on our earlier acknowledgement for {reference}, this is to confirm the request has been fulfilled and is now marked as closed.\n\n"
            "Advisory Only: Please do not reply to this closed request thread for new support needs.\n\n"
            "View request details: {detail_url}\n\n"
            "If you believe further action is required or have additional questions, please submit a new one via Request Hub.\n\n"
            "Thank you for your cooperation."
        )

        body = quote(
            body_template.format(
                requestor_name=requestor_name,
                description=description_line,
                reference=request_obj.reference_code,
                detail_url=detail_url,
            )
        )

        RequestCommunication.objects.create(
            request=request_obj,
            user=request.user,
            channel=RequestCommunication.Channel.OUTLOOK,
        )

        mailto_parts = [f"mailto:{recipients}", f"subject={subject}", f"body={body}"]
        if cc_field:
            mailto_parts.insert(1, f"cc={quote(cc_field)}")

        outlook_url = "?".join([mailto_parts[0], "&".join(mailto_parts[1:])])
        messages.info(request, "Drafting closing email in your default mail client…")
        return render(
            request,
            "hub/outlook_redirect.html",
            {
                "mailto_url": outlook_url,
                "launch_url": outlook_url,
                "formatted_draft_url": "",
                "return_url": redirect_target,
            },
        )


class RequestTeamsRedirectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(
            Request.objects.select_related("engineer", "backup_engineer", "requestor", "account"),
            pk=pk,
        )

        redirect_target = request.META.get("HTTP_REFERER") or reverse("hub:dashboard")

        role = request.user.role
        if role not in ({User.Roles.ADMIN, PM_ESG_ROLE, PM_ESS_ROLE} | ENGINEER_ACCESS_ROLES):
            messages.error(request, "You are not allowed to start a Teams chat for this request.")
            return redirect(redirect_target)

        if role in ENGINEER_ACCESS_ROLES:
            if request.user != request_obj.engineer and request.user != request_obj.backup_engineer:
                messages.error(request, "You are not allowed to start a Teams chat for this request.")
                return redirect(redirect_target)
            already_launched = RequestCommunication.objects.filter(
                request=request_obj,
                user=request.user,
                channel=RequestCommunication.Channel.TEAMS,
            ).exists()
            if already_launched:
                messages.warning(request, "You already launched the Teams chat for this request.")
                return redirect(redirect_target)

        if role == PM_ESS_ROLE:
            if not request_obj.requestor or request_obj.requestor.role != User.Roles.REQUESTOR_ESS:
                messages.error(request, "Teams chat is only available for Requestor-ESS requests.")
                return redirect(redirect_target)
            if not request_obj.engineer or not request_obj.engineer.email:
                messages.error(request, "Assign an engineer with an email before starting a Teams chat.")
                return redirect(redirect_target)
            if not request_obj.requestor.email:
                messages.error(request, "Requestor email is missing; unable to start a Teams chat.")
                return redirect(redirect_target)
            already_launched = RequestCommunication.objects.filter(
                request=request_obj,
                user=request.user,
                channel=RequestCommunication.Channel.TEAMS,
            ).exists()
            if already_launched:
                messages.warning(request, "You already launched the Teams chat for this request.")
                return redirect(redirect_target)

            topic = request_obj._build_teams_chat_topic(reference_code=request_obj.reference_code)
            participants = ",".join(sorted({email for email in [request_obj.engineer.email, request_obj.requestor.email] if email}))
            if not participants:
                messages.error(request, "Unable to start a Teams chat. Missing participant emails.")
                return redirect(redirect_target)
            teams_url = (
                "https://teams.microsoft.com/l/chat/0/0?users="
                f"{quote(participants)}&topicName={quote(topic)}"
            )
        else:
            teams_url = request_obj.teams_chat_url

        if not teams_url:
            messages.error(
                request,
                "Unable to start a Teams chat. Ensure the engineer and requestor have email addresses configured.",
            )
            return redirect(redirect_target)

        RequestCommunication.objects.create(
            request=request_obj,
            user=request.user,
            channel=RequestCommunication.Channel.TEAMS,
        )

        messages.info(request, "Launching Microsoft Teams…")
        return render(
            request,
            "hub/teams_redirect.html",
            {"teams_url": teams_url},
        )


class RequestExportCSVView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, View):
    """Allow administrators to export all requests to a CSV download."""

    columns = (
        "Reference",
        "Account",
        "Requestor",
        "Requestor Email",
        "Engineer",
        "Engineer Email",
        "Priority",
        "Status",
        "Engagement",
        "Start Date",
        "Due Date",
        "End Date",
        "Days",
        "Description",
        "Created",
        "Updated",
    )

    def get(self, request, *args, **kwargs):
        timestamp = timezone.now().strftime("%Y%m%d-%H%M%S")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="requests-{timestamp}.csv"'

        writer = csv.writer(response)
        writer.writerow(self.columns)

        queryset = Request.objects.select_related("account", "requestor", "engineer").order_by("reference_code")
        for req in queryset:
            requestor = req.requestor
            engineer = req.engineer
            writer.writerow(
                [
                    req.reference_code,
                    req.account.name if req.account else "",
                    requestor.get_full_name() or requestor.username if requestor else "",
                    requestor.email if requestor else "",
                    engineer.get_full_name() or engineer.username if engineer else "",
                    engineer.email if engineer else "",
                    req.get_priority_display(),
                    req.get_status_display(),
                    req.get_engagement_type_display(),
                    req.start_date.strftime("%Y-%m-%d") if req.start_date else "",
                    req.due_date.strftime("%Y-%m-%d") if req.due_date else "",
                    req.end_date.strftime("%Y-%m-%d") if req.end_date else "",
                    req.days_since_creation,
                    (req.description or "").replace("\r\n", " ").replace("\n", " "),
                    req.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    req.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        return response


class SqrExportView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, View):
    """Export all SQR submissions to an Excel (.xlsx) file."""

    def get(self, request, *args, **kwargs):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        qs = SqrSubmission.objects.select_related(
            "engineer", "pm_esg_reviewer", "reviewed_by",
            "assigned_pm", "assigned_sse", "linked_request",
        ).order_by("-created_at", "-pk")

        report_type = request.GET.get("report", "sqr")
        if report_type not in {"sqr", "esg"}:
            report_type = "sqr"

        report_config = {
            "sqr": {
                "worksheet_title": "SQR",
                "filename_prefix": "sqr-export",
                "columns": _get_sqr_export_columns,
            },
            "esg": {
                "worksheet_title": "ESG Performance Tracker",
                "filename_prefix": "esg-performance-tracker",
                "columns": _get_sqr_esg_export_columns,
            },
        }[report_type]

        wb = Workbook()
        ws = wb.active
        ws.title = report_config["worksheet_title"]

        # ── Styles ──────────────────────────────────────────────
        hdr_font  = Font(bold=True, color="FFFFFF", size=9)
        hdr_fill  = PatternFill("solid", fgColor="1A1F2E")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        columns = report_config["columns"]()

        # ── Header row ──────────────────────────────────────────
        ws.row_dimensions[1].height = 36
        for col_idx, (label, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

        # ── Data rows ───────────────────────────────────────────
        for row_idx, submission in enumerate(qs, start=2):
            ws.row_dimensions[row_idx].height = 18
            for col_idx, (_, extractor) in enumerate(columns, start=1):
                try:
                    val = extractor(submission)
                except Exception:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = cell_align
                cell.border = border

        # ── Auto-fit column widths (header text drives minimum) ─
        for col_idx, (label, _) in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(label)
            for row_idx in range(2, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v:
                    max_len = max(max_len, min(len(str(v)), 40))
            ws.column_dimensions[col_letter].width = max_len + 2

        # ── Freeze top row ──────────────────────────────────────
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        timestamp = timezone.now().strftime("%Y%m%d-%H%M%S")
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename_prefix = report_config["filename_prefix"]
        response["Content-Disposition"] = f'attachment; filename="{filename_prefix}-{timestamp}.xlsx"'
        return response


class SqrImportView(AdminRequiredMixin, LoginRequiredMixin, View):
    """Import SQR data from Excel and overwrite the current SQR table."""

    EXPECTED_HEADERS = [label for label, _ in _get_sqr_export_columns()]
    STATUS_MAP = _build_choice_import_map(SqrSubmission.Status.choices)
    PROPOSAL_STATUS_MAP = _build_choice_import_map(SqrSubmission.ProposalStatus.choices)
    OVERALL_STATUS_MAP = _build_choice_import_map(SqrSubmission.OverallStatus.choices)
    DELIVERY_HEALTH_MAP = _build_choice_import_map(SqrSubmission.DeliveryHealth.choices)
    REVENUE_DECLARATION_MAP = _build_choice_import_map(SqrSubmission.RevenueDeclaration.choices)
    IMPORT_STATUS_TIMEOUT = 60 * 60
    MANAGED_SCOPES = frozenset([
        "Deployment Only", "Deployment and Project Management", "Maintenance", "On-Call Services", "Implementation", "Implementation and Project Management", "Managed Support and Maintenance Service",
    ])

    def post(self, request, *args, **kwargs):
        form = SqrImportForm(request.POST, request.FILES)
        if not form.is_valid():
            error_list = form.errors.get("import_file", form.non_field_errors())
            return JsonResponse(
                {
                    "ok": False,
                    "errors": [str(error) for error in error_list],
                },
                status=400,
            )

        should_overwrite = str(request.POST.get("confirm_overwrite", "")).strip().lower()
        if should_overwrite not in {"1", "true", "yes", "on"}:
            return JsonResponse(
                {
                    "ok": False,
                    "errors": ["Please confirm that the current SQR table will be replaced before importing."],
                },
                status=400,
            )

        upload = form.cleaned_data["import_file"]
        try:
            file_bytes = upload.read()
        except Exception:
            return JsonResponse(
                {
                    "ok": False,
                    "errors": ["The selected file could not be read. Please upload a valid Excel file."],
                },
                status=400,
            )

        if not file_bytes:
            return JsonResponse(
                {
                    "ok": False,
                    "errors": ["The selected file is empty."],
                },
                status=400,
            )

        job_id = uuid.uuid4().hex
        self._set_status(
            request.user,
            job_id,
            {
                "state": "queued",
                "percent": 0,
                "processed": 0,
                "total": 0,
                "message": "Preparing import…",
                "errors": [],
                "result": None,
            },
        )

        worker = threading.Thread(
            target=self._run_import_job,
            kwargs={
                "job_id": job_id,
                "user_id": request.user.pk,
                "file_bytes": file_bytes,
            },
            daemon=True,
        )
        worker.start()

        return JsonResponse({
            "ok": True,
            "job_id": job_id,
            "message": "Import started.",
        })

    def get(self, request, *args, **kwargs):
        job_id = (request.GET.get("job_id") or "").strip()
        if not job_id:
            return JsonResponse({"ok": False, "error": "Missing import job id."}, status=400)
        status = self._get_status(request.user, job_id)
        if status is None:
            return JsonResponse({"ok": False, "error": "Import job not found."}, status=404)
        return JsonResponse({"ok": True, "status": status})

    @classmethod
    def _status_cache_key(cls, user_id, job_id):
        return f"sqr-import:{user_id}:{job_id}"

    @classmethod
    def _set_status(cls, user, job_id, payload):
        user_id = user.pk if hasattr(user, "pk") else user
        cache.set(cls._status_cache_key(user_id, job_id), payload, timeout=cls.IMPORT_STATUS_TIMEOUT)

    @classmethod
    def _get_status(cls, user, job_id):
        user_id = user.pk if hasattr(user, "pk") else user
        return cache.get(cls._status_cache_key(user_id, job_id))

    @classmethod
    def _update_status(cls, user_id, job_id, **changes):
        current = cache.get(cls._status_cache_key(user_id, job_id), {}) or {}
        current.update(changes)
        cache.set(cls._status_cache_key(user_id, job_id), current, timeout=cls.IMPORT_STATUS_TIMEOUT)

    def _run_import_job(self, *, job_id, user_id, file_bytes):
        close_old_connections()
        try:
            actor = User.objects.get(pk=user_id)
            row_payloads = self._parse_import_file(file_bytes=file_bytes, actor=actor, user_id=user_id, job_id=job_id)
            total = len(row_payloads)
            if total <= 0:
                raise ValueError("Import failed: the file does not contain any data rows.")

            self._update_status(
                user_id,
                job_id,
                state="running",
                total=total,
                processed=0,
                created=0,
                percent=5,
                message="Replacing current SQR table…",
            )

            created_count = 0
            with transaction.atomic():
                SqrSubmission.objects.all().delete()
                for index, payload in enumerate(row_payloads, start=1):
                    submission = SqrSubmission()
                    data = payload["data"].copy()
                    created_at_value = data.pop("created_at", None)
                    reference_code = data.pop("reference_code", "")
                    year = data.pop("year", None)
                    sequence_number = data.pop("sequence_number", None)

                    for field_name, field_value in data.items():
                        setattr(submission, field_name, field_value)

                    if year:
                        submission.year = year
                    if sequence_number:
                        submission.sequence_number = sequence_number
                    if reference_code:
                        submission.reference_code = reference_code

                    submission.save()

                    updates = {}
                    if created_at_value:
                        updates["created_at"] = created_at_value
                    if reference_code:
                        updates["reference_code"] = reference_code
                    if sequence_number:
                        updates["sequence_number"] = sequence_number
                    if year:
                        updates["year"] = year
                    if updates:
                        SqrSubmission.objects.filter(pk=submission.pk).update(**updates)

                    created_count += 1
                    percent = min(100, 5 + int((index / total) * 95))
                    self._update_status(
                        user_id,
                        job_id,
                        state="running",
                        processed=index,
                        created=created_count,
                        total=total,
                        percent=percent,
                        message=f"Imported {index} of {total} row(s)…",
                    )

            self._update_status(
                user_id,
                job_id,
                state="completed",
                processed=created_count,
                created=created_count,
                total=created_count,
                percent=100,
                message="Import completed successfully.",
                result={
                    "created": created_count,
                    "updated": 0,
                },
            )
        except Exception as exc:
            logger.exception("SQR import job failed", extra={"job_id": job_id, "user_id": user_id})
            self._update_status(
                user_id,
                job_id,
                state="failed",
                percent=0,
                message=str(exc) or "Import failed.",
                errors=[str(exc) or "Import failed."],
            )
        finally:
            close_old_connections()

    def _parse_import_file(self, *, file_bytes, actor, user_id, job_id):
        import io
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            raise ValueError("The selected file could not be read. Please upload a valid Excel file.") from exc

        request_map = {
            _normalize_import_text(item.reference_code): item
            for item in Request.objects.select_related("account", "engineer").all()
            if item.reference_code
        }
        pm_users = list(User.objects.filter(role=User.Roles.PM_ESG).order_by("first_name", "last_name", "username"))
        engineer_users = list(User.objects.filter(role__in=User.ENGINEER_ACCESS_ROLES).order_by("first_name", "last_name", "username"))
        pm_lookup = self._build_user_lookup(pm_users)
        engineer_lookup = self._build_user_lookup(engineer_users)

        start_row = self._detect_data_start_row(worksheet)
        candidate_rows = list(
            worksheet.iter_rows(
                min_row=start_row,
                max_row=worksheet.max_row,
                max_col=len(self.EXPECTED_HEADERS),
                values_only=True,
            )
        )
        non_blank_rows = [row for row in candidate_rows if not self._is_blank_row(row)]

        self._update_status(
            user_id,
            job_id,
            state="running",
            processed=0,
            total=len(non_blank_rows),
            percent=1,
            message="Reading Excel rows…",
            errors=[],
        )

        row_payloads = []
        row_errors = []
        total_rows = max(len(non_blank_rows), 1)
        processed_rows = 0
        for excel_row_number, row in enumerate(candidate_rows, start=start_row):
            if self._is_blank_row(row):
                continue
            row_data = {
                header: row[idx] if idx < len(row) else None
                for idx, header in enumerate(self.EXPECTED_HEADERS)
            }
            try:
                row_payloads.append(
                    self._prepare_row_payload(
                        row_data=row_data,
                        request_map=request_map,
                        pm_lookup=pm_lookup,
                        engineer_lookup=engineer_lookup,
                        actor=actor,
                    )
                )
            except ValueError as exc:
                row_errors.append(f"Row {excel_row_number}: {exc}")

            processed_rows += 1
            self._update_status(
                user_id,
                job_id,
                state="running",
                processed=processed_rows,
                total=len(non_blank_rows),
                percent=min(45, int((processed_rows / total_rows) * 45)),
                message=f"Validated {processed_rows} of {len(non_blank_rows)} row(s)…",
            )

        if not row_payloads and not row_errors:
            raise ValueError("Import failed: the file does not contain any data rows.")

        if row_errors:
            preview_errors = row_errors[:5]
            suffix = f" ...and {len(row_errors) - 5} more error(s)." if len(row_errors) > 5 else ""
            raise ValueError(f"Import failed with {len(row_errors)} row error(s). {' | '.join(preview_errors)}{suffix}")

        return row_payloads

    def _detect_data_start_row(self, worksheet):
        first_row = [self._stringify(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        normalized_first_row = [_normalize_import_text(value) for value in first_row]
        expected_prefix = [_normalize_import_text(label) for label in self.EXPECTED_HEADERS[:3]]
        if normalized_first_row[:3] == expected_prefix:
            return 2
        return 1

    @staticmethod
    def _stringify(value) -> str:
        return str(value or "").strip()

    @staticmethod
    def _is_blank_row(row) -> bool:
        return all(str(value or "").strip() == "" for value in row)

    @staticmethod
    def _build_user_lookup(users):
        lookup = {}
        for user in users:
            keys = {
                _normalize_import_text(user.username),
                _normalize_import_text(user.get_full_name()),
                _normalize_import_text(f"{user.first_name} {user.last_name}"),
            }
            for key in {item for item in keys if item}:
                lookup.setdefault(key, []).append(user)
        return lookup

    @staticmethod
    def _first_lookup_user(lookup):
        for users in lookup.values():
            if users:
                return users[0]
        return None

    def _parse_user(self, raw_value, lookup, label):
        normalized = _normalize_import_text(raw_value)
        if not normalized:
            return None
        matches = lookup.get(normalized, [])
        if not matches:
            raise ValueError(f"{label} was not found: {raw_value}")
        unique_matches = {user.pk: user for user in matches}
        if len(unique_matches) > 1:
            raise ValueError(f"{label} matches multiple users: {raw_value}")
        return next(iter(unique_matches.values()))

    def _parse_user_lenient(self, raw_value, lookup):
        normalized = _normalize_import_text(raw_value)
        if not normalized:
            return None
        matches = lookup.get(normalized, [])
        if not matches:
            return None
        unique_matches = {user.pk: user for user in matches}
        if len(unique_matches) > 1:
            return None
        return next(iter(unique_matches.values()))

    @staticmethod
    def _parse_choice(raw_value, mapping, label, default_blank=True):
        normalized = _normalize_import_text(raw_value)
        if not normalized:
            return "" if default_blank else None
        value = mapping.get(normalized)
        if value is None:
            raise ValueError(f"Invalid {label}: {raw_value}")
        return value

    @staticmethod
    def _parse_choice_lenient(raw_value, mapping, default_blank=True):
        normalized = _normalize_import_text(raw_value)
        if not normalized:
            return "" if default_blank else None
        return mapping.get(normalized, "" if default_blank else None)

    @staticmethod
    def _parse_decimal(raw_value, label):
        if raw_value in (None, ""):
            return None
        text_value = str(raw_value).replace(",", "").strip()
        if not text_value:
            return None
        try:
            return Decimal(text_value)
        except Exception as exc:
            raise ValueError(f"Invalid {label}: {raw_value}") from exc

    @classmethod
    def _parse_decimal_lenient(cls, raw_value, label):
        try:
            return cls._parse_decimal(raw_value, label)
        except ValueError:
            return None

    @staticmethod
    def _parse_int(raw_value, label):
        if raw_value in (None, ""):
            return None
        try:
            return int(Decimal(str(raw_value)))
        except Exception as exc:
            raise ValueError(f"Invalid {label}: {raw_value}") from exc

    @classmethod
    def _parse_int_lenient(cls, raw_value, label):
        try:
            return cls._parse_int(raw_value, label)
        except ValueError:
            return None

    @staticmethod
    def _parse_date(raw_value, label):
        if raw_value in (None, ""):
            return None
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value
        text_value = str(raw_value).strip()
        if not text_value:
            return None
        try:
            return date.fromisoformat(text_value)
        except Exception:
            pass
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text_value, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Invalid {label}: {raw_value}")

    @classmethod
    def _parse_date_lenient(cls, raw_value, label):
        try:
            return cls._parse_date(raw_value, label)
        except ValueError:
            return None

    @staticmethod
    def _to_datetime(value):
        if not value:
            return None
        return timezone.make_aware(datetime.combine(value, datetime.min.time()), MANILA_TZ)

    def _prepare_row_payload(self, *, row_data, request_map, pm_lookup, engineer_lookup, actor):
        sqr_id = self._stringify(row_data.get("SQR ID"))
        request_code = self._stringify(row_data.get("RQ ID"))
        linked_request = request_map.get(_normalize_import_text(request_code)) if request_code else None

        default_engineer = getattr(linked_request, "engineer", None) or self._first_lookup_user(engineer_lookup) or actor
        default_pm = self._first_lookup_user(pm_lookup) or actor

        engineer = self._parse_user_lenient(row_data.get("Requester Name"), engineer_lookup)
        pm_reviewer = self._parse_user_lenient(row_data.get("Approver Name"), pm_lookup)
        assigned_pm = self._parse_user_lenient(row_data.get("Assigned PM"), pm_lookup)
        assigned_sse = self._parse_user_lenient(row_data.get("Assigned SSE"), engineer_lookup)

        status = self._parse_choice_lenient(row_data.get("SQR Status"), self.STATUS_MAP) or SqrSubmission.Status.FOR_PROCESSING
        proposal_status = self._parse_choice_lenient(row_data.get("Proposal Status"), self.PROPOSAL_STATUS_MAP)
        overall_status = self._parse_choice_lenient(row_data.get("Overall Status"), self.OVERALL_STATUS_MAP)
        delivery_health = self._parse_choice_lenient(row_data.get("Health Status"), self.DELIVERY_HEALTH_MAP)
        revenue_declaration = self._parse_choice_lenient(row_data.get("Revenue Declaration"), self.REVENUE_DECLARATION_MAP)

        customer_name = self._stringify(row_data.get("Account Name"))
        project_title = self._stringify(row_data.get("Service Description"))
        project_details = self._stringify(row_data.get("Scope of Services"))
        sqr_folder_link = self._stringify(row_data.get("SQR Doc. Ref. Link"))

        sse_manhrs = self._parse_decimal_lenient(row_data.get("SSE Man-hrs"), "SSE Man-hrs")
        pm_manhrs = self._parse_decimal_lenient(row_data.get("PM Man-hrs"), "PM Man-hrs")
        discount_rate = self._parse_int_lenient(row_data.get("Discount Rate (%)"), "Discount Rate (%)")
        discount_rate = 0 if discount_rate is None else max(0, min(100, discount_rate))
        delivery_progress = self._parse_int_lenient(row_data.get("Overall Progress %"), "Overall Progress %")
        if delivery_progress is not None:
            delivery_progress = max(0, min(100, delivery_progress))

        reviewed_date_value = row_data.get("Approval Date", row_data.get("Reviewed Date"))
        approval_date = self._parse_date_lenient(reviewed_date_value, "Approval Date")
        validity_due_date = self._parse_date_lenient(row_data.get("Validity Due Date"), "Validity Due Date")
        created_date = self._parse_date_lenient(row_data.get("SQR Date"), "SQR Date")

        instance = SqrSubmission()
        reference_meta = self._parse_reference_code(sqr_id)
        data = {
            "linked_request": linked_request,
            "engineer": engineer or default_engineer,
            "pm_esg_reviewer": pm_reviewer or default_pm,
            "customer_name": customer_name or "Legacy Imported Record",
            "customer_company": self._stringify(row_data.get("Group Name")),
            "customer_contact": self._stringify(row_data.get("Account Manager")),
            "project_title": project_title or "Legacy Imported Service",
            "project_details": project_details or "Imported legacy data with incomplete details.",
            "sse_manhrs": sse_manhrs,
            "sqr_folder_link": sqr_folder_link,
            "discount_rate": discount_rate,
            "pm_manhrs": pm_manhrs,
            "proposal_status": proposal_status,
            "po_pnl_date": self._parse_date_lenient(row_data.get("PO / PNL Date"), "PO / PNL Date"),
            "assigned_pm": assigned_pm or pm_reviewer or default_pm,
            "assigned_sse": assigned_sse or engineer or default_engineer,
            "delivery_start_date": self._parse_date_lenient(row_data.get("Start Date"), "Start Date"),
            "delivery_target_finish_date": self._parse_date_lenient(row_data.get("Target Finish Date"), "Target Finish Date"),
            "overall_status": overall_status,
            "delivery_health": delivery_health,
            "delivery_progress": delivery_progress,
            "key_updates_risks": self._stringify(row_data.get("Key Updates / Risks")),
            "delivery_actual_finish_date": self._parse_date_lenient(row_data.get("Actual Finish Date"), "Actual Finish Date"),
            "delivery_completion_signed_date": self._parse_date_lenient(row_data.get("Completion Signed Date"), "Completion Signed Date"),
            "revenue_date": self._parse_date_lenient(
                row_data.get("Billed Date")
                or row_data.get("Date")
                or row_data.get("SI / Revenue Date"),
                "Billed Date",
            ),
            "revenue_source": self._stringify(
                row_data.get("Billing Type") or row_data.get("Source")
            ),
            "revenue_reference_no": self._stringify(
                row_data.get("Billing Reference") or row_data.get("Reference No.")
            ),
            "revenue_status": self._stringify(
                row_data.get("Billing Status") or row_data.get("Revenue Status")
            ),
            "revenue_remarks": self._stringify(
                row_data.get("PO Remarks") or row_data.get("Remarks")
            ),
            "revenue_overview": self._stringify(
                row_data.get("Billing Remarks") or row_data.get("Revenue Declaration")
            ),
            "revenue_declaration": revenue_declaration,
            "status": status,
            "documentation_links": "",
            "remarks": "",
            "reference_code": sqr_id,
            "year": reference_meta["year"],
            "sequence_number": reference_meta["sequence_number"],
        }

        if status == SqrSubmission.Status.APPROVED:
            reviewed_at = self._to_datetime(approval_date) or timezone.now()
            data["reviewed_at"] = reviewed_at
            data["reviewed_by"] = actor
            data["validity_due_date"] = validity_due_date or (reviewed_at.date() + timedelta(days=45))
        else:
            data["reviewed_at"] = None
            data["reviewed_by"] = None
            data["validity_due_date"] = validity_due_date

        scope = data["project_details"]
        group = data["customer_company"]
        data["managed_support_amount"] = (
            Decimal("149000.00") if group == "ESS" else Decimal("192000.00")
        ) if scope in self.MANAGED_SCOPES else None

        if created_date:
            data["created_at"] = self._to_datetime(created_date)

        return {"instance": instance, "data": data}

    @staticmethod
    def _parse_reference_code(reference_code):
        cleaned = str(reference_code or "").strip()
        if not cleaned:
            current_year = timezone.now().astimezone(MANILA_TZ).year
            return {"year": current_year, "sequence_number": None}
        match = re.fullmatch(r"SQR-(\d{4})-(\d{4})", cleaned, flags=re.IGNORECASE)
        if not match:
            raise ValueError(f"Invalid SQR ID format: {reference_code}")
        return {
            "year": int(match.group(1)),
            "sequence_number": int(match.group(2)),
        }


class RequestReportView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, TemplateView):
    template_name = "hub/report.html"

    @staticmethod
    def _normalize_report_view(value: str | None) -> str:
        report_view = (value or "operational").lower()
        if report_view not in {"operational", "activity"}:
            return "operational"
        return report_view

    def _get_requested_edit_log(self) -> Optional[EngineerActivityLog]:
        edit_log_id = (self.request.GET.get("edit_activity") or "").strip()
        if not edit_log_id:
            return None
        try:
            return EngineerActivityLog.objects.select_related("engineer", "account", "request").get(pk=edit_log_id)
        except (EngineerActivityLog.DoesNotExist, ValueError):
            messages.error(self.request, "We could not find that activity log to edit.")
            return None

    @staticmethod
    def _normalize_month_value(value: str | None) -> str:
        month_value = (value or "").strip()
        if not month_value:
            return ""
        try:
            datetime.strptime(month_value, "%Y-%m")
        except ValueError:
            return ""
        return month_value

    @staticmethod
    def _month_start(month_value: str | None) -> Optional[date]:
        if not month_value:
            return None
        parsed = datetime.strptime(month_value, "%Y-%m")
        return date(parsed.year, parsed.month, 1)

    @staticmethod
    def _next_month_start(month_start: date) -> date:
        if month_start.month == 12:
            return date(month_start.year + 1, 1, 1)
        return date(month_start.year, month_start.month + 1, 1)

    def _resolve_activity_month_filters(
        self,
        *,
        start_month_value: str | None = None,
        end_month_value: str | None = None,
    ) -> dict:
        start_month = self._normalize_month_value(
            self.request.GET.get("start_month") if start_month_value is None else start_month_value
        )
        end_month = self._normalize_month_value(
            self.request.GET.get("end_month") if end_month_value is None else end_month_value
        )

        start_month_date = self._month_start(start_month)
        end_month_date = self._month_start(end_month)

        if start_month_date and end_month_date and start_month_date > end_month_date:
            start_month, end_month = end_month, start_month
            start_month_date, end_month_date = end_month_date, start_month_date

        end_exclusive_date = self._next_month_start(end_month_date) if end_month_date else None

        if start_month_date and end_month_date:
            if start_month_date == end_month_date:
                label = start_month_date.strftime("%B %Y")
            else:
                label = f"{start_month_date.strftime('%B %Y')} to {end_month_date.strftime('%B %Y')}"
        elif start_month_date:
            label = f"{start_month_date.strftime('%B %Y')} onwards"
        elif end_month_date:
            label = f"Up to {end_month_date.strftime('%B %Y')}"
        else:
            label = "All months"

        return {
            "start_month": start_month,
            "end_month": end_month,
            "start_month_date": start_month_date,
            "end_exclusive_date": end_exclusive_date,
            "label": label,
        }

    @staticmethod
    def _build_activity_report_url(
        *,
        start_month: str = "",
        end_month: str = "",
        edit_activity: Optional[int] = None,
    ) -> str:
        params = {"report_view": "activity"}
        if start_month:
            params["start_month"] = start_month
        if end_month:
            params["end_month"] = end_month
        if edit_activity is not None:
            params["edit_activity"] = str(edit_activity)
        return f"{reverse('hub:report')}?{urlencode(params)}"

    def post(self, request, *args, **kwargs):
        report_view = self._normalize_report_view(request.POST.get("report_view"))
        if report_view != "activity":
            return redirect("hub:report")

        start_month = self._normalize_month_value(request.POST.get("start_month"))
        end_month = self._normalize_month_value(request.POST.get("end_month"))
        activity_report_url = self._build_activity_report_url(start_month=start_month, end_month=end_month)

        log_id = (request.POST.get("log_id") or "").strip()
        if not log_id:
            messages.error(request, "No activity log was selected for editing.")
            return redirect(activity_report_url)

        try:
            editing_activity_log = EngineerActivityLog.objects.select_related("engineer", "account", "request").get(pk=log_id)
        except (EngineerActivityLog.DoesNotExist, ValueError):
            messages.error(request, "Unable to update the selected activity log.")
            return redirect(activity_report_url)

        activity_form = EngineerActivityLogForm(
            data=request.POST,
            engineer=editing_activity_log.engineer,
            instance=editing_activity_log,
        )
        if activity_form.is_valid():
            activity_form.save()
            engineer_name = editing_activity_log.engineer.get_full_name() or editing_activity_log.engineer.username
            messages.success(request, f"Activity log for {engineer_name} updated successfully.")
            return redirect(activity_report_url)

        context = self.get_context_data(
            report_view="activity",
            editing_activity_log=editing_activity_log,
            activity_form=activity_form,
            activity_start_month=start_month,
            activity_end_month=end_month,
        )
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report_view = self._normalize_report_view(kwargs.get("report_view") or self.request.GET.get("report_view"))
        context["report_view"] = report_view

        if report_view == "activity":
            editing_activity_log = kwargs.get("editing_activity_log")
            activity_form = kwargs.get("activity_form")
            activity_start_month = kwargs.get("activity_start_month")
            activity_end_month = kwargs.get("activity_end_month")
            if editing_activity_log is None:
                editing_activity_log = self._get_requested_edit_log()
            if editing_activity_log is not None and activity_form is None:
                activity_form = EngineerActivityLogForm(
                    engineer=editing_activity_log.engineer,
                    instance=editing_activity_log,
                )
            try:
                activity_log_page = max(1, int(self.request.GET.get("activity_log_page") or 1))
            except (TypeError, ValueError):
                activity_log_page = 1
            context.update(
                self._build_activity_log_context(
                    start_month_value=activity_start_month,
                    end_month_value=activity_end_month,
                    page_num=activity_log_page,
                )
            )
            context["editing_activity_log"] = editing_activity_log
            context["activity_form"] = activity_form
        else:
            context.update(self._build_operational_context())
        return context

    def _build_operational_context(self):
        total_requests = Request.objects.count()
        account_manager_qs = (
            Request.objects.values("account_manager")
            .annotate(
                total=Count("id"),
                ongoing=Count("id", filter=Q(status=Request.Status.ONGOING)),
                completed=Count("id", filter=Q(status=Request.Status.COMPLETED)),
            )
            .order_by("-total", "account_manager")
        )
        account_manager_data = []
        for item in account_manager_qs:
            label = item["account_manager"] or "Unspecified"
            account_manager_data.append(
                {
                    "name": label,
                    "total": item["total"],
                    "ongoing": item["ongoing"],
                    "completed": item["completed"],
                }
            )

        engineer_qs = (
            Request.objects.values(
                "engineer",
                "engineer__first_name",
                "engineer__last_name",
                "engineer__username",
            )
            .annotate(
                total=Count("id"),
                ongoing=Count("id", filter=Q(status=Request.Status.ONGOING)),
                completed=Count("id", filter=Q(status=Request.Status.COMPLETED)),
            )
            .order_by("-total", "engineer__username")
        )
        engineer_data = []
        for item in engineer_qs:
            if item["engineer"] is None:
                name = "Unassigned"
            else:
                first = item.get("engineer__first_name") or ""
                last = item.get("engineer__last_name") or ""
                full_name = f"{first} {last}".strip()
                name = full_name or item.get("engineer__username") or "Engineer"
            engineer_data.append(
                {
                    "name": name,
                    "total": item["total"],
                    "ongoing": item["ongoing"],
                    "completed": item["completed"],
                }
            )

        status_labels = dict(Request.Status.choices)
        status_counts = {status: 0 for status, _ in Request.Status.choices}
        for item in Request.objects.values("status").annotate(total=Count("id")):
            status_counts[item["status"]] = item["total"]
        status_breakdown = [
            {
                "label": status_labels.get(status, status.title()),
                "total": total,
            }
            for status, total in status_counts.items()
            if total > 0
        ]

        engagement_counts = {
            value: {"total": 0, "ongoing": 0, "completed": 0}
            for value, _ in Request.Engagement.choices
        }
        for item in Request.objects.values("engagement_type", "status").annotate(total=Count("id")):
            key = item["engagement_type"]
            bucket = engagement_counts.setdefault(
                key,
                {"total": 0, "ongoing": 0, "completed": 0},
            )
            bucket["total"] += item["total"]
            if item["status"] == Request.Status.ONGOING:
                bucket["ongoing"] += item["total"]
            elif item["status"] == Request.Status.COMPLETED:
                bucket["completed"] += item["total"]

        engagement_labels_map = dict(Request.Engagement.choices)
        engagement_order = [
            Request.Engagement.OPPORTUNITY,
            Request.Engagement.SUPPORT,
            Request.Engagement.TRAINING,
            Request.Engagement.INQUIRY,
        ]

        product_categories = ["Azure", "M365", "Others"]
        product_buckets = {
            category: {"total": 0, "ongoing": 0, "completed": 0}
            for category in product_categories
        }
        for item in Request.objects.values("product_category", "status").annotate(total=Count("id")):
            category = item["product_category"] or "Others"
            normalized = (category or "").lower()
            if normalized == "azure" or category == "Azure":
                key = "Azure"
            elif normalized in {"m365", "microsoft 365"}:
                key = "M365"
            else:
                key = "Others"

            bucket = product_buckets.setdefault(key, {"total": 0, "ongoing": 0, "completed": 0})
            bucket["total"] += item["total"]
            if item["status"] == Request.Status.ONGOING:
                bucket["ongoing"] += item["total"]
            elif item["status"] == Request.Status.COMPLETED:
                bucket["completed"] += item["total"]

        engagement_chart_payload = {
            "labels": [engagement_labels_map.get(value, value.title()) for value in engagement_order],
            "totals": [engagement_counts.get(value, {"total": 0})["total"] for value in engagement_order],
            "ongoing": [engagement_counts.get(value, {"ongoing": 0})["ongoing"] for value in engagement_order],
            "completed": [engagement_counts.get(value, {"completed": 0})["completed"] for value in engagement_order],
        }
        product_chart_payload = {
            "labels": list(product_buckets.keys()),
            "totals": [bucket["total"] for bucket in product_buckets.values()],
            "ongoing": [bucket["ongoing"] for bucket in product_buckets.values()],
            "completed": [bucket["completed"] for bucket in product_buckets.values()],
        }
        engagement_chart_has_data = any(value > 0 for value in engagement_chart_payload["totals"])
        product_chart_has_data = any(value > 0 for value in product_chart_payload["totals"])

        return {
            "totals": {
                "requests": total_requests,
                "account_managers": (
                    Request.objects.exclude(account_manager__isnull=True)
                    .exclude(account_manager__exact="")
                    .values("account_manager")
                    .distinct()
                    .count()
                ),
                "engineers": (
                    Request.objects.exclude(engineer__isnull=True)
                    .values("engineer")
                    .distinct()
                    .count()
                ),
                "ongoing": Request.objects.filter(status=Request.Status.ONGOING).count(),
                "completed": Request.objects.filter(status=Request.Status.COMPLETED).count(),
            },
            "account_manager_chart": {
                "labels": [item["name"] for item in account_manager_data],
                "totals": [item["total"] for item in account_manager_data],
                "ongoing": [item["ongoing"] for item in account_manager_data],
                "completed": [item["completed"] for item in account_manager_data],
            },
            "engineer_chart": {
                "labels": [item["name"] for item in engineer_data],
                "totals": [item["total"] for item in engineer_data],
                "ongoing": [item["ongoing"] for item in engineer_data],
                "completed": [item["completed"] for item in engineer_data],
            },
            "engagement_chart": engagement_chart_payload,
            "engagement_chart_has_data": engagement_chart_has_data,
            "product_chart": product_chart_payload,
            "product_chart_has_data": product_chart_has_data,
            "account_manager_stats": account_manager_data,
            "engineer_stats": engineer_data,
            "status_breakdown": status_breakdown,
        }

    def _build_activity_log_context(self, *, start_month_value: str | None = None, end_month_value: str | None = None, page_num: int = 1):
        month_filters = self._resolve_activity_month_filters(
            start_month_value=start_month_value,
            end_month_value=end_month_value,
        )
        logs_qs = EngineerActivityLog.objects.select_related("engineer", "account", "request")
        if month_filters["start_month_date"]:
            logs_qs = logs_qs.filter(request_date__gte=month_filters["start_month_date"])
        if month_filters["end_exclusive_date"]:
            logs_qs = logs_qs.filter(request_date__lt=month_filters["end_exclusive_date"])

        total_hours = logs_qs.aggregate(total=Sum("actual_hours")) or {}
        billable_hours = logs_qs.filter(is_billable=True).aggregate(total=Sum("actual_hours")) or {}
        non_billable_hours = logs_qs.filter(is_billable=False).aggregate(total=Sum("actual_hours")) or {}

        total_hours_value = total_hours.get("total") or Decimal("0")
        billable_hours_value = billable_hours.get("total") or Decimal("0")
        non_billable_hours_value = non_billable_hours.get("total") or Decimal("0")

        engineer_hours = (
            logs_qs.values(
                "engineer",
                "engineer__first_name",
                "engineer__last_name",
                "engineer__username",
            )
            .annotate(
                total_hours=Sum("actual_hours"),
                billable_hours=Sum("actual_hours", filter=Q(is_billable=True)),
            )
            .order_by("engineer__first_name", "engineer__last_name", "engineer__username")
        )

        engineer_chart = {
            "labels": [],
            "billable": [],
            "non_billable": [],
        }
        engineer_table = []
        for row in engineer_hours:
            if row["engineer"] is None:
                display_name = "Unassigned"
            else:
                first = (row.get("engineer__first_name") or "").strip()
                last = (row.get("engineer__last_name") or "").strip()
                full_name = f"{first} {last}".strip()
                display_name = full_name or row.get("engineer__username") or "Engineer"
            total_val = row.get("total_hours") or Decimal("0")
            billable_val = row.get("billable_hours") or Decimal("0")
            non_billable_val = total_val - billable_val
            engineer_chart["labels"].append(display_name)
            engineer_chart["billable"].append(float(billable_val))
            engineer_chart["non_billable"].append(float(non_billable_val))
            engineer_table.append(
                {
                    "name": display_name,
                    "total_hours": total_val,
                    "billable_hours": billable_val,
                    "non_billable_hours": non_billable_val,
                }
            )

        activity_label_map = dict(EngineerActivityLog.ActivityType.choices)
        activity_hours = (
            logs_qs.values("activity_type")
            .annotate(
                total_hours=Sum("actual_hours"),
                billable_hours=Sum("actual_hours", filter=Q(is_billable=True)),
            )
            .order_by("activity_type")
        )

        activity_chart = {
            "labels": [],
            "billable": [],
            "non_billable": [],
        }
        activity_table = []
        for row in activity_hours:
            activity_key = row.get("activity_type")
            label = activity_label_map.get(activity_key, "Unspecified")
            total_val = row.get("total_hours") or Decimal("0")
            billable_val = row.get("billable_hours") or Decimal("0")
            non_billable_val = total_val - billable_val
            activity_chart["labels"].append(label)
            activity_chart["billable"].append(float(billable_val))
            activity_chart["non_billable"].append(float(non_billable_val))
            activity_table.append(
                {
                    "label": label,
                    "total_hours": total_val,
                    "billable_hours": billable_val,
                    "non_billable_hours": non_billable_val,
                }
            )

        location_label_map = dict(EngineerActivityLog.Location.choices)
        location_hours = (
            logs_qs.values("location")
            .annotate(total_hours=Sum("actual_hours"))
            .order_by("location")
        )
        location_chart = {
            "labels": [],
            "totals": [],
        }
        for row in location_hours:
            location_key = row.get("location")
            label = location_label_map.get(location_key, "Unspecified")
            total_val = row.get("total_hours") or Decimal("0")
            location_chart["labels"].append(label)
            location_chart["totals"].append(float(total_val))

        billable_chart = {
            "labels": ["Billable", "Not Billable"],
            "totals": [
                float(billable_hours_value),
                float(non_billable_hours_value),
            ],
        }

        paginator = Paginator(logs_qs.order_by("-request_date", "-created_at"), 50)
        try:
            page_obj = paginator.page(page_num)
        except InvalidPage:
            page_obj = paginator.page(1)
        recent_logs = list(page_obj.object_list)

        return {
            "activity_totals": {
                "entries": logs_qs.count(),
                "total_hours": total_hours_value,
                "billable_hours": billable_hours_value,
                "non_billable_hours": non_billable_hours_value,
                "unique_accounts": logs_qs.values("account").distinct().count(),
            },
            "activity_start_month": month_filters["start_month"],
            "activity_end_month": month_filters["end_month"],
            "activity_month_filter_label": month_filters["label"],
            "activity_engineer_chart": engineer_chart,
            "activity_engineer_table": engineer_table,
            "activity_type_chart": activity_chart,
            "activity_type_table": activity_table,
            "activity_location_chart": location_chart,
            "activity_billable_chart": billable_chart,
            "activity_logs": recent_logs,
            "activity_logs_page_obj": page_obj,
        }


class UserManagementView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = "hub/management.html"
    formset_class = modelformset_factory(User, form=UserManagementForm, extra=0, can_delete=False)
    account_form_class = modelformset_factory(Account, form=AccountManagementForm, extra=1, can_delete=True)

    def get_queryset(self):
        return User.objects.order_by("date_joined", "username")

    def get(self, request, *args, **kwargs):
        self._sync_account_baseline()
        formset = self.formset_class(queryset=self.get_queryset())
        account_formset = self.account_form_class(queryset=Account.objects.order_by("name"))
        create_user_form = UserManagementForm(prefix="create_user")
        self._prepare_formset(formset)
        self._prepare_account_formset(account_formset)
        return render(request, self.template_name, self._build_context(formset, account_formset, create_user_form=create_user_form))

    def post(self, request, *args, **kwargs):
        active_tab = request.POST.get("active_tab", "users")
        self._sync_account_baseline()
        if active_tab != "accounts":
            user_action_value = request.POST.get("user_action")
            if user_action_value:
                action_response = self._handle_user_action_request(request, user_action_value)
                if action_response:
                    return action_response

            if request.POST.get("create_user_submit") == "1":
                create_user_form = UserManagementForm(request.POST, prefix="create_user")
                formset = self.formset_class(queryset=self.get_queryset())
                account_formset = self.account_form_class(queryset=Account.objects.order_by("name"))
                self._prepare_formset(formset)
                self._prepare_account_formset(account_formset)
                if create_user_form.is_valid():
                    new_user = create_user_form.save()
                    display_name = new_user.get_full_name() or new_user.username
                    messages.success(request, f"Created new user account for {display_name}.")
                    return redirect("hub:management")
                return render(
                    request,
                    self.template_name,
                    self._build_context(
                        formset,
                        account_formset,
                        create_user_form=create_user_form,
                        show_create_user_modal=True,
                    ),
                )

        formset = self.formset_class(request.POST, queryset=self.get_queryset())
        account_formset = self.account_form_class(request.POST, queryset=Account.objects.order_by("name"))
        create_user_form = UserManagementForm(prefix="create_user")
        self._prepare_formset(formset)
        self._prepare_account_formset(account_formset)

        if active_tab == "accounts":
            return self._handle_account_submission(request, account_formset, formset)

        if formset.is_valid():
            current_admins = User.objects.filter(role=User.Roles.ADMIN).count()
            admin_delta = 0
            pending_error = False
            admin_removal_candidates = []

            for form in formset:
                if not form.cleaned_data:
                    continue

                is_delete = form.cleaned_data.get("DELETE")
                if is_delete:
                    if form.instance.pk and form.instance.is_superuser:
                        form.add_error("DELETE", "Superuser accounts cannot be deleted.")
                        pending_error = True
                        continue
                    if form.instance.pk and form.instance.role == User.Roles.ADMIN:
                        admin_delta -= 1
                        admin_removal_candidates.append((form, "delete"))
                    continue

                if not form.instance.pk and not form.has_changed():
                    continue

                new_role = form.cleaned_data.get("role")
                original_role = form.instance.role if form.instance.pk else None

                if form.instance.pk and form.instance.is_superuser and new_role != User.Roles.ADMIN:
                    form.add_error("role", "Superuser accounts must remain administrators.")
                    pending_error = True

                if form.instance.pk:
                    if original_role == User.Roles.ADMIN and new_role != User.Roles.ADMIN:
                        admin_delta -= 1
                        admin_removal_candidates.append((form, "role"))
                    elif original_role != User.Roles.ADMIN and new_role == User.Roles.ADMIN:
                        admin_delta += 1
                else:
                    if new_role == User.Roles.ADMIN:
                        admin_delta += 1

            if pending_error:
                return render(request, self.template_name, self._build_context(formset, account_formset, create_user_form=create_user_form))

            if current_admins + admin_delta <= 0:
                if admin_removal_candidates:
                    form, field = admin_removal_candidates[0]
                    if field == "delete":
                        form.add_error(None, "At least one administrator must remain.")
                    else:
                        form.add_error("role", "At least one administrator must remain.")
                else:
                    messages.error(request, "At least one administrator must remain.")
                return render(request, self.template_name, self._build_context(formset, account_formset, create_user_form=create_user_form))

            created_count = 0
            updated_count = 0
            deleted_count = 0

            with transaction.atomic():
                for form in formset:
                    if not form.cleaned_data:
                        continue

                    if not form.has_changed() and form.instance.pk:
                        continue

                    is_new = not form.instance.pk
                    form.save()
                    if is_new:
                        created_count += 1
                    else:
                        updated_count += 1

            if created_count or updated_count or deleted_count:
                parts = []
                if created_count:
                    parts.append(f"created {created_count} user account{'s' if created_count != 1 else ''}")
                if updated_count:
                    parts.append(f"updated {updated_count} user account{'s' if updated_count != 1 else ''}")
                if deleted_count:
                    parts.append(f"removed {deleted_count} user account{'s' if deleted_count != 1 else ''}")
                messages.success(request, ", ".join(parts).capitalize() + ".")
            else:
                messages.info(request, "No changes detected.")
            return redirect("hub:management")

        return render(request, self.template_name, self._build_context(formset, account_formset, create_user_form=create_user_form))

    def _handle_account_submission(self, request, account_formset, user_formset):
        if account_formset.is_valid():
            created = updated = deleted = 0
            with transaction.atomic():
                for form in account_formset:
                    if not form.cleaned_data:
                        continue
                    if form.cleaned_data.get("DELETE"):
                        if form.instance.pk:
                            form.instance.delete()
                            deleted += 1
                        continue
                    if not form.instance.pk and not form.has_changed():
                        continue
                    is_new = not form.instance.pk
                    form.save()
                    if is_new:
                        created += 1
                    else:
                        updated += 1
            if created or updated or deleted:
                parts = []
                if created:
                    parts.append(f"created {created} account{'s' if created != 1 else ''}")
                if updated:
                    parts.append(f"updated {updated} account{'s' if updated != 1 else ''}")
                if deleted:
                    parts.append(f"removed {deleted} account{'s' if deleted != 1 else ''}")
                messages.success(request, ", ".join(parts).capitalize() + ".")
            else:
                messages.info(request, "No account changes detected.")
            return redirect("hub:management")
        return render(request, self.template_name, self._build_context(user_formset, account_formset, active_tab="accounts"))

    def _build_context(self, formset, account_formset, active_tab="users", create_user_form=None, show_create_user_modal=False):
        if create_user_form is None:
            create_user_form = UserManagementForm(prefix="create_user")
        return {
            "formset": formset,
            "account_formset": account_formset,
            "create_user_form": create_user_form,
            "show_create_user_modal": show_create_user_modal,
            "total_users": User.objects.count(),
            "total_accounts": Account.objects.count(),
            "active_tab": active_tab,
            "default_password": getattr(settings, "DEFAULT_USER_PASSWORD", "@Password"),
            "users": User.objects.select_related().order_by("date_joined", "username"),
            "role_choices": User.Roles.choices,
        }

    @staticmethod
    def _prepare_formset(formset):
        for form in formset:
            delete_field = form.fields.get("DELETE")
            if delete_field:
                existing_class = delete_field.widget.attrs.get("class", "")
                delete_field.widget.attrs["class"] = (existing_class + " form-check-input").strip()

    @staticmethod
    def _prepare_account_formset(formset):
        for form in formset:
            delete_field = form.fields.get("DELETE")
            if delete_field:
                existing_class = delete_field.widget.attrs.get("class", "")
                delete_field.widget.attrs["class"] = (existing_class + " form-check-input").strip()

    @staticmethod
    def _sync_account_baseline():
        if Account.objects.exists():
            return

        seed_accounts = []
        for raw_name in ACCOUNT_NAME_RAW:
            normalized = (raw_name or "").strip()
            if not normalized:
                continue
            seed_accounts.append(Account(name=normalized))

        if seed_accounts:
            Account.objects.bulk_create(seed_accounts, ignore_conflicts=True)

    def _handle_user_action_request(self, request, action_value):
        action, separator, raw_user_id = (action_value or "").partition(":")
        if not separator:
            messages.error(request, "We could not determine the requested action.")
            return redirect("hub:management")

        try:
            target_user = User.objects.get(pk=int(raw_user_id))
        except (User.DoesNotExist, ValueError):
            messages.error(request, "We could not find the selected user account.")
            return redirect("hub:management")

        if action == "reset_password":
            return self._reset_user_password(request, target_user)
        if action == "delete_user":
            return self._delete_user_account(request, target_user)

        messages.error(request, "Unknown action requested.")
        return redirect("hub:management")

    @staticmethod
    def _clear_user_sessions(target_user: User) -> int:
        removed = 0
        user_id = str(target_user.pk)
        session_qs = Session.objects.filter(expire_date__gte=timezone.now())
        for session in session_qs:
            try:
                data = session.get_decoded()
            except Exception:
                continue
            if str(data.get("_auth_user_id")) == user_id:
                session.delete()
                removed += 1
        return removed

    def _reset_user_password(self, request, target_user: User):
        default_password = getattr(settings, "DEFAULT_USER_PASSWORD", "@Password")
        target_user.set_password(default_password)
        target_user.must_change_password = True
        target_user.save(update_fields=["password", "must_change_password"])
        removed_sessions = self._clear_user_sessions(target_user)
        display_name = target_user.get_full_name() or target_user.username
        message = (
            f"Reset password for {display_name}. The default password has been restored and they must set a new password on next sign-in."
        )
        if removed_sessions:
            message += f" Ended {removed_sessions} session{'s' if removed_sessions != 1 else ''}."
        messages.success(request, message)
        return redirect("hub:management")

    def _delete_user_account(self, request, target_user: User):
        if target_user.is_superuser:
            messages.error(request, "Superuser accounts cannot be deleted.")
            return redirect("hub:management")

        if target_user.pk == request.user.pk:
            messages.error(request, "You cannot delete your own account.")
            return redirect("hub:management")

        if target_user.role == User.Roles.ADMIN:
            admin_count = User.objects.filter(role=User.Roles.ADMIN).count()
            if admin_count <= 1:
                messages.error(request, "At least one administrator must remain.")
                return redirect("hub:management")

        display_name = target_user.get_full_name() or target_user.username
        target_user.delete()
        messages.success(request, f"Deleted user account for {display_name}.")
        return redirect("hub:management")


class UserEditView(AdminRequiredMixin, LoginRequiredMixin, View):
    """Handle the per-user Manage modal: update info and optionally change password."""

    def post(self, request, pk):
        target_user = get_object_or_404(User, pk=pk)
        form = UserManagementForm(request.POST, instance=target_user)
        if form.is_valid():
            form.save()
            display_name = target_user.get_full_name() or target_user.username
            messages.success(request, f"Updated user account for {display_name}.")
        else:
            for field_name, field_errors in form.errors.items():
                label = form.fields[field_name].label if field_name in form.fields else field_name
                for error in field_errors:
                    messages.error(request, f"{label}: {error}")
        return redirect("hub:management")


class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = "hub/notifications.html"
    context_object_name = "notifications"
    paginate_by = 50

    def get_queryset(self):
        queryset = (
            self.request.user.notifications.select_related("related_request")
            .order_by("-created_at")
        )
        user = self.request.user
        if getattr(user, "role", None) in ADMIN_PANEL_ROLES:
            queryset = queryset.filter(source__icontains="new request")

        # ── Filter: read/unread ──
        self._filter_read = self.request.GET.get("filter", "all")
        if self._filter_read == "unread":
            queryset = queryset.filter(is_read=False)
        elif self._filter_read == "read":
            queryset = queryset.filter(is_read=True)

        # ── Filter: category ──
        self._filter_category = self.request.GET.get("category", "")
        if self._filter_category:
            # We can't directly filter on computed properties, so filter on message/source patterns
            cat = self._filter_category
            if cat == "new_request":
                queryset = queryset.filter(source__icontains="new request")
            elif cat == "assignment":
                queryset = queryset.filter(message__icontains="assigned to request")
            elif cat == "update":
                queryset = queryset.filter(
                    Q(source__icontains="status update") |
                    Q(message__icontains="posted an update") |
                    Q(message__icontains=" updated ")
                )
            elif cat == "completion":
                queryset = queryset.filter(
                    Q(message__icontains="completed") |
                    Q(message__icontains="closed") |
                    Q(source__icontains="close")
                )
            elif cat == "reminder":
                queryset = queryset.filter(
                    Q(source__icontains="nudge") |
                    Q(message__icontains="reminder")
                )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        base_qs = user.notifications.all()
        if getattr(user, "role", None) in ADMIN_PANEL_ROLES:
            base_qs = base_qs.filter(source__icontains="new request")
        context["unread_count"] = base_qs.filter(is_read=False).count()
        context["total_count"] = base_qs.count()
        context["current_filter"] = self._filter_read
        context["current_category"] = self._filter_category

        # ── Category breakdown for the summary bar ──
        category_counts = {
            "new_request": base_qs.filter(source__icontains="new request").count(),
            "assignment": base_qs.filter(message__icontains="assigned to request").count(),
            "update": base_qs.filter(
                Q(source__icontains="status update") |
                Q(message__icontains="posted an update") |
                Q(message__icontains=" updated ")
            ).count(),
            "completion": base_qs.filter(
                Q(message__icontains="completed") |
                Q(message__icontains="closed") |
                Q(source__icontains="close")
            ).count(),
            "reminder": base_qs.filter(
                Q(source__icontains="nudge") |
                Q(message__icontains="reminder")
            ).count(),
        }
        context["category_counts"] = category_counts

        # ── Date-grouped notifications for timeline display ──
        notifications_list = list(context["notifications"])
        context["grouped_notifications"] = self._group_by_date(notifications_list)

        return context

    def _group_by_date(self, notifications):
        """Group notifications by Today, Yesterday, This Week, Older."""
        now = timezone.now().date()
        groups = {"Today": [], "Yesterday": [], "This Week": [], "Older": []}

        for n in notifications:
            created_date = timezone.localtime(n.created_at, MANILA_TZ).date()
            if created_date == now:
                groups["Today"].append(n)
            elif created_date == now - timedelta(days=1):
                groups["Yesterday"].append(n)
            elif (now - created_date).days < 7:
                groups["This Week"].append(n)
            else:
                groups["Older"].append(n)

        return {k: v for k, v in groups.items() if v}


class NotificationReadView(LoginRequiredMixin, View):
    def post(self, request, pk):
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.mark_read()
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", reverse("hub:notifications")))


class NotificationFollowRedirectView(LoginRequiredMixin, View):
    def get(self, request, pk):
        notification = get_object_or_404(
            Notification.objects.select_related("related_request"),
            pk=pk,
            recipient=request.user,
        )
        notification.mark_read()
        if notification.related_request:
            req = notification.related_request
            # Route to manage views instead of read-only detail.
            if request.user.role in ADMIN_PANEL_ROLES:
                return redirect("hub:request-manage", pk=req.pk)
            return redirect("hub:request-manage-collab", pk=req.pk)
        return redirect("hub:notifications")


class NotificationDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.delete()
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", reverse("hub:notifications")))


class NotificationMarkAllReadView(LoginRequiredMixin, View):
    def post(self, request):
        request.user.notifications.filter(is_read=False).update(is_read=True)
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", reverse("hub:notifications")))


class RequestNudgeView(AdminOrPmEsgRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(Request.objects.select_related("engineer", "backup_engineer"), pk=pk)

        if not request_obj.engineer:
            messages.error(request, "Cannot nudge: No engineer assigned.")
            return redirect("hub:request-manage", pk=pk)

        Notification.objects.create(
            recipient=request_obj.engineer,
            message=f"Reminder: Please update the status for {request_obj.reference_code}.",
            related_request=request_obj,
            actor="Admin",
            source="Request Nudge",
        )

        if request_obj.backup_engineer:
            Notification.objects.create(
                recipient=request_obj.backup_engineer,
                message=f"Reminder: Backup request {request_obj.reference_code} needs attention.",
                related_request=request_obj,
                actor="Admin",
                source="Request Nudge",
            )

        messages.success(
            request,
            f"Nudge sent to {request_obj.engineer.get_full_name() or request_obj.engineer.username}.",
        )
        return redirect("hub:request-manage", pk=pk)


class RequestStatusUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        request_obj = get_object_or_404(
            Request.objects.select_related("engineer", "requestor"),
            pk=pk,
        )

        if request.user.role not in ENGINEER_ACCESS_ROLES or request_obj.engineer_id != request.user.id:
            messages.error(request, "You are not allowed to update this request's status.")
            return redirect("hub:request-manage-collab", pk=pk)

        original = Request.objects.get(pk=request_obj.pk)
        form = RequestStatusForm(request.POST, instance=request_obj)
        if form.is_valid():
            request_obj._actor_user = request.user
            request_obj._actor_source = "Engineer · Status Update"
            form.save()
            changed_fields = []
            if original.status != request_obj.status:
                changed_fields.append("status")
            if original.end_date != request_obj.end_date:
                changed_fields.append("end_date")
            if changed_fields:
                summary_text = summarize_request_changes(original, request_obj, changed_fields)
                create_change_status_log(request_obj, request.user, "Engineer · Status Update", summary_text)
            messages.success(request, "Request status updated.")
        else:
            messages.error(request, "Unable to update status. Please try again.")
        return redirect("hub:request-manage-collab", pk=pk)
